from __future__ import annotations

import sys
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Font


ROOT = Path(__file__).resolve().parents[1]
ENCODING = "utf-8-sig"

ASSET_MASTER_PRIMARY = ROOT / "data" / "stage5_mapping" / "final" / "asset_master_for_mapping_patched.csv"
ASSET_MASTER_FALLBACK = ROOT / "data" / "stage4_etf_index" / "final" / "asset_master_with_etf_index_cleaned.csv"
UNDERLYING_MAPPING = ROOT / "data" / "stage5_mapping" / "final" / "underlying_mapping_fixed.csv"
DOMESTIC_MAPPING = ROOT / "data" / "stage5_mapping" / "final" / "domestic_cross_market_mapping_fixed.csv"

STAGE6_DIR = ROOT / "data" / "stage6_industry_chain"
FINAL_DIR = STAGE6_DIR / "final"
PROCESSED_DIR = STAGE6_DIR / "processed"
MANUAL_DIR = STAGE6_DIR / "manual"
OUTPUT_DIR = ROOT / "output" / "stage6_industry_chain"

CHAIN_MAPPING_OUT = FINAL_DIR / "industry_chain_mapping.csv"
ASSET_POOL_OUT = FINAL_DIR / "industry_chain_asset_pool.csv"
QUALITY_REPORT_OUT = PROCESSED_DIR / "industry_chain_quality_report.csv"
MISSING_REVIEW_OUT = PROCESSED_DIR / "industry_chain_missing_asset_review.csv"
EXCEL_OUT = ROOT / "output" / "mapping_info_stage6_industry_chain.xlsx"

MISSING_VALUES = {"", "CHECK", "TODO", "NA", "NAN", "NONE", "CHECK_WAIT_ETF_MASTER"}


CHAIN_RULES: dict[str, dict[str, Any]] = {
    "BLACK": {
        "industry_name": "黑色产业链",
        "core_groups": [
            "IRON_ORE", "COKING_COAL", "COKE", "REBAR", "HOT_ROLLED_COIL",
            "STAINLESS_STEEL", "FERROSILICON", "SILICOMANGANESE",
        ],
        "relations": [
            ("IRON_ORE", "REBAR", "A_TO_B", "UPSTREAM_DOWNSTREAM", "HIGH"),
            ("IRON_ORE", "HOT_ROLLED_COIL", "A_TO_B", "UPSTREAM_DOWNSTREAM", "HIGH"),
            ("COKING_COAL", "COKE", "A_TO_B", "UPSTREAM_DOWNSTREAM", "HIGH"),
            ("COKE", "REBAR", "A_TO_B", "UPSTREAM_DOWNSTREAM", "HIGH"),
            ("COKE", "HOT_ROLLED_COIL", "A_TO_B", "UPSTREAM_DOWNSTREAM", "HIGH"),
            ("FERROSILICON", "REBAR", "A_TO_B", "COST_DRIVEN", "MEDIUM"),
            ("SILICOMANGANESE", "REBAR", "A_TO_B", "COST_DRIVEN", "MEDIUM"),
            ("REBAR", "HOT_ROLLED_COIL", "BIDIRECTIONAL", "SUBSTITUTE", "HIGH"),
        ],
    },
    "NONFERROUS": {
        "industry_name": "有色产业链",
        "core_groups": ["COPPER", "ALUMINUM", "ZINC", "LEAD", "NICKEL", "TIN", "ALUMINA", "CAST_ALUMINUM_ALLOY"],
        "relations": [
            ("ALUMINA", "ALUMINUM", "A_TO_B", "UPSTREAM_DOWNSTREAM", "HIGH"),
            ("ALUMINUM", "CAST_ALUMINUM_ALLOY", "A_TO_B", "UPSTREAM_DOWNSTREAM", "HIGH"),
            ("COPPER", "ALUMINUM", "BIDIRECTIONAL", "SAME_SECTOR_RELATED", "MEDIUM"),
            ("ZINC", "LEAD", "BIDIRECTIONAL", "SAME_SECTOR_RELATED", "MEDIUM"),
            ("NICKEL", "STAINLESS_STEEL", "BIDIRECTIONAL", "COST_DRIVEN", "HIGH"),
            ("TIN", "COPPER", "BIDIRECTIONAL", "SAME_SECTOR_RELATED", "LOW"),
        ],
    },
    "PRECIOUS_METAL": {
        "industry_name": "贵金属",
        "core_groups": ["GOLD", "SILVER", "PALLADIUM", "PLATINUM"],
        "relations": [
            ("GOLD", "SILVER", "BIDIRECTIONAL", "SUBSTITUTE", "HIGH"),
            ("PALLADIUM", "PLATINUM", "BIDIRECTIONAL", "SUBSTITUTE", "HIGH"),
            ("GOLD", "PLATINUM", "BIDIRECTIONAL", "SAME_SECTOR_RELATED", "MEDIUM"),
        ],
    },
    "ENERGY_CHEMICAL": {
        "industry_name": "能源化工",
        "core_groups": [
            "CRUDE_OIL", "FUEL_OIL", "LOW_SULFUR_FUEL_OIL", "BITUMEN", "RUBBER",
            "BUTADIENE_RUBBER", "PTA", "METHANOL", "LLDPE", "PVC", "POLYPROPYLENE",
            "ETHYLENE_GLYCOL", "STYRENE", "POLYESTER_STAPLE_FIBER", "PARAXYLENE",
            "BOTTLE_GRADE_PET", "PROPYLENE", "BENZENE", "UREA", "SODA_ASH", "GLASS",
            "CAUSTIC_SODA", "PULP", "OFFSET_PAPER",
        ],
        "relations": [
            ("CRUDE_OIL", "FUEL_OIL", "A_TO_B", "UPSTREAM_DOWNSTREAM", "HIGH"),
            ("CRUDE_OIL", "LOW_SULFUR_FUEL_OIL", "A_TO_B", "UPSTREAM_DOWNSTREAM", "HIGH"),
            ("CRUDE_OIL", "BITUMEN", "A_TO_B", "UPSTREAM_DOWNSTREAM", "HIGH"),
            ("CRUDE_OIL", "PTA", "A_TO_B", "COST_DRIVEN", "MEDIUM"),
            ("CRUDE_OIL", "LLDPE", "A_TO_B", "COST_DRIVEN", "MEDIUM"),
            ("CRUDE_OIL", "POLYPROPYLENE", "A_TO_B", "COST_DRIVEN", "MEDIUM"),
            ("CRUDE_OIL", "PVC", "A_TO_B", "COST_DRIVEN", "LOW"),
            ("PARAXYLENE", "PTA", "A_TO_B", "UPSTREAM_DOWNSTREAM", "HIGH"),
            ("PTA", "POLYESTER_STAPLE_FIBER", "A_TO_B", "UPSTREAM_DOWNSTREAM", "HIGH"),
            ("PTA", "BOTTLE_GRADE_PET", "A_TO_B", "UPSTREAM_DOWNSTREAM", "HIGH"),
            ("ETHYLENE_GLYCOL", "POLYESTER_STAPLE_FIBER", "A_TO_B", "UPSTREAM_DOWNSTREAM", "HIGH"),
            ("STYRENE", "BENZENE", "A_TO_B", "COST_DRIVEN", "MEDIUM"),
            ("METHANOL", "POLYPROPYLENE", "A_TO_B", "COST_DRIVEN", "MEDIUM"),
            ("SODA_ASH", "GLASS", "A_TO_B", "COST_DRIVEN", "HIGH"),
            ("CAUSTIC_SODA", "PVC", "BIDIRECTIONAL", "SAME_SECTOR_RELATED", "MEDIUM"),
            ("RUBBER", "BUTADIENE_RUBBER", "BIDIRECTIONAL", "SUBSTITUTE", "MEDIUM"),
            ("PULP", "OFFSET_PAPER", "A_TO_B", "UPSTREAM_DOWNSTREAM", "HIGH"),
        ],
    },
    "OILSEED": {
        "industry_name": "油脂油料",
        "core_groups": [
            "SOYBEAN", "SOYBEAN_MEAL", "SOYBEAN_OIL", "PALM_OIL", "RAPESEED",
            "RAPESEED_MEAL", "RAPESEED_OIL", "PEANUT",
        ],
        "relations": [
            ("SOYBEAN", "SOYBEAN_MEAL", "A_TO_B", "UPSTREAM_DOWNSTREAM", "HIGH"),
            ("SOYBEAN", "SOYBEAN_OIL", "A_TO_B", "UPSTREAM_DOWNSTREAM", "HIGH"),
            ("RAPESEED", "RAPESEED_MEAL", "A_TO_B", "UPSTREAM_DOWNSTREAM", "HIGH"),
            ("RAPESEED", "RAPESEED_OIL", "A_TO_B", "UPSTREAM_DOWNSTREAM", "HIGH"),
            ("SOYBEAN_OIL", "PALM_OIL", "BIDIRECTIONAL", "SUBSTITUTE", "HIGH"),
            ("SOYBEAN_MEAL", "RAPESEED_MEAL", "BIDIRECTIONAL", "SUBSTITUTE", "HIGH"),
            ("SOYBEAN_OIL", "RAPESEED_OIL", "BIDIRECTIONAL", "SUBSTITUTE", "HIGH"),
            ("PEANUT", "SOYBEAN_OIL", "BIDIRECTIONAL", "SUBSTITUTE", "MEDIUM"),
        ],
    },
    "AGRICULTURE": {
        "industry_name": "农产品",
        "core_groups": [
            "CORN", "CORN_STARCH", "EGG", "LIVE_HOG", "SUGAR", "COTTON", "COTTON_YARN",
            "APPLE", "JUJUBE", "PEANUT", "JAPONICA_RICE", "EARLY_INDICA_RICE",
            "LATE_INDICA_RICE", "COMMON_WHEAT", "STRONG_WHEAT", "HARD_WHEAT",
        ],
        "relations": [
            ("CORN", "CORN_STARCH", "A_TO_B", "UPSTREAM_DOWNSTREAM", "HIGH"),
            ("CORN", "LIVE_HOG", "A_TO_B", "COST_DRIVEN", "HIGH"),
            ("CORN", "EGG", "A_TO_B", "COST_DRIVEN", "MEDIUM"),
            ("COTTON", "COTTON_YARN", "A_TO_B", "UPSTREAM_DOWNSTREAM", "HIGH"),
            ("SUGAR", "APPLE", "BIDIRECTIONAL", "DEMAND_LINKED", "LOW"),
            ("JAPONICA_RICE", "EARLY_INDICA_RICE", "BIDIRECTIONAL", "SUBSTITUTE", "MEDIUM"),
            ("EARLY_INDICA_RICE", "LATE_INDICA_RICE", "BIDIRECTIONAL", "SUBSTITUTE", "MEDIUM"),
            ("COMMON_WHEAT", "STRONG_WHEAT", "BIDIRECTIONAL", "SUBSTITUTE", "MEDIUM"),
            ("STRONG_WHEAT", "HARD_WHEAT", "BIDIRECTIONAL", "SUBSTITUTE", "MEDIUM"),
        ],
    },
    "NEW_ENERGY": {
        "industry_name": "新能源",
        "core_groups": ["INDUSTRIAL_SILICON", "LITHIUM_CARBONATE", "POLYSILICON"],
        "relations": [
            ("INDUSTRIAL_SILICON", "POLYSILICON", "A_TO_B", "UPSTREAM_DOWNSTREAM", "HIGH"),
            ("LITHIUM_CARBONATE", "INDUSTRIAL_SILICON", "BIDIRECTIONAL", "SAME_SECTOR_RELATED", "LOW"),
        ],
    },
    "BUILDING_MATERIAL": {
        "industry_name": "建材",
        "core_groups": ["PLYWOOD", "FIBERBOARD", "LOG", "GLASS", "REBAR", "HOT_ROLLED_COIL"],
        "relations": [
            ("LOG", "PLYWOOD", "A_TO_B", "UPSTREAM_DOWNSTREAM", "HIGH"),
            ("LOG", "FIBERBOARD", "A_TO_B", "UPSTREAM_DOWNSTREAM", "HIGH"),
            ("GLASS", "REBAR", "BIDIRECTIONAL", "DEMAND_LINKED", "LOW"),
            ("GLASS", "HOT_ROLLED_COIL", "BIDIRECTIONAL", "DEMAND_LINKED", "LOW"),
        ],
    },
}

CHAIN_PREFIX = {
    "BLACK": "CHAIN_BLACK",
    "NONFERROUS": "CHAIN_NONFERROUS",
    "PRECIOUS_METAL": "CHAIN_PRECIOUS",
    "ENERGY_CHEMICAL": "CHAIN_ENERGY_CHEM",
    "OILSEED": "CHAIN_OILSEED",
    "AGRICULTURE": "CHAIN_AGRI",
    "NEW_ENERGY": "CHAIN_NEW_ENERGY",
    "BUILDING_MATERIAL": "CHAIN_BUILDING",
}


def ensure_dirs() -> None:
    for path in [FINAL_DIR, PROCESSED_DIR, MANUAL_DIR, OUTPUT_DIR, EXCEL_OUT.parent]:
        path.mkdir(parents=True, exist_ok=True)


def clean_text(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    return str(value).strip()


def is_missing(value: Any) -> bool:
    return clean_text(value).upper() in MISSING_VALUES


def read_csv(path: Path, required: bool = True) -> pd.DataFrame:
    if not path.exists():
        if required:
            raise FileNotFoundError(f"缺少输入文件: {path}")
        return pd.DataFrame()
    return pd.read_csv(path, dtype=str, encoding=ENCODING).fillna("")


def save_csv(df: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(path, index=False, encoding=ENCODING)


def choose_asset_master() -> tuple[pd.DataFrame, str]:
    if ASSET_MASTER_PRIMARY.exists():
        return read_csv(ASSET_MASTER_PRIMARY), str(ASSET_MASTER_PRIMARY)
    return read_csv(ASSET_MASTER_FALLBACK), str(ASSET_MASTER_FALLBACK)


def all_rule_groups() -> set[str]:
    groups: set[str] = set()
    for rule in CHAIN_RULES.values():
        groups.update(rule["core_groups"])
        for a, b, *_ in rule["relations"]:
            groups.add(a)
            groups.add(b)
    return groups


def normalize_asset_type(asset_type: str) -> str:
    return "SPOT" if asset_type == "INDEX" else asset_type


def asset_rank(row: pd.Series) -> tuple[int, int, str]:
    asset_type = clean_text(row.get("asset_type"))
    exchange_code = clean_text(row.get("exchange_code"))
    asset_id = clean_text(row.get("asset_id"))
    type_rank = {"FUTURE": 0, "OPTION": 1, "ETF": 2, "SPOT": 3, "INDEX": 3}.get(asset_type, 9)
    domestic_rank = 0 if exchange_code in {"CFFEX", "SHFE", "INE", "DCE", "CZCE", "GFEX", "SSE", "SZSE"} else 1
    return type_rank, domestic_rank, asset_id


def build_asset_pool(asset_master: pd.DataFrame, underlying_mapping: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, pd.Series]]:
    base = underlying_mapping[
        [
            "asset_id", "symbol", "name_cn", "asset_type", "subtype", "exchange_code",
            "underlying_group", "underlying_name_cn", "underlying_name_en",
        ]
    ].copy()
    enrich_cols = ["asset_id", "tradable", "can_long", "can_short", "exchange_name", "country", "currency"]
    enrich = asset_master[[c for c in enrich_cols if c in asset_master.columns]].drop_duplicates("asset_id")
    base = base.merge(enrich, on="asset_id", how="left")
    base = base[~base["underlying_group"].map(is_missing)].copy()

    industry_by_group: dict[str, list[str]] = {}
    for rule in CHAIN_RULES.values():
        for group in rule["core_groups"]:
            industry_by_group.setdefault(group, []).append(rule["industry_name"])

    preferred: dict[str, pd.Series] = {}
    rows = []
    for group in sorted(all_rule_groups()):
        group_assets = base[base["underlying_group"] == group].copy()
        industry_name = " / ".join(industry_by_group.get(group, []))
        if group_assets.empty:
            rows.append(
                {
                    "underlying_group": group,
                    "underlying_name_cn": "",
                    "industry_name": industry_name,
                    "asset_count": 0,
                    "preferred_asset_id": "",
                    "preferred_symbol": "",
                    "preferred_name_cn": "",
                    "preferred_asset_type": "",
                    "all_asset_ids": "",
                    "mapping_status": "MISSING_ASSET",
                }
            )
            continue
        group_assets["_rank"] = group_assets.apply(asset_rank, axis=1)
        group_assets = group_assets.sort_values("_rank")
        best = group_assets.iloc[0].drop(labels=["_rank"], errors="ignore")
        preferred[group] = best
        rows.append(
            {
                "underlying_group": group,
                "underlying_name_cn": clean_text(best.get("underlying_name_cn")),
                "industry_name": industry_name,
                "asset_count": len(group_assets),
                "preferred_asset_id": clean_text(best.get("asset_id")),
                "preferred_symbol": clean_text(best.get("symbol")),
                "preferred_name_cn": clean_text(best.get("name_cn")),
                "preferred_asset_type": normalize_asset_type(clean_text(best.get("asset_type"))),
                "all_asset_ids": ";".join(group_assets["asset_id"].astype(str).tolist()),
                "mapping_status": "MAPPED",
            }
        )
    pool = pd.DataFrame(rows)
    save_csv(pool, ASSET_POOL_OUT)
    return pool, preferred


def segment_for(direction: str, relation_type: str, side: str) -> str:
    if relation_type == "UPSTREAM_DOWNSTREAM":
        return "UPSTREAM" if side == "A" else "DOWNSTREAM"
    if relation_type == "SUBSTITUTE":
        return "SUBSTITUTE"
    if relation_type == "BY_PRODUCT":
        return "BY_PRODUCT"
    if relation_type in {"COST_DRIVEN", "DEMAND_LINKED", "INVENTORY_LINKED"}:
        return "RELATED" if direction == "BIDIRECTIONAL" else ("UPSTREAM" if side == "A" else "DOWNSTREAM")
    return "RELATED"


def strategy_for(relation_type: str, strength: str) -> str:
    if relation_type in {"UPSTREAM_DOWNSTREAM", "COST_DRIVEN"}:
        return "CROSS_PRODUCT_ARBITRAGE" if strength in {"HIGH", "MEDIUM"} else "CHAIN_RELATED"
    if relation_type == "SUBSTITUTE":
        return "PAIR_TRADING"
    if relation_type in {"DEMAND_LINKED", "INVENTORY_LINKED", "SAME_SECTOR_RELATED"}:
        return "SPREAD_RELATED"
    return "CHAIN_RELATED"


def direction_supported(direction: str) -> str:
    if direction == "BIDIRECTIONAL":
        return "BIDIRECTIONAL"
    return "A_TO_B_CHAIN_DIRECTION"


def tradable_check(a: pd.Series, b: pd.Series) -> str:
    return "BOTH_TRADABLE" if clean_text(a.get("tradable")) == "Y" and clean_text(b.get("tradable")) == "Y" else "CHECK"


def long_short_note(direction: str, group_a: str, group_b: str) -> str:
    if direction == "BIDIRECTIONAL":
        return f"{group_a} 与 {group_b} 为双向相关品种，套利方向需结合价差、库存和期限结构确认。"
    return f"{group_a} -> {group_b} 为产业链方向，价差交易需结合成本传导和库存周期确认。"


def build_chain_mapping(preferred: dict[str, pd.Series]) -> tuple[pd.DataFrame, pd.DataFrame]:
    rows = []
    missing_rows = []
    seen: set[tuple[str, str, str]] = set()
    counters = {key: 0 for key in CHAIN_RULES}

    for chain_key, rule in CHAIN_RULES.items():
        industry = rule["industry_name"]
        for group_a, group_b, rel_direction, rel_type, strength in rule["relations"]:
            asset_a = preferred.get(group_a)
            asset_b = preferred.get(group_b)
            missing = []
            if asset_a is None:
                missing.append(group_a)
            if asset_b is None:
                missing.append(group_b)
            if missing:
                missing_rows.append(
                    {
                        "industry_name": industry,
                        "underlying_group_a": group_a,
                        "underlying_group_b": group_b,
                        "missing_underlying_groups": ";".join(missing),
                        "relation_direction": rel_direction,
                        "relation_type": rel_type,
                        "issue_reason": "产业链规则中的 underlying_group 在当前资产库中没有可用资产。",
                        "suggested_next_step": "后续如新增相关资产，可重新运行 Step 6 生成产业链映射。",
                    }
                )
                continue

            asset_id_a = clean_text(asset_a.get("asset_id"))
            asset_id_b = clean_text(asset_b.get("asset_id"))
            if not asset_id_a or not asset_id_b or asset_id_a == asset_id_b:
                missing_rows.append(
                    {
                        "industry_name": industry,
                        "underlying_group_a": group_a,
                        "underlying_group_b": group_b,
                        "missing_underlying_groups": "",
                        "relation_direction": rel_direction,
                        "relation_type": rel_type,
                        "issue_reason": "产业链关系无法形成有效资产对，可能存在空 asset_id 或自映射。",
                        "suggested_next_step": "检查 underlying_mapping_fixed 中对应品种的资产选择。",
                    }
                )
                continue

            pair_key = tuple(sorted([asset_id_a, asset_id_b]) + [rel_type])
            if pair_key in seen:
                continue
            seen.add(pair_key)
            counters[chain_key] += 1
            chain_id = f"{CHAIN_PREFIX[chain_key]}_{counters[chain_key]:03d}"
            segment_a = segment_for(rel_direction, rel_type, "A")
            segment_b = segment_for(rel_direction, rel_type, "B")
            rows.append(
                {
                    "chain_id": chain_id,
                    "industry_name": industry,
                    "chain_segment": f"{segment_a}_TO_{segment_b}" if segment_a != segment_b else segment_a,
                    "asset_id_a": asset_id_a,
                    "asset_id_b": asset_id_b,
                    "symbol_a": clean_text(asset_a.get("symbol")),
                    "symbol_b": clean_text(asset_b.get("symbol")),
                    "name_a": clean_text(asset_a.get("name_cn")),
                    "name_b": clean_text(asset_b.get("name_cn")),
                    "asset_type_a": normalize_asset_type(clean_text(asset_a.get("asset_type"))),
                    "asset_type_b": normalize_asset_type(clean_text(asset_b.get("asset_type"))),
                    "exchange_a": clean_text(asset_a.get("exchange_code")),
                    "exchange_b": clean_text(asset_b.get("exchange_code")),
                    "underlying_group_a": group_a,
                    "underlying_group_b": group_b,
                    "relation_direction": rel_direction,
                    "relation_type": rel_type,
                    "relation_strength": strength,
                    "strategy_type": strategy_for(rel_type, strength),
                    "mapping_confidence": strength,
                    "tradable_check": tradable_check(asset_a, asset_b),
                    "direction_supported": direction_supported(rel_direction),
                    "long_short_note": long_short_note(rel_direction, group_a, group_b),
                    "data_source": "STEP6_BUILT_IN_INDUSTRY_CHAIN_RULES",
                    "source_status": "INDUSTRY_CHAIN_MAPPING_CANDIDATE_NEED_REVIEW",
                    "notes": "基于 Step 5 fixed underlying_group 和内置产业链关系生成，非套利计算结果。",
                }
            )

    mapping = pd.DataFrame(rows)
    missing_review = pd.DataFrame(missing_rows)
    if not mapping.empty:
        mapping = mapping.drop_duplicates(["asset_id_a", "asset_id_b", "relation_type"]).reset_index(drop=True)
    save_csv(mapping, CHAIN_MAPPING_OUT)
    save_csv(missing_review, MISSING_REVIEW_OUT)
    return mapping, missing_review


def count_duplicate_relations(mapping: pd.DataFrame) -> int:
    if mapping.empty:
        return 0
    return int(mapping.duplicated(["asset_id_a", "asset_id_b", "relation_type"]).sum())


def count_self_mappings(mapping: pd.DataFrame) -> int:
    if mapping.empty:
        return 0
    return int((mapping["asset_id_a"] == mapping["asset_id_b"]).sum())


def build_quality_report(mapping: pd.DataFrame, asset_pool: pd.DataFrame, missing_review: pd.DataFrame) -> pd.DataFrame:
    duplicate_relation_count = count_duplicate_relations(mapping)
    self_mapping_count = count_self_mappings(mapping)
    rows = [
        {
            "section": "overall",
            "industry_name": "ALL",
            "relation_type": "ALL",
            "total_chain_mapping_count": len(mapping),
            "total_industry_count": mapping["industry_name"].nunique() if not mapping.empty else 0,
            "missing_underlying_group_count": int((asset_pool["mapping_status"] == "MISSING_ASSET").sum()) if not asset_pool.empty else 0,
            "missing_asset_relation_count": len(missing_review),
            "duplicate_relation_count": duplicate_relation_count,
            "self_mapping_count": self_mapping_count,
            "high_confidence_count": int((mapping["mapping_confidence"] == "HIGH").sum()) if not mapping.empty else 0,
            "medium_confidence_count": int((mapping["mapping_confidence"] == "MEDIUM").sum()) if not mapping.empty else 0,
            "low_confidence_count": int((mapping["mapping_confidence"] == "LOW").sum()) if not mapping.empty else 0,
            "mapping_count": len(mapping),
            "missing_asset_count": len(missing_review),
            "count": len(mapping),
        }
    ]

    industries = sorted({rule["industry_name"] for rule in CHAIN_RULES.values()})
    for industry in industries:
        group = mapping[mapping["industry_name"] == industry] if not mapping.empty else pd.DataFrame()
        missing = missing_review[missing_review["industry_name"] == industry] if not missing_review.empty else pd.DataFrame()
        rows.append(
            {
                "section": "by_industry_name",
                "industry_name": industry,
                "relation_type": "",
                "mapping_count": len(group),
                "missing_asset_count": len(missing),
                "high_confidence_count": int((group["mapping_confidence"] == "HIGH").sum()) if not group.empty else 0,
                "medium_confidence_count": int((group["mapping_confidence"] == "MEDIUM").sum()) if not group.empty else 0,
                "low_confidence_count": int((group["mapping_confidence"] == "LOW").sum()) if not group.empty else 0,
            }
        )

    if not mapping.empty:
        for relation_type, group in mapping.groupby("relation_type"):
            rows.append(
                {
                    "section": "by_relation_type",
                    "industry_name": "",
                    "relation_type": relation_type,
                    "count": len(group),
                }
            )

    report = pd.DataFrame(rows).fillna("")
    save_csv(report, QUALITY_REPORT_OUT)
    return report


def readme_df() -> pd.DataFrame:
    return pd.DataFrame(
        [
            ("项目", "Step 6 产业链映射表"),
            ("数据基础", "基于已有 asset_master 和 underlying_mapping_fixed 构建。"),
            ("数据获取", "不调用外部 API。"),
            ("用途", "产业链映射用于后续跨品种套利和产业链套利。"),
            ("范围", "当前以国内期货品种为主，ETF/期权可作为辅助标的。"),
            ("后续", "后续可根据研究需要扩展更细的产业链关系。"),
        ],
        columns=["item", "description"],
    )


def write_excel(mapping: pd.DataFrame, asset_pool: pd.DataFrame, quality: pd.DataFrame, missing_review: pd.DataFrame) -> None:
    sheets = {
        "README": readme_df(),
        "industry_chain_mapping": mapping,
        "industry_chain_asset_pool": asset_pool,
        "industry_chain_quality_report": quality,
        "industry_chain_missing_asset_review": missing_review,
    }
    with pd.ExcelWriter(EXCEL_OUT, engine="openpyxl") as writer:
        for name, df in sheets.items():
            sheet = name[:31]
            df.to_excel(writer, sheet_name=sheet, index=False)
            ws = writer.book[sheet]
            ws.freeze_panes = "A2"
            if ws.max_row > 1 and ws.max_column > 1:
                ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.font = Font(bold=True)
            for col_cells in ws.columns:
                width = 10
                letter = col_cells[0].column_letter
                for cell in col_cells[:200]:
                    width = max(width, min(len("" if cell.value is None else str(cell.value)) + 2, 40))
                ws.column_dimensions[letter].width = width


def main() -> None:
    try:
        ensure_dirs()
        asset_master, source_path = choose_asset_master()
        underlying = read_csv(UNDERLYING_MAPPING)
        read_csv(DOMESTIC_MAPPING)

        asset_pool, preferred = build_asset_pool(asset_master, underlying)
        mapping, missing_review = build_chain_mapping(preferred)
        quality = build_quality_report(mapping, asset_pool, missing_review)
        write_excel(mapping, asset_pool, quality, missing_review)

        self_mapping_count = count_self_mappings(mapping)
        duplicate_relation_count = count_duplicate_relations(mapping)
        print(f"使用的 asset_master 来源: {source_path}")
        print(f"industry_chain_mapping 行数: {len(mapping)}")
        print(f"industry_chain_asset_pool 行数: {len(asset_pool)}")
        print(f"missing review 行数: {len(missing_review)}")
        print("各 industry_name 数量:")
        if mapping.empty:
            print("  无")
        else:
            for industry, count in mapping["industry_name"].value_counts().sort_index().items():
                print(f"  {industry}: {count}")
        print("各 relation_type 数量:")
        if mapping.empty:
            print("  无")
        else:
            for relation_type, count in mapping["relation_type"].value_counts().sort_index().items():
                print(f"  {relation_type}: {count}")
        print(f"self_mapping_count: {self_mapping_count}")
        print(f"duplicate_relation_count: {duplicate_relation_count}")
        print(f"Excel 输出路径: {EXCEL_OUT}")
    except Exception as exc:
        print(f"Step 6 产业链映射生成失败: {exc}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
