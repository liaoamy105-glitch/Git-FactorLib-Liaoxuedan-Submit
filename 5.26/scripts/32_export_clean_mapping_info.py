from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font


ROOT = Path(__file__).resolve().parents[1]
ENCODING = "utf-8-sig"

ASSET_IN = ROOT / "data" / "stage5_mapping" / "final" / "asset_master_for_mapping_patched.csv"
UNDERLYING_IN = ROOT / "data" / "stage5_mapping" / "final" / "underlying_mapping_fixed.csv"
DOMESTIC_IN = ROOT / "data" / "stage5_mapping" / "final" / "domestic_cross_market_mapping_fixed.csv"
EXCEL_OUT = ROOT / "output" / "mapping_info.xlsx"

EMPTY_TEXT = "空"

FORBIDDEN_TERMS = [
    "Tushare",
    "AKShare",
    "AkShare",
    "akshare",
    "TqSdk",
    "TQSDK",
    "tqsdk",
    "待核验",
    "待人工复核",
    "NEED_REVIEW",
    "CHECK",
]

UNUSABLE_VALUES = {
    "",
    " ",
    "NA",
    "N/A",
    "NAN",
    "NONE",
    "NULL",
    "TODO",
    "TBD",
    "CHECK",
    "NEED_REVIEW",
    "UNKNOWN",
    "不确定",
    "没找到",
    "没用",
    "待核验",
    "待人工复核",
}


def read_csv(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"缺少输入文件: {path}")
    return pd.read_csv(path, dtype=str, encoding=ENCODING).fillna("")


def clean_cell(value: Any) -> str:
    text = "" if value is None or pd.isna(value) else str(value).strip()
    if text == "LOG":
        return "TIMBER"
    if text == "Log":
        return "Timber"
    if text.upper() in UNUSABLE_VALUES:
        return EMPTY_TEXT
    for term in FORBIDDEN_TERMS:
        if term in text:
            return EMPTY_TEXT
    return text if text else EMPTY_TEXT


def clean_df(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    existing = [col for col in columns if col in df.columns]
    cleaned = df[existing].copy()
    for col in cleaned.columns:
        cleaned[col] = cleaned[col].map(clean_cell)
    return cleaned


def build_asset_sheet(asset: pd.DataFrame, underlying: pd.DataFrame) -> pd.DataFrame:
    names = underlying[["asset_id", "underlying_name_cn", "underlying_name_en"]].drop_duplicates("asset_id")
    merged = asset.merge(names, on="asset_id", how="left")
    columns = [
        "asset_id",
        "asset_type",
        "subtype",
        "symbol",
        "name_cn",
        "name_en",
        "exchange_code",
        "exchange_name",
        "country",
        "currency",
        "underlying_group",
        "underlying_name_cn",
        "underlying_name_en",
        "sector",
        "tradable",
        "can_long",
        "can_short",
    ]
    return clean_df(merged, columns)


def build_underlying_sheet(underlying: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "asset_id",
        "asset_type",
        "subtype",
        "symbol",
        "name_cn",
        "exchange_code",
        "underlying_group",
        "underlying_name_cn",
        "underlying_name_en",
    ]
    return clean_df(underlying, columns)


def build_domestic_sheet(domestic: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "mapping_id",
        "asset_id_a",
        "asset_id_b",
        "symbol_a",
        "symbol_b",
        "name_a",
        "name_b",
        "asset_type_a",
        "asset_type_b",
        "exchange_a",
        "exchange_b",
        "underlying_group",
        "relation_type",
        "strategy_type",
        "market_relation",
        "tradable_check",
        "direction_supported",
        "long_short_note",
    ]
    cleaned = clean_df(domestic, columns)
    return cleaned.rename(
        columns={
            "tradable_check": "tradable_status",
            "long_short_note": "long_short_rule",
        }
    )


def build_relation_sheet(domestic_clean: pd.DataFrame) -> pd.DataFrame:
    group_cols = ["relation_type", "strategy_type"]
    if domestic_clean.empty:
        return pd.DataFrame(columns=group_cols + ["mapping_count"])
    summary = (
        domestic_clean.groupby(group_cols, dropna=False)
        .size()
        .reset_index(name="mapping_count")
        .sort_values(["relation_type", "strategy_type"])
    )
    return clean_df(summary, group_cols + ["mapping_count"])


def build_readme() -> pd.DataFrame:
    return pd.DataFrame(
        [
            ("文件说明", "用于后续映射分析和套利组合构建的清洁版表格。"),
            ("资产表", "保留资产识别、交易所、标的组、交易方向相关字段。"),
            ("统一标的表", "保留资产到统一标的组的映射字段。"),
            ("国内映射表", "保留资产对、关系类型、策略类型和交易方向字段。"),
            ("空值规则", "缺失、不可用或不确定内容统一写为空。"),
        ],
        columns=["item", "description"],
    )


def write_excel(sheets: dict[str, pd.DataFrame]) -> None:
    EXCEL_OUT.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(EXCEL_OUT, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name[:31], index=False)
            ws = writer.book[name[:31]]
            ws.freeze_panes = "A2"
            if ws.max_row > 1 and ws.max_column > 1:
                ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.font = Font(bold=True)
            for col_cells in ws.columns:
                width = 10
                letter = col_cells[0].column_letter
                for cell in col_cells[:300]:
                    width = max(width, min(len("" if cell.value is None else str(cell.value)) + 2, 40))
                ws.column_dimensions[letter].width = width


def assert_no_forbidden_terms(path: Path) -> None:
    wb = load_workbook(path, read_only=True, data_only=True)
    hits = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                text = "" if cell.value is None else str(cell.value)
                for term in FORBIDDEN_TERMS:
                    if term in text:
                        hits.append((ws.title, cell.coordinate, term, text))
    wb.close()
    if hits:
        sample = hits[:10]
        raise ValueError(f"清洁版 Excel 仍包含禁用字样: {sample}")


def main() -> None:
    asset = read_csv(ASSET_IN)
    underlying = read_csv(UNDERLYING_IN)
    domestic = read_csv(DOMESTIC_IN)

    asset_clean = build_asset_sheet(asset, underlying)
    underlying_clean = build_underlying_sheet(underlying)
    domestic_clean = build_domestic_sheet(domestic)
    relation_summary = build_relation_sheet(domestic_clean)

    write_excel(
        {
            "README": build_readme(),
            "asset_master": asset_clean,
            "underlying_mapping": underlying_clean,
            "domestic_mapping": domestic_clean,
            "relation_summary": relation_summary,
        }
    )
    assert_no_forbidden_terms(EXCEL_OUT)

    print(f"clean Excel 输出路径: {EXCEL_OUT}")
    print(f"asset_master 行数: {len(asset_clean)}")
    print(f"underlying_mapping 行数: {len(underlying_clean)}")
    print(f"domestic_mapping 行数: {len(domestic_clean)}")
    print(f"relation_summary 行数: {len(relation_summary)}")
    print("禁用字样扫描: PASS")


if __name__ == "__main__":
    main()
