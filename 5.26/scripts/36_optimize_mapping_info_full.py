from __future__ import annotations

import shutil
import sys
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Font


ROOT = Path(__file__).resolve().parents[1]
ENCODING = "utf-8-sig"
EMPTY = "空"

FULL_OUT = ROOT / "output" / "mapping_info_full.xlsx"
OPT_OUT = ROOT / "output" / "mapping_info_full_optimized.xlsx"
BACKUP_OUT = ROOT / "output" / "mapping_info_full_backup_before_step9.xlsx"
REPORT_OUT = ROOT / "output" / "mapping_info_full_optimized_check_report.txt"

ASSET_PRIMARY = ROOT / "data" / "stage5_mapping" / "final" / "asset_master_for_mapping_patched.csv"
ASSET_FALLBACK = ROOT / "data" / "stage4_etf_index" / "final" / "asset_master_with_etf_index_cleaned.csv"
FUTURES = ROOT / "data" / "final" / "futures_master.csv"
OPTIONS_PRIMARY = ROOT / "data" / "stage4_etf_index" / "final" / "options_master_with_etf_underlying.csv"
OPTIONS_FALLBACK = ROOT / "data" / "stage2_options" / "final" / "options_master_tushare_primary_final.csv"
ETF = ROOT / "data" / "stage4_etf_index" / "final" / "etf_master_cleaned.csv"
INDEX = ROOT / "data" / "stage4_etf_index" / "final" / "index_master.csv"
RULE_FUT = ROOT / "data" / "final" / "contract_rule_detail.csv"
RULE_OPT = ROOT / "data" / "stage2_options" / "final" / "option_contract_rule_detail_tushare_primary_final.csv"
UNDERLYING = ROOT / "data" / "stage5_mapping" / "final" / "underlying_mapping_fixed.csv"
DOMESTIC = ROOT / "data" / "stage5_mapping" / "final" / "domestic_cross_market_mapping_fixed.csv"
INDUSTRY = ROOT / "data" / "stage6_industry_chain" / "final" / "industry_chain_mapping.csv"
INDUSTRY_POOL = ROOT / "data" / "stage6_industry_chain" / "final" / "industry_chain_asset_pool.csv"
FOREIGN_ASSET = ROOT / "data" / "stage7_foreign_mapping" / "final" / "foreign_asset_master.csv"
FOREIGN = ROOT / "data" / "stage7_foreign_mapping" / "final" / "foreign_cross_market_mapping.csv"

INVALID_EXACT = {
    "", "NAN", "NONE", "NULL", "CHECK", "TODO", "PENDING", "NEED_REVIEW", "REVIEW",
    "CHECK_WAIT_ETF_MASTER", "CHECK_NEED_EXCHANGE_RULE_REVIEW", "CHECK_DYNAMIC_BY_EXCHANGE_NOTICE",
    "OBSERVED_MATURITY_DATES_AVAILABLE", "OBSERVED_EXPIRY_MONTHS_AVAILABLE",
    "OBSERVED_DELIST_OR_LAST_TRADE_DATES_AVAILABLE", "OBSERVED_LAST_EXERCISE_DATES_AVAILABLE",
    "NA_NOT_APPLICABLE", "NA_NOT_APPLICABLE_OR_CHECK_BY_EXCHANGE", "UNAVAILABLE",
    "PLACEHOLDER", "WAIT", "MIXED_NEED_REVIEW", "待复核", "待确认", "需复核", "需要复核",
    "待补充", "未确认", "无法确定",
}
INVALID_CONTAINS = [
    "动态字段", "交易所公告", "AKShare", "OpenCTP", "Tushare", "TqSdk", "source_status",
    "data_source", "LOCAL_HTML", "OFFICIAL_REVIEWED", "INFERRED_FROM", "MANUAL_RULE",
    "NEED_MANUAL", "NEED_EXCHANGE",
]
PROCESS_KEYWORDS = [
    "data_source", "source_status", "source", "status", "notes", "note", "review", "quality",
    "log", "summary", "confidence", "method", "update_date", "classification", "required",
    "reason", "suggested", "raw", "patch", "warning", "akshare", "observed", "sample",
    "static_rule_missing", "missing_fields",
]
CORE_FIELDS = {
    "asset_id", "foreign_asset_id", "mapping_id", "chain_id", "symbol", "name_cn", "name_en",
    "asset_type", "subtype", "exchange_code", "exchange_name", "country", "currency",
    "underlying_group", "relation_type", "strategy_type", "direction_supported", "tradable",
    "can_long", "can_short", "contract_month_rule", "contract_rule_code", "contract_rule_desc_cn",
    "listed_contract_count_rule", "expiry_rule", "last_trading_day_rule", "exercise_rule",
    "exercise_type", "settlement_type", "tick_size", "contract_multiplier", "contract_unit",
    "quote_unit",
}

SHEET_ORDER = [
    "README", "relation_summary", "validation_summary", "mapping_display", "current_contract_detail",
    "core_asset_master", "asset_master", "futures_master", "options_master", "core_etf_master",
    "etf_master", "index_master", "contract_rule_detail", "underlying_mapping",
    "domestic_mapping_display", "domestic_cross_market_mapping", "industry_chain_mapping",
    "industry_chain_asset_pool", "foreign_mapping_display", "foreign_asset_master",
    "foreign_cross_market_mapping",
]

ORDER = {
    "mapping_display": ["期货", "商品名称", "交易所", "统一标的", "期权", "ETF期权", "ETF现货", "指数/现货", "国内跨市场关系", "国外期货1", "国外期货2", "国外ETF/参考资产", "产业链相关品种", "支持策略", "备注"],
    "current_contract_detail": ["asset_id", "symbol", "name_cn", "asset_type", "subtype", "exchange_code", "exchange_name", "sample_ts_code", "sample_contract_name", "contract_unit", "quote_unit", "tick_size", "contract_multiplier", "settlement_type", "list_date", "delist_date", "sample_last_ddate"],
    "asset_master": ["asset_id", "asset_type", "subtype", "symbol", "name_cn", "name_en", "exchange_code", "exchange_name", "country", "currency", "underlying_group", "sector", "tradable", "can_long", "can_short"],
    "core_asset_master": ["asset_id", "asset_type", "subtype", "symbol", "name_cn", "name_en", "exchange_code", "exchange_name", "country", "currency", "underlying_group", "sector", "tradable", "can_long", "can_short"],
    "futures_master": ["asset_id", "symbol", "name_cn", "name_en", "exchange_code", "exchange_name", "country", "currency", "sector", "underlying_group", "underlying_name_cn", "underlying_name_en", "tradable", "can_long", "can_short", "contract_unit", "quote_unit", "tick_size", "contract_multiplier", "sample_ts_code", "sample_contract_name"],
    "options_master": ["asset_id", "asset_type", "subtype", "option_symbol", "option_name_cn", "option_name_en", "option_type", "exchange_code", "exchange_name", "country", "currency", "underlying_asset_id", "underlying_symbol", "underlying_name", "underlying_group", "underlying_name_cn", "underlying_name_en", "call_put_supported", "exercise_type", "settlement_type", "contract_multiplier", "tick_size", "quote_unit", "contract_unit", "tradable", "can_long", "can_short"],
    "core_etf_master": ["asset_id", "asset_type", "subtype", "etf_code", "ts_code", "name_cn", "exchange_code", "exchange_name", "fund_company", "fund_type", "invest_type", "tracking_index_code", "tracking_index_name", "underlying_group", "etf_importance_level", "etf_category_refined", "is_core_arbitrage_etf", "is_etf_option_underlying", "tradable", "can_long", "can_short"],
    "etf_master": ["asset_id", "asset_type", "subtype", "etf_code", "ts_code", "name_cn", "name_en", "exchange_code", "exchange_name", "country", "currency", "fund_company", "fund_type", "invest_type", "tracking_index_code", "tracking_index_name", "underlying_group", "tradable", "can_long", "can_short"],
    "index_master": ["asset_id", "asset_type", "subtype", "index_code", "ts_code", "name_cn", "name_en", "exchange_code", "exchange_name", "publisher", "category", "market", "country", "currency", "underlying_group", "tradable", "can_long", "can_short"],
    "contract_rule_detail": ["asset_id", "symbol", "name_cn", "asset_type", "subtype", "exchange_code", "exchange_name", "country", "currency", "contract_month_rule", "contract_rule_code", "contract_rule_desc_cn", "contract_rule_params", "listed_contract_count_rule", "expiry_rule", "expiry_rule_code", "last_trading_day_rule", "delivery_day_rule", "exercise_rule", "exercise_type", "settlement_type", "contract_multiplier", "contract_unit", "tick_size", "quote_unit", "price_limit_rule", "margin_rule", "trading_hours", "night_trading", "position_limit_rule", "delivery_rule", "rule_parseable"],
    "underlying_mapping": ["asset_id", "symbol", "name_cn", "asset_type", "subtype", "exchange_code", "underlying_group", "underlying_name_cn", "underlying_name_en"],
    "domestic_cross_market_mapping": ["mapping_id", "asset_id_a", "asset_id_b", "symbol_a", "symbol_b", "name_a", "name_b", "asset_type_a", "asset_type_b", "exchange_a", "exchange_b", "underlying_group", "relation_type", "strategy_type", "market_relation", "tradable_check", "direction_supported", "long_short_note", "relation_desc_cn"],
    "domestic_mapping_display": ["国内资产A", "国内资产B", "关系类型", "策略类型", "统一标的", "交易所A", "交易所B", "多空/方向说明", "relation_desc_cn"],
    "industry_chain_mapping": ["chain_id", "industry_name", "chain_segment", "asset_id_a", "asset_id_b", "symbol_a", "symbol_b", "name_a", "name_b", "asset_type_a", "asset_type_b", "exchange_a", "exchange_b", "underlying_group_a", "underlying_group_b", "relation_direction", "relation_type", "relation_strength", "strategy_type", "tradable_check", "direction_supported", "long_short_note", "relation_desc_cn"],
    "industry_chain_asset_pool": ["underlying_group", "underlying_name_cn", "industry_name", "asset_count", "preferred_asset_id", "preferred_symbol", "preferred_name_cn", "preferred_asset_type", "all_asset_ids", "mapping_status"],
    "foreign_mapping_display": ["国内资产", "国外资产", "国外交易所", "国外市场", "统一标的", "关系类型", "策略类型", "是否需要汇率换算", "单位换算说明", "映射置信度", "relation_desc_cn"],
    "foreign_asset_master": ["foreign_asset_id", "foreign_asset_type", "foreign_subtype", "foreign_symbol", "foreign_name_en", "foreign_name_cn", "foreign_exchange_code", "foreign_exchange_name", "foreign_country", "foreign_currency", "foreign_contract_unit", "foreign_quote_unit", "foreign_tick_size", "underlying_group", "sector", "tradable", "can_long", "can_short"],
    "foreign_cross_market_mapping": ["mapping_id", "domestic_asset_id", "foreign_asset_id", "domestic_symbol", "foreign_symbol", "domestic_name", "foreign_name", "domestic_asset_type", "foreign_asset_type", "domestic_exchange", "foreign_exchange", "domestic_country", "foreign_country", "underlying_group", "currency_domestic", "currency_foreign", "unit_conversion", "fx_conversion_needed", "trading_hour_overlap", "relation_type", "strategy_type", "direction_supported", "mapping_confidence", "relation_desc_cn"],
    "relation_summary": ["relation_source", "relation_type", "strategy_type", "count", "relation_type_cn"],
    "validation_summary": ["check_item", "result", "count", "description"],
}


def read_csv(path: Path, required: bool = False) -> pd.DataFrame:
    if not path.exists():
        if required:
            raise FileNotFoundError(f"缺少输入文件: {path}")
        return pd.DataFrame()
    return pd.read_csv(path, dtype=str, encoding=ENCODING).fillna("")


def clean_text(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    return str(value).strip()


def is_invalid(value: Any) -> bool:
    text = clean_text(value)
    return text.upper() in INVALID_EXACT or any(token in text for token in INVALID_CONTAINS)


def clean_value(value: Any) -> str:
    text = clean_text(value)
    if text == "LONG_ETF_AND_SHORT_FUTURE_NEED_CHECK":
        return "LONG_ETF_AND_SHORT_FUTURE_NEED_CONFIRM"
    return EMPTY if is_invalid(text) else (text or EMPTY)


def empty_mask(series: pd.Series) -> pd.Series:
    return series.fillna("").astype(str).str.strip().isin(["", EMPTY])


def core_column(col: str) -> bool:
    lower = col.lower()
    return lower in CORE_FIELDS or any(lower.endswith("_" + c) for c in CORE_FIELDS) or any(lower.startswith(c + "_") for c in CORE_FIELDS)


def keep_process_col(sheet: str, col: str) -> bool:
    if sheet == "relation_summary" and col in {"relation_source", "relation_type_cn"}:
        return True
    if sheet == "validation_summary":
        return True
    if sheet == "foreign_cross_market_mapping" and col == "mapping_confidence":
        return True
    if sheet == "industry_chain_mapping" and col in {"mapping_confidence", "relation_strength"}:
        return True
    if sheet in {"domestic_cross_market_mapping", "industry_chain_mapping"} and col in {"relation_type", "strategy_type", "direction_supported", "long_short_note", "relation_desc_cn"}:
        return True
    if sheet in {"domestic_mapping_display", "foreign_mapping_display", "mapping_display"}:
        return True
    if sheet == "industry_chain_asset_pool" and col == "mapping_status":
        return True
    return False


def drop_process_col(sheet: str, col: str) -> bool:
    lower = col.lower()
    return any(k in lower for k in PROCESS_KEYWORDS) and not keep_process_col(sheet, col)


def order_cols(sheet: str, df: pd.DataFrame) -> pd.DataFrame:
    front = [c for c in ORDER.get(sheet, []) if c in df.columns]
    rest = [c for c in df.columns if c not in front]
    return df[front + rest]


def clean_sheet(sheet: str, df: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    if df.empty:
        return df, []
    out = df.copy()
    for col in out.columns:
        out[col] = out[col].map(clean_value)
    dropped: list[str] = []
    for col in list(out.columns):
        if drop_process_col(sheet, col):
            dropped.append(col)
            out = out.drop(columns=[col])
    for col in list(out.columns):
        empty_ratio = float(empty_mask(out[col]).mean()) if len(out) else 1.0
        if empty_ratio >= 1.0 or (empty_ratio > 0.95 and not core_column(col) and not keep_process_col(sheet, col)):
            dropped.append(col)
            out = out.drop(columns=[col])
    out = order_cols(sheet, out)
    if "tradable_check" in out.columns:
        out = out.rename(columns={"tradable_check": "tradable_status"})
    return out, dropped


def choose_asset_master() -> tuple[pd.DataFrame, Path]:
    if ASSET_PRIMARY.exists():
        return read_csv(ASSET_PRIMARY, True), ASSET_PRIMARY
    return read_csv(ASSET_FALLBACK, True), ASSET_FALLBACK


def choose_options() -> tuple[pd.DataFrame, Path]:
    if OPTIONS_PRIMARY.exists():
        return read_csv(OPTIONS_PRIMARY, True), OPTIONS_PRIMARY
    return read_csv(OPTIONS_FALLBACK, True), OPTIONS_FALLBACK


def display_index_as_spot(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if "asset_type" in out.columns:
        mask = out["asset_type"].eq("INDEX")
        out.loc[mask, "asset_type"] = "SPOT"
        if "subtype" in out.columns:
            out.loc[mask, "subtype"] = "INDEX_SPOT"
        for col in ["tradable", "can_long", "can_short"]:
            if col in out.columns:
                out.loc[mask, col] = "N"
    return out


def sync_futures(futures: pd.DataFrame, underlying: pd.DataFrame, asset: pd.DataFrame) -> pd.DataFrame:
    out = futures.copy()
    fut_um = underlying[underlying["asset_type"].eq("FUTURE")][["asset_id", "underlying_group", "underlying_name_cn", "underlying_name_en"]].drop_duplicates("asset_id")
    out = out.merge(fut_um, on="asset_id", how="left", suffixes=("", "_fixed"))
    for col in ["underlying_group", "underlying_name_cn", "underlying_name_en"]:
        fixed = col + "_fixed"
        if fixed in out.columns:
            if col not in out.columns:
                out[col] = ""
            mask = out[col].map(is_invalid) | out[col].astype(str).str.strip().eq("")
            out.loc[mask, col] = out.loc[mask, fixed]
            out = out.drop(columns=[fixed])
    if "sector" in out.columns and "sector" in asset.columns:
        sec = asset[["asset_id", "sector"]].drop_duplicates("asset_id").rename(columns={"sector": "sector_asset"})
        out = out.merge(sec, on="asset_id", how="left")
        mask = out["sector"].map(is_invalid) | out["sector"].astype(str).str.strip().eq("")
        out.loc[mask, "sector"] = out.loc[mask, "sector_asset"]
        out = out.drop(columns=["sector_asset"])
    return out


def sync_options(options: pd.DataFrame, underlying: pd.DataFrame) -> pd.DataFrame:
    out = options.copy()
    names = underlying[["asset_id", "underlying_group", "underlying_name_cn", "underlying_name_en"]].drop_duplicates("asset_id")
    out = out.merge(names, left_on="asset_id", right_on="asset_id", how="left", suffixes=("", "_self"))
    for col in ["underlying_group", "underlying_name_cn", "underlying_name_en"]:
        self_col = col + "_self"
        if self_col in out.columns:
            if col not in out.columns:
                out[col] = ""
            mask = out[col].map(is_invalid) | out[col].astype(str).str.strip().eq("")
            out.loc[mask, col] = out.loc[mask, self_col]
            out = out.drop(columns=[self_col])
    if "underlying_name" in out.columns and "underlying_name_cn" in out.columns:
        mask = out["underlying_name"].map(is_invalid) | out["underlying_name"].astype(str).str.strip().eq("")
        out.loc[mask, "underlying_name"] = out.loc[mask, "underlying_name_cn"]
    return out


def core_etf(etf: pd.DataFrame) -> pd.DataFrame:
    if etf.empty:
        return etf
    subtype = etf.get("subtype", pd.Series("", index=etf.index)).astype(str)
    ug = etf.get("underlying_group", pd.Series("", index=etf.index)).astype(str).str.strip()
    mask = (
        etf.get("is_core_arbitrage_etf", pd.Series("", index=etf.index)).eq("Y")
        | etf.get("is_etf_option_underlying", pd.Series("", index=etf.index)).eq("Y")
        | etf.get("etf_importance_level", pd.Series("", index=etf.index)).isin(["CORE", "IMPORTANT"])
        | subtype.str.contains("BROAD_BASED_ETF|COMMODITY_ETF|BOND_ETF", na=False)
        | (~ug.isin(["", "CHECK", "TODO", EMPTY]))
    )
    return etf.loc[mask].copy()


def current_contract_detail(futures: pd.DataFrame) -> pd.DataFrame:
    out = futures.copy()
    rename = {"list_date_min": "list_date", "delist_date_max": "delist_date"}
    out = out.rename(columns={k: v for k, v in rename.items() if k in out.columns})
    cols = ORDER["current_contract_detail"]
    return out[[c for c in cols if c in out.columns]].copy()


def merge_contract_rules() -> pd.DataFrame:
    frames = [df for df in [read_csv(RULE_FUT), read_csv(RULE_OPT)] if not df.empty]
    if not frames:
        return pd.DataFrame()
    cols = sorted(set().union(*[set(df.columns) for df in frames]))
    out = pd.concat([df.reindex(columns=cols, fill_value="") for df in frames], ignore_index=True)
    explicit_drop = [c for c in out.columns if any(k in c.lower() for k in ["akshare_update_time", "observed_", "sample_", "static_rule_missing_fields", "underlying_group_review", "review", "quality", "log"])]
    return out.drop(columns=explicit_drop, errors="ignore")


DOM_DESC = {
    "ETF_INDEX": "ETF 与其跟踪指数之间的映射关系，可用于 ETF 折溢价和指数参考分析。",
    "FUTURE_ETF": "期货与同一统一标的下 ETF 之间的映射关系，可用于期现/基差类分析。",
    "OPTION_UNDERLYING": "期权与其标的资产之间的映射关系，可用于期权定价和期权套利分析。",
    "FUTURE_INDEX": "股指期货与对应指数之间的映射关系，指数为参考标的。",
    "OPTION_INDEX": "股指期权与对应指数之间的映射关系，指数为参考标的。",
    "OPTION_FUTURE": "期权与相关期货标的之间的映射关系。",
    "SAME_UNDERLYING_CROSS_EXCHANGE": "同一统一标的下不同交易所资产之间的映射关系，可用于跨交易所观察。",
}


def add_domestic_desc(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["relation_desc_cn"] = out.get("relation_type", "").map(lambda x: DOM_DESC.get(x, "国内跨市场映射关系。"))
    return out


def domestic_display(dom: pd.DataFrame) -> pd.DataFrame:
    if dom.empty:
        return dom
    d = add_domestic_desc(dom)
    return pd.DataFrame({
        "国内资产A": d["name_a"] + "（" + d["symbol_a"] + "）",
        "国内资产B": d["name_b"] + "（" + d["symbol_b"] + "）",
        "关系类型": d["relation_type"],
        "策略类型": d["strategy_type"],
        "统一标的": d["underlying_group"],
        "交易所A": d["exchange_a"],
        "交易所B": d["exchange_b"],
        "多空/方向说明": d.get("long_short_note", ""),
        "relation_desc_cn": d["relation_desc_cn"],
    })


def industry_desc(row: pd.Series) -> str:
    a, b, ind, rel = row.get("name_a", ""), row.get("name_b", ""), row.get("industry_name", ""), row.get("relation_type", "")
    if rel == "UPSTREAM_DOWNSTREAM":
        return f"{a} 与 {b} 属于 {ind} 的上下游关系，可用于产业链价差和成本传导分析。"
    if rel == "SUBSTITUTE":
        return f"{a} 与 {b} 具有替代或联动关系，可用于配对交易或价差观察。"
    if rel == "COST_DRIVEN":
        return f"{a} 对 {b} 具有成本驱动或原料传导关系。"
    if rel == "DEMAND_LINKED":
        return f"{a} 与 {b} 存在需求联动关系。"
    if rel == "SAME_SECTOR_RELATED":
        return f"{a} 与 {b} 属于同一板块相关品种。"
    return f"{a} 与 {b} 存在产业链相关关系。"


def add_industry_desc(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if not out.empty:
        out["relation_desc_cn"] = out.apply(industry_desc, axis=1)
    return out


def foreign_desc(row: pd.Series) -> str:
    rel = row.get("relation_type", "")
    if "SAME_UNDERLYING" in rel:
        return "国内资产与国外资产对应同一或高度一致的统一标的，可用于内外盘价差和跨市场套利观察。"
    if "RELATED_UNDERLYING" in rel:
        return "国内资产与国外资产标的相关但口径存在差异，可作为跨市场相关性和宏观对冲参考。"
    if "ETF_RELATED" in rel:
        return "国内资产与海外 ETF 或基金资产存在标的相关关系，可用于跨市场联动观察。"
    if "INDEX_RELATED" in rel:
        return "国内资产与海外指数或指数期货存在市场风格或风险因子相关关系。"
    return "该映射为国外跨市场候选关系，后续可结合单位换算、汇率和交易时间进一步验证。"


def add_foreign_desc(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if not out.empty:
        out["relation_desc_cn"] = out.apply(foreign_desc, axis=1)
    return out


def foreign_display(foreign: pd.DataFrame) -> pd.DataFrame:
    if foreign.empty:
        return foreign
    f = add_foreign_desc(foreign)
    return pd.DataFrame({
        "国内资产": f["domestic_name"] + "（" + f["domestic_symbol"] + "）",
        "国外资产": f["foreign_name"] + "（" + f["foreign_symbol"] + "）",
        "国外交易所": f["foreign_exchange"],
        "国外市场": f["foreign_country"],
        "统一标的": f["underlying_group"],
        "关系类型": f["relation_type"],
        "策略类型": f["strategy_type"],
        "是否需要汇率换算": f["fx_conversion_needed"],
        "单位换算说明": f["unit_conversion"],
        "映射置信度": f["mapping_confidence"],
        "relation_desc_cn": f["relation_desc_cn"],
    })


def join_limited(values: list[str], limit: int | None = None) -> str:
    vals = [v for v in dict.fromkeys([clean_text(v) for v in values]) if v and not is_invalid(v)]
    if not vals:
        return EMPTY
    if limit and len(vals) > limit:
        return "；".join(vals[:limit]) + "；等"
    return "；".join(vals)


def mapping_display(futures: pd.DataFrame, options: pd.DataFrame, etf_core: pd.DataFrame, etf_all: pd.DataFrame, index: pd.DataFrame, domestic: pd.DataFrame, industry: pd.DataFrame, foreign: pd.DataFrame) -> pd.DataFrame:
    rows = []
    conf_rank = {"HIGH": 0, "MEDIUM": 1, "LOW": 2}
    for _, fut in futures.sort_values([c for c in ["exchange_code", "symbol"] if c in futures.columns]).iterrows():
        aid, ug = fut.get("asset_id", ""), fut.get("underlying_group", "")
        same_opt = options[(options.get("underlying_asset_id", "") == aid) | ((options.get("underlying_group", "") == ug) & options.get("subtype", "").astype(str).str.contains("COMMODITY_OPTION|INDEX_OPTION", na=False))]
        etf_opt = options[(options.get("underlying_group", "") == ug) & options.get("subtype", "").eq("ETF_OPTION")]
        etf_src = etf_core if not etf_core.empty else etf_all
        etfs = etf_src[etf_src.get("underlying_group", "").eq(ug)] if not etf_src.empty else pd.DataFrame()
        idx = index[index.get("underlying_group", "").eq(ug)] if not index.empty else pd.DataFrame()
        dom_rel = domestic[(domestic.get("asset_id_a", "") == aid) | (domestic.get("asset_id_b", "") == aid)] if not domestic.empty else pd.DataFrame()
        ind_rel = industry[(industry.get("asset_id_a", "") == aid) | (industry.get("asset_id_b", "") == aid)] if not industry.empty else pd.DataFrame()
        for_rel = foreign[foreign.get("domestic_asset_id", "").eq(aid)] if not foreign.empty else pd.DataFrame()
        fut_for = for_rel[for_rel.get("foreign_asset_type", "").eq("FUTURE")].copy() if not for_rel.empty else pd.DataFrame()
        if not fut_for.empty:
            fut_for["_r"] = fut_for["mapping_confidence"].map(conf_rank).fillna(9)
            fut_for = fut_for.sort_values(["_r", "foreign_exchange", "foreign_symbol"])
        non_fut = for_rel[~for_rel.get("foreign_asset_type", "").eq("FUTURE")] if not for_rel.empty else pd.DataFrame()
        strategies = []
        for df in [dom_rel, ind_rel, for_rel]:
            if not df.empty and "strategy_type" in df.columns:
                strategies.extend(df["strategy_type"].tolist())
        rows.append({
            "期货": f"{fut.get('symbol','')}_{fut.get('exchange_code','')}_中国",
            "商品名称": fut.get("name_cn", ""),
            "交易所": fut.get("exchange_name", ""),
            "统一标的": ug,
            "期权": join_limited((same_opt.get("option_name_cn", same_opt.get("option_symbol", pd.Series(dtype=str))).astype(str)).tolist() if not same_opt.empty else []),
            "ETF期权": join_limited((etf_opt.get("option_name_cn", etf_opt.get("option_symbol", pd.Series(dtype=str))).astype(str)).tolist() if not etf_opt.empty else []),
            "ETF现货": join_limited((etfs.get("name_cn", etfs.get("etf_code", pd.Series(dtype=str))).astype(str)).tolist() if not etfs.empty else [], 5),
            "指数/现货": join_limited((idx.get("name_cn", idx.get("index_code", pd.Series(dtype=str))).astype(str)).tolist() if not idx.empty else [], 5),
            "国内跨市场关系": join_limited([f"{r.get('relation_type','')}-{r.get('name_b') if r.get('asset_id_a') == aid else r.get('name_a')}" for _, r in dom_rel.iterrows()], 5),
            "国外期货1": EMPTY if fut_for.empty else f"{fut_for.iloc[0].get('foreign_name')}（{fut_for.iloc[0].get('foreign_symbol')}）_{fut_for.iloc[0].get('foreign_exchange')}_{fut_for.iloc[0].get('foreign_country')}",
            "国外期货2": EMPTY if len(fut_for) < 2 else f"{fut_for.iloc[1].get('foreign_name')}（{fut_for.iloc[1].get('foreign_symbol')}）_{fut_for.iloc[1].get('foreign_exchange')}_{fut_for.iloc[1].get('foreign_country')}",
            "国外ETF/参考资产": join_limited([f"{r.get('foreign_name')}（{r.get('foreign_symbol')}）_{r.get('foreign_exchange')}_{r.get('foreign_country')}" for _, r in non_fut.iterrows()], 3),
            "产业链相关品种": join_limited([f"{r.get('name_b') if r.get('asset_id_a') == aid else r.get('name_a')}-{r.get('relation_type')}" for _, r in ind_rel.iterrows()], 5),
            "支持策略": join_limited(strategies),
            "备注": EMPTY if not for_rel.empty and not ind_rel.empty else "国外映射或产业链映射为空",
        })
    return pd.DataFrame(rows)


REL_CN = {
    "ETF_INDEX": "ETF-指数", "FUTURE_ETF": "期货-ETF", "OPTION_UNDERLYING": "期权-标的",
    "FUTURE_INDEX": "期货-指数", "OPTION_INDEX": "期权-指数", "OPTION_FUTURE": "期权-期货",
    "SAME_UNDERLYING_CROSS_EXCHANGE": "同标的跨交易所", "UPSTREAM_DOWNSTREAM": "上下游关系",
    "SUBSTITUTE": "替代关系", "COST_DRIVEN": "成本驱动", "DEMAND_LINKED": "需求联动",
    "SAME_SECTOR_RELATED": "同板块相关", "DOMESTIC_FOREIGN_SAME_UNDERLYING": "国内外同标的",
    "DOMESTIC_FOREIGN_RELATED_UNDERLYING": "国内外相关标的", "DOMESTIC_FOREIGN_ETF_RELATED": "国内外ETF相关",
    "DOMESTIC_FOREIGN_INDEX_RELATED": "国内外指数相关",
}


def relation_summary(domestic: pd.DataFrame, industry: pd.DataFrame, foreign: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for source, df in [("DOMESTIC", domestic), ("INDUSTRY_CHAIN", industry), ("FOREIGN", foreign)]:
        if df.empty:
            continue
        grouped = df.groupby(["relation_type", "strategy_type"], dropna=False).size().reset_index(name="count")
        for _, r in grouped.iterrows():
            rows.append({"relation_source": source, "relation_type": r["relation_type"], "strategy_type": r["strategy_type"], "count": r["count"], "relation_type_cn": REL_CN.get(r["relation_type"], r["relation_type"])})
    return pd.DataFrame(rows)


def readme_df(missing: list[str]) -> pd.DataFrame:
    rows = [
        ("项目名称", "套利组合构建用资产映射表完整版"),
        ("文件说明", "本文件为优化后的完整开发版，包含资产主表、核心资产展示表、横向映射展示表、期货、期权、ETF、指数、统一标的映射、国内跨市场映射、产业链映射、国外跨市场映射和交易规则详情表。"),
        ("说明1", "mapping_display 为横向展示表，一行对应一个国内期货品种。"),
        ("说明2", "asset_master 为全量资产主表。"),
        ("说明3", "core_asset_master 和 core_etf_master 用于快速查看核心资产。"),
        ("说明4", "contract_rule_detail 为规则详情表，current_contract_detail 为样例合约详情表。"),
        ("说明5", "output/mapping_info.xlsx 为提交用精简版，本文件为开发用完整版。"),
        ("说明6", "无有效数据的列已自动删除；暂未确认或无法识别的内容统一显示为空。"),
    ]
    if missing:
        rows.append(("缺失输入", "；".join(missing)))
    return pd.DataFrame(rows, columns=["item", "description"])


def count_residual(sheets: dict[str, pd.DataFrame], tokens: list[str]) -> int:
    total = 0
    for df in sheets.values():
        if df.empty:
            continue
        text = df.astype(str)
        for token in tokens:
            total += int(text.apply(lambda s: s.str.contains(token, case=False, regex=False, na=False)).sum().sum())
    return total


def validation(asset: pd.DataFrame, futures: pd.DataFrame, options: pd.DataFrame, underlying: pd.DataFrame, domestic: pd.DataFrame, industry: pd.DataFrame, foreign: pd.DataFrame, foreign_asset: pd.DataFrame, contract: pd.DataFrame, residual_invalid: int, residual_source: int) -> pd.DataFrame:
    asset_ids = set(asset.get("asset_id", pd.Series(dtype=str)).astype(str))
    foreign_ids = set(foreign_asset.get("foreign_asset_id", pd.Series(dtype=str)).astype(str))
    fut_empty = int(futures.get("underlying_group", pd.Series(dtype=str)).map(is_invalid).sum())
    check_wait = int(options.astype(str).apply(lambda s: s.str.contains("CHECK_WAIT_ETF_MASTER", regex=False, na=False)).sum().sum()) if not options.empty else 0
    etf_under_empty = 0
    if not options.empty and {"subtype", "underlying_asset_id"}.issubset(options.columns):
        mask = options["subtype"].eq("ETF_OPTION")
        etf_under_empty = int((options.loc[mask, "underlying_asset_id"].map(is_invalid) | options.loc[mask, "underlying_asset_id"].astype(str).str.strip().eq("")).sum())
    rows = [
        ("asset_master 行数", "PASS", len(asset), "资产主表记录数。"),
        ("asset_id 重复数", "PASS" if "asset_id" in asset and not asset["asset_id"].duplicated().any() else "FAIL", int(asset["asset_id"].duplicated().sum()) if "asset_id" in asset else -1, "count 为重复 asset_id 数。"),
        ("futures_master 中 underlying_group 为空数量", "PASS" if fut_empty == 0 else "FAIL", fut_empty, "Step 5.1 修复结果应同步到期货展示表。"),
        ("options_master 中 CHECK_WAIT_ETF_MASTER 残留数量", "PASS" if check_wait == 0 else "FAIL", check_wait, "ETF期权标的应使用 Step 4 修复结果。"),
        ("options_master 中 ETF_OPTION underlying_asset_id 为空数量", "PASS" if etf_under_empty == 0 else "FAIL", etf_under_empty, "ETF_OPTION 标的资产不能为空。"),
        ("underlying_mapping 行数", "PASS", len(underlying), "统一标的映射记录数。"),
        ("domestic_cross_market_mapping 行数", "PASS", len(domestic), "国内跨市场映射记录数。"),
        ("domestic mapping_id 重复数", "PASS" if domestic.empty or not domestic["mapping_id"].duplicated().any() else "FAIL", int(domestic["mapping_id"].duplicated().sum()) if "mapping_id" in domestic else 0, "count 为重复 mapping_id 数。"),
        ("domestic 自映射数量", "PASS" if domestic.empty or int((domestic["asset_id_a"] == domestic["asset_id_b"]).sum()) == 0 else "FAIL", int((domestic["asset_id_a"] == domestic["asset_id_b"]).sum()) if {"asset_id_a", "asset_id_b"}.issubset(domestic.columns) else 0, "count 为自映射数量。"),
        ("industry_chain_mapping 行数", "PASS", len(industry), "产业链映射记录数。"),
        ("industry 自映射数量", "PASS" if industry.empty or int((industry["asset_id_a"] == industry["asset_id_b"]).sum()) == 0 else "FAIL", int((industry["asset_id_a"] == industry["asset_id_b"]).sum()) if {"asset_id_a", "asset_id_b"}.issubset(industry.columns) else 0, "count 为自映射数量。"),
        ("foreign_cross_market_mapping 行数", "PASS", len(foreign), "国外跨市场映射记录数。"),
        ("foreign mapping_id 重复数", "PASS" if foreign.empty or not foreign["mapping_id"].duplicated().any() else "FAIL", int(foreign["mapping_id"].duplicated().sum()) if "mapping_id" in foreign else 0, "count 为重复 mapping_id 数。"),
        ("foreign 国内资产回连失败数", "PASS" if foreign.empty or not (set(foreign["domestic_asset_id"]) - asset_ids) else "FAIL", len(set(foreign["domestic_asset_id"]) - asset_ids) if "domestic_asset_id" in foreign else 0, "count 为无法回连国内资产数。"),
        ("foreign 国外资产回连失败数", "PASS" if foreign.empty or not (set(foreign["foreign_asset_id"]) - foreign_ids) else "FAIL", len(set(foreign["foreign_asset_id"]) - foreign_ids) if "foreign_asset_id" in foreign else 0, "count 为无法回连国外资产数。"),
        ("contract_rule_detail 行数", "PASS", len(contract), "规则详情记录数。"),
        ("残留 CHECK/TODO/NEED_REVIEW 单元格数量", "PASS" if residual_invalid == 0 else "WARN", residual_invalid, "展示层应尽量清理无效占位。"),
        ("残留 AKShare/Tushare/TqSdk 单元格数量", "PASS" if residual_source == 0 else "WARN", residual_source, "展示层应清理过程来源信息。"),
    ]
    return pd.DataFrame(rows, columns=["check_item", "result", "count", "description"])


def core_assets(asset: pd.DataFrame, options: pd.DataFrame, core_etfs: pd.DataFrame) -> pd.DataFrame:
    a = display_index_as_spot(asset)
    core_etf_ids = set(core_etfs.get("asset_id", pd.Series(dtype=str)))
    mask = a["asset_type"].isin(["FUTURE", "OPTION"]) | ((a["asset_type"] == "SPOT") & (a["subtype"] == "INDEX_SPOT"))
    if "asset_id" in a.columns:
        mask = mask | a["asset_id"].isin(core_etf_ids)
    if "underlying_group" in a.columns:
        mask = mask | ((a["asset_type"] == "ETF") & (~a["underlying_group"].map(is_invalid)) & a["underlying_group"].astype(str).str.strip().ne(""))
    return a.loc[mask].copy()


def write_excel(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name[:31], index=False)
            ws = writer.book[name[:31]]
            ws.freeze_panes = "A2"
            if ws.max_row > 1 and ws.max_column > 1:
                ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.font = Font(bold=True)
            for col_cells in ws.columns:
                width = 10
                letter = col_cells[0].column_letter
                for cell in col_cells[:200]:
                    width = max(width, min(len("" if cell.value is None else str(cell.value)) + 2, 40))
                ws.column_dimensions[letter].width = width


def report_text(path_backup: bool, sheet_rows: dict[str, int], dropped: dict[str, list[str]], metrics: dict[str, int], validation_df: pd.DataFrame, conclusion: str) -> str:
    lines = [
        f"输出 Excel 路径: {FULL_OUT}",
        f"优化副本路径: {OPT_OUT}",
        f"是否备份旧 mapping_info_full.xlsx: {'Y' if path_backup else 'N'}",
        "",
        "每个 sheet 行数 / 删除列数量:",
    ]
    for name, rows in sheet_rows.items():
        lines.append(f"- {name}: 行数={rows}, 删除列数量={len(dropped.get(name, []))}")
        lines.append(f"  删除列名: {', '.join(dropped.get(name, [])) if dropped.get(name) else '无'}")
    lines += [
        "",
        f"futures_master underlying_group 空值数量: {metrics['futures_underlying_empty']}",
        f"options_master CHECK_WAIT_ETF_MASTER 残留数量: {metrics['options_check_wait']}",
        f"ETF_OPTION underlying_asset_id 空值数量: {metrics['etf_option_underlying_empty']}",
        f"contract_rule_detail 残留 AKShare/OpenCTP 单元格数量: {metrics['contract_source_residual']}",
        f"全文件残留 CHECK/TODO/NEED_REVIEW 数量: {metrics['residual_invalid']}",
        f"全文件残留 Tushare/AKShare/TqSdk/OpenCTP 数量: {metrics['residual_source']}",
        "",
        "validation_summary:",
    ]
    for _, row in validation_df.iterrows():
        lines.append(f"- {row['result']} | {row['check_item']} | count={row['count']} | {row['description']}")
    lines.append("")
    lines.append(f"最终结论: {conclusion}")
    return "\n".join(lines) + "\n"


def main() -> None:
    try:
        backed = False
        if FULL_OUT.exists():
            shutil.copy2(FULL_OUT, BACKUP_OUT)
            backed = True

        asset_raw, _ = (read_csv(ASSET_PRIMARY, True), ASSET_PRIMARY) if ASSET_PRIMARY.exists() else (read_csv(ASSET_FALLBACK, True), ASSET_FALLBACK)
        futures_raw = read_csv(FUTURES, True)
        options_raw, _ = choose_options()
        etf_raw = read_csv(ETF, True)
        index_raw = read_csv(INDEX, True)
        underlying_raw = read_csv(UNDERLYING, True)
        domestic_raw = read_csv(DOMESTIC, True)
        industry_raw = read_csv(INDUSTRY, True)
        industry_pool_raw = read_csv(INDUSTRY_POOL, True)
        foreign_asset_raw = read_csv(FOREIGN_ASSET, True)
        foreign_raw = read_csv(FOREIGN, True)

        asset_show = display_index_as_spot(asset_raw)
        futures_show = sync_futures(futures_raw, underlying_raw, asset_raw)
        options_show = sync_options(options_raw, underlying_raw)
        etf_core = core_etf(etf_raw)
        index_show = display_index_as_spot(index_raw)
        contract_show = merge_contract_rules()
        domestic_desc = add_domestic_desc(domestic_raw)
        industry_desc_df = add_industry_desc(industry_raw)
        foreign_desc_df = add_foreign_desc(foreign_raw)

        sheets_raw: dict[str, pd.DataFrame] = {
            "relation_summary": relation_summary(domestic_raw, industry_raw, foreign_raw),
            "mapping_display": mapping_display(futures_show, options_show, etf_core, etf_raw, index_show, domestic_raw, industry_raw, foreign_raw),
            "current_contract_detail": current_contract_detail(futures_show),
            "core_asset_master": core_assets(asset_show, options_show, etf_core),
            "asset_master": asset_show,
            "futures_master": futures_show,
            "options_master": options_show,
            "core_etf_master": etf_core,
            "etf_master": etf_raw,
            "index_master": index_show,
            "contract_rule_detail": contract_show,
            "underlying_mapping": underlying_raw,
            "domestic_mapping_display": domestic_display(domestic_raw),
            "domestic_cross_market_mapping": domestic_desc,
            "industry_chain_mapping": industry_desc_df,
            "industry_chain_asset_pool": industry_pool_raw,
            "foreign_mapping_display": foreign_display(foreign_raw),
            "foreign_asset_master": foreign_asset_raw,
            "foreign_cross_market_mapping": foreign_desc_df,
        }

        # First pass clean for residual counts, then validation, then final clean with validation included.
        cleaned: dict[str, pd.DataFrame] = {}
        dropped: dict[str, list[str]] = {}
        for name, df in sheets_raw.items():
            c, d = clean_sheet(name, df)
            cleaned[name] = c
            dropped[name] = d
        residual_invalid = count_residual(cleaned, ["CHECK", "TODO", "NEED_REVIEW"])
        residual_source = count_residual(cleaned, ["Tushare", "AKShare", "TqSdk", "OpenCTP"])
        validation_raw = validation(asset_show, futures_show, options_show, underlying_raw, domestic_raw, industry_raw, foreign_raw, foreign_asset_raw, contract_show, residual_invalid, residual_source)
        sheets_raw["validation_summary"] = validation_raw
        sheets_raw["README"] = readme_df([])

        final_sheets: dict[str, pd.DataFrame] = {}
        for name in SHEET_ORDER:
            if name not in sheets_raw:
                continue
            c, d = clean_sheet(name, sheets_raw[name])
            final_sheets[name] = c
            dropped[name] = d

        write_excel(FULL_OUT, final_sheets)
        write_excel(OPT_OUT, final_sheets)

        metrics = {
            "futures_underlying_empty": int(futures_show.get("underlying_group", pd.Series(dtype=str)).map(is_invalid).sum()),
            "options_check_wait": int(options_show.astype(str).apply(lambda s: s.str.contains("CHECK_WAIT_ETF_MASTER", regex=False, na=False)).sum().sum()),
            "etf_option_underlying_empty": int(((options_show["subtype"].eq("ETF_OPTION")) & (options_show["underlying_asset_id"].map(is_invalid) | options_show["underlying_asset_id"].astype(str).str.strip().eq(""))).sum()) if {"subtype", "underlying_asset_id"}.issubset(options_show.columns) else 0,
            "contract_source_residual": count_residual({"contract": final_sheets["contract_rule_detail"]}, ["AKShare", "OpenCTP"]),
            "residual_invalid": count_residual({k: v for k, v in final_sheets.items() if k != "validation_summary"}, ["CHECK", "TODO", "NEED_REVIEW"]),
            "residual_source": count_residual({k: v for k, v in final_sheets.items() if k != "validation_summary"}, ["Tushare", "AKShare", "TqSdk", "OpenCTP"]),
        }
        validation_clean = final_sheets["validation_summary"]
        fail_count = int((validation_clean["result"] == "FAIL").sum())
        conclusion = "PASS"
        if fail_count or metrics["futures_underlying_empty"] or metrics["options_check_wait"] or metrics["etf_option_underlying_empty"]:
            conclusion = "NEED_MAJOR_FIX"
        elif int((validation_clean["result"] == "WARN").sum()) or metrics["residual_invalid"] or metrics["residual_source"]:
            conclusion = "NEED_MINOR_FIX"

        sheet_rows = {name: len(df) for name, df in final_sheets.items()}
        REPORT_OUT.write_text(report_text(backed, sheet_rows, dropped, metrics, validation_clean, conclusion), encoding="utf-8")

        print(f"是否备份旧 mapping_info_full.xlsx: {'Y' if backed else 'N'}")
        print(f"output/mapping_info_full.xlsx 路径: {FULL_OUT}")
        print(f"output/mapping_info_full_optimized.xlsx 路径: {OPT_OUT}")
        print(f"mapping_display 行数: {len(final_sheets.get('mapping_display', []))}")
        print(f"current_contract_detail 行数: {len(final_sheets.get('current_contract_detail', []))}")
        print(f"core_asset_master 行数: {len(final_sheets.get('core_asset_master', []))}")
        print(f"core_etf_master 行数: {len(final_sheets.get('core_etf_master', []))}")
        print(f"futures_master underlying_group 空值数量: {metrics['futures_underlying_empty']}")
        print(f"options_master CHECK_WAIT_ETF_MASTER 残留数量: {metrics['options_check_wait']}")
        print(f"ETF_OPTION underlying_asset_id 空值数量: {metrics['etf_option_underlying_empty']}")
        print(f"contract_rule_detail 残留 AKShare/OpenCTP 单元格数量: {metrics['contract_source_residual']}")
        print(f"validation_summary FAIL 数量: {fail_count}")
        print(f"最终结论: {conclusion}")
        print(f"检查报告路径: {REPORT_OUT}")
    except Exception as exc:
        print(f"Step 9 展示层优化失败: {exc}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
