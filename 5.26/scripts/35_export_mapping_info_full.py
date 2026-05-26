from __future__ import annotations

import re
import sys
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Font


ROOT = Path(__file__).resolve().parents[1]
ENCODING = "utf-8-sig"
EMPTY = "空"
EXCEL_OUT = ROOT / "output" / "mapping_info_full.xlsx"
REPORT_OUT = ROOT / "output" / "mapping_info_full_check_report.txt"

ASSET_MASTER_PRIMARY = ROOT / "data" / "stage5_mapping" / "final" / "asset_master_for_mapping_patched.csv"
ASSET_MASTER_FALLBACK = ROOT / "data" / "stage4_etf_index" / "final" / "asset_master_with_etf_index_cleaned.csv"

INPUTS = {
    "futures_master": ROOT / "data" / "final" / "futures_master.csv",
    "options_master": ROOT / "data" / "stage2_options" / "final" / "options_master_tushare_primary_final.csv",
    "etf_master": ROOT / "data" / "stage4_etf_index" / "final" / "etf_master_cleaned.csv",
    "index_master": ROOT / "data" / "stage4_etf_index" / "final" / "index_master.csv",
    "contract_rule_detail_future": ROOT / "data" / "final" / "contract_rule_detail.csv",
    "contract_rule_detail_option": ROOT / "data" / "stage2_options" / "final" / "option_contract_rule_detail_tushare_primary_final.csv",
    "underlying_mapping": ROOT / "data" / "stage5_mapping" / "final" / "underlying_mapping_fixed.csv",
    "domestic_cross_market_mapping": ROOT / "data" / "stage5_mapping" / "final" / "domestic_cross_market_mapping_fixed.csv",
    "industry_chain_mapping": ROOT / "data" / "stage6_industry_chain" / "final" / "industry_chain_mapping.csv",
    "industry_chain_asset_pool": ROOT / "data" / "stage6_industry_chain" / "final" / "industry_chain_asset_pool.csv",
    "foreign_asset_master": ROOT / "data" / "stage7_foreign_mapping" / "final" / "foreign_asset_master.csv",
    "foreign_cross_market_mapping": ROOT / "data" / "stage7_foreign_mapping" / "final" / "foreign_cross_market_mapping.csv",
    "industry_chain_quality_report": ROOT / "data" / "stage6_industry_chain" / "processed" / "industry_chain_quality_report.csv",
    "foreign_mapping_quality_report": ROOT / "data" / "stage7_foreign_mapping" / "processed" / "foreign_mapping_quality_report.csv",
}

INVALID_EXACT = {
    "", "NAN", "NONE", "NULL", "CHECK", "TODO", "PENDING", "NEED_REVIEW", "REVIEW",
    "CHECK_NEED_EXCHANGE_RULE_REVIEW", "CHECK_DYNAMIC_BY_EXCHANGE_NOTICE",
    "OBSERVED_MATURITY_DATES_AVAILABLE", "OBSERVED_EXPIRY_MONTHS_AVAILABLE",
    "OBSERVED_DELIST_OR_LAST_TRADE_DATES_AVAILABLE", "OBSERVED_LAST_EXERCISE_DATES_AVAILABLE",
    "NA_NOT_APPLICABLE", "NA_NOT_APPLICABLE_OR_CHECK_BY_EXCHANGE", "UNAVAILABLE",
    "PLACEHOLDER", "WAIT", "MIXED_NEED_REVIEW", "待复核", "待确认", "需复核",
    "需要复核", "待补充", "未确认", "无法确定",
}
INVALID_CONTAINS = ["NEED_REVIEW", "待复核", "待确认", "需复核", "需要复核", "待补充", "未确认", "无法确定", "动态字段", "交易所公告"]
DISPLAY_TOKEN_RE = re.compile(r"\b(?:CHECK|TODO|NEED_REVIEW)\b", re.IGNORECASE)

PROCESS_KEYWORDS = [
    "data_source", "source_status", "source", "status", "notes", "note", "review",
    "quality", "log", "summary", "confidence", "method", "update_date",
    "classification", "required", "reason", "suggested", "raw", "patch", "warning",
]

CORE_FIELDS = {
    "asset_id", "foreign_asset_id", "mapping_id", "chain_id", "symbol", "name_cn", "name_en",
    "asset_type", "subtype", "exchange_code", "exchange_name", "country", "currency",
    "underlying_group", "relation_type", "strategy_type", "direction_supported", "tradable",
    "can_long", "can_short", "tick_size", "contract_multiplier", "contract_unit", "quote_unit",
}

ORDER = {
    "asset_master": ["asset_id", "asset_type", "subtype", "symbol", "name_cn", "name_en", "exchange_code", "exchange_name", "country", "currency", "underlying_group", "sector", "tradable", "can_long", "can_short"],
    "futures_master": ["asset_id", "symbol", "name_cn", "name_en", "exchange_code", "exchange_name", "country", "currency", "future_type", "sector", "underlying_group", "tradable", "can_long", "can_short", "contract_unit", "quote_unit", "tick_size", "contract_multiplier", "trading_hours", "night_trading"],
    "options_master": ["asset_id", "asset_type", "subtype", "option_symbol", "option_name_cn", "option_name_en", "option_type", "exchange_code", "exchange_name", "country", "currency", "underlying_asset_id", "underlying_symbol", "underlying_name", "underlying_group", "call_put_supported", "exercise_type", "settlement_type", "contract_multiplier", "tick_size", "quote_unit", "contract_unit", "tradable", "can_long", "can_short"],
    "etf_master": ["asset_id", "asset_type", "subtype", "etf_code", "ts_code", "name_cn", "name_en", "exchange_code", "exchange_name", "country", "currency", "fund_company", "fund_type", "invest_type", "tracking_index_code", "tracking_index_name", "underlying_group", "tradable", "can_long", "can_short"],
    "index_master": ["asset_id", "asset_type", "subtype", "index_code", "ts_code", "name_cn", "name_en", "exchange_code", "exchange_name", "publisher", "category", "market", "country", "currency", "underlying_group", "tradable", "can_long", "can_short"],
    "contract_rule_detail": ["asset_id", "symbol", "name_cn", "asset_type", "subtype", "exchange_code", "exchange_name", "country", "currency", "contract_month_rule", "contract_rule_code", "contract_rule_desc_cn", "contract_rule_params", "listed_contract_count_rule", "expiry_rule", "expiry_rule_code", "last_trading_day_rule", "delivery_day_rule", "exercise_rule", "exercise_type", "settlement_type", "contract_multiplier", "contract_unit", "tick_size", "quote_unit", "price_limit_rule", "margin_rule", "trading_hours", "night_trading", "position_limit_rule", "delivery_rule", "rule_parseable"],
    "underlying_mapping": ["asset_id", "symbol", "name_cn", "asset_type", "subtype", "exchange_code", "underlying_group", "underlying_name_cn", "underlying_name_en"],
    "domestic_cross_market_mapping": ["mapping_id", "asset_id_a", "asset_id_b", "symbol_a", "symbol_b", "name_a", "name_b", "asset_type_a", "asset_type_b", "exchange_a", "exchange_b", "underlying_group", "relation_type", "strategy_type", "market_relation", "tradable_check", "direction_supported", "long_short_note"],
    "industry_chain_mapping": ["chain_id", "industry_name", "chain_segment", "asset_id_a", "asset_id_b", "symbol_a", "symbol_b", "name_a", "name_b", "asset_type_a", "asset_type_b", "exchange_a", "exchange_b", "underlying_group_a", "underlying_group_b", "relation_direction", "relation_type", "relation_strength", "strategy_type", "tradable_check", "direction_supported", "long_short_note"],
    "industry_chain_asset_pool": ["underlying_group", "underlying_name_cn", "industry_name", "asset_count", "preferred_asset_id", "preferred_symbol", "preferred_name_cn", "preferred_asset_type", "all_asset_ids", "mapping_status"],
    "foreign_asset_master": ["foreign_asset_id", "foreign_asset_type", "foreign_subtype", "foreign_symbol", "foreign_name_en", "foreign_name_cn", "foreign_exchange_code", "foreign_exchange_name", "foreign_country", "foreign_currency", "foreign_contract_unit", "foreign_quote_unit", "foreign_tick_size", "underlying_group", "sector", "tradable", "can_long", "can_short"],
    "foreign_cross_market_mapping": ["mapping_id", "domestic_asset_id", "foreign_asset_id", "domestic_symbol", "foreign_symbol", "domestic_name", "foreign_name", "domestic_asset_type", "foreign_asset_type", "domestic_exchange", "foreign_exchange", "domestic_country", "foreign_country", "underlying_group", "currency_domestic", "currency_foreign", "unit_domestic", "unit_foreign", "unit_conversion", "fx_conversion_needed", "trading_hour_overlap", "relation_type", "strategy_type", "direction_supported", "mapping_confidence"],
}


def read_csv(path: Path, required: bool = False) -> pd.DataFrame:
    if not path.exists():
        if required:
            raise FileNotFoundError(f"缺少输入文件: {path}")
        return pd.DataFrame()
    return pd.read_csv(path, dtype=str, encoding=ENCODING).fillna("")


def choose_asset_master() -> tuple[pd.DataFrame, Path]:
    if ASSET_MASTER_PRIMARY.exists():
        return read_csv(ASSET_MASTER_PRIMARY, required=True), ASSET_MASTER_PRIMARY
    return read_csv(ASSET_MASTER_FALLBACK, required=True), ASSET_MASTER_FALLBACK


def clean_value(value: Any) -> str:
    text = "" if value is None or pd.isna(value) else str(value).strip()
    upper = text.upper()
    if upper in INVALID_EXACT:
        return EMPTY
    if any(token in upper for token in ["CHECK_NEED", "CHECK_DYNAMIC", "NEED_REVIEW"]):
        return EMPTY
    if any(token in text for token in INVALID_CONTAINS):
        return EMPTY
    return text if text else EMPTY


def is_empty_display(series: pd.Series) -> pd.Series:
    return series.fillna("").astype(str).str.strip().isin(["", EMPTY])


def core_column(col: str) -> bool:
    lower = col.lower()
    return lower in CORE_FIELDS or any(lower.endswith("_" + c) for c in CORE_FIELDS) or any(lower.startswith(c + "_") for c in CORE_FIELDS)


def should_keep_process_col(sheet: str, col: str) -> bool:
    if sheet == "relation_summary" and col == "relation_source":
        return True
    if sheet == "industry_chain_asset_pool" and col == "mapping_status":
        return True
    if sheet in {"domestic_cross_market_mapping", "industry_chain_mapping"} and col == "long_short_note":
        return True
    if sheet in {"relation_summary", "validation_summary"} and "summary" in col.lower():
        return True
    if sheet == "foreign_cross_market_mapping" and col == "mapping_confidence":
        return True
    if sheet == "industry_chain_mapping" and col in {"mapping_confidence", "relation_strength"}:
        return True
    if sheet == "domestic_cross_market_mapping" and col in {"relation_type", "strategy_type", "direction_supported"}:
        return True
    return False


def process_column_drop(sheet: str, col: str) -> bool:
    lower = col.lower()
    return any(k in lower for k in PROCESS_KEYWORDS) and not should_keep_process_col(sheet, col)


def order_columns(sheet: str, df: pd.DataFrame) -> pd.DataFrame:
    preferred = [c for c in ORDER.get(sheet, []) if c in df.columns]
    rest = [c for c in df.columns if c not in preferred]
    return df[preferred + rest]


def clean_sheet(sheet: str, df: pd.DataFrame) -> tuple[pd.DataFrame, list[str], dict[str, int]]:
    if df.empty:
        return df, [], {"invalid_token_cells": 0, "empty_columns": 0}
    cleaned = df.copy()
    for col in cleaned.columns:
        cleaned[col] = cleaned[col].map(clean_value)

    dropped: list[str] = []
    for col in list(cleaned.columns):
        if process_column_drop(sheet, col):
            dropped.append(col)
            cleaned = cleaned.drop(columns=[col])

    for col in list(cleaned.columns):
        empty_ratio = float(is_empty_display(cleaned[col]).mean()) if len(cleaned) else 1.0
        if empty_ratio >= 1.0:
            dropped.append(col)
            cleaned = cleaned.drop(columns=[col])
        elif empty_ratio > 0.95 and not core_column(col):
            dropped.append(col)
            cleaned = cleaned.drop(columns=[col])

    if cleaned.empty:
        return cleaned, dropped, {"invalid_token_cells": 0, "empty_columns": 0}
    cleaned = order_columns(sheet, cleaned)
    invalid_cells = int(cleaned.astype(str).apply(lambda s: s.str.contains(DISPLAY_TOKEN_RE, na=False)).sum().sum())
    empty_columns = int(sum(is_empty_display(cleaned[col]).all() for col in cleaned.columns))
    return cleaned, dropped, {"invalid_token_cells": invalid_cells, "empty_columns": empty_columns}


def concat_contract_rules() -> pd.DataFrame:
    future = read_csv(INPUTS["contract_rule_detail_future"])
    option = read_csv(INPUTS["contract_rule_detail_option"])
    frames = [df for df in [future, option] if not df.empty]
    if not frames:
        return pd.DataFrame()
    columns = sorted(set().union(*[set(df.columns) for df in frames]))
    return pd.concat([df.reindex(columns=columns, fill_value="") for df in frames], ignore_index=True)


def build_relation_summary(domestic: pd.DataFrame, industry: pd.DataFrame, foreign: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for source, df in [("DOMESTIC", domestic), ("INDUSTRY_CHAIN", industry), ("FOREIGN", foreign)]:
        if df.empty or "relation_type" not in df.columns:
            continue
        cols = ["relation_type"] + (["strategy_type"] if "strategy_type" in df.columns else [])
        grouped = df.groupby(cols, dropna=False).size().reset_index(name="count")
        for _, row in grouped.iterrows():
            rows.append({
                "relation_source": source,
                "relation_type": row.get("relation_type", ""),
                "strategy_type": row.get("strategy_type", ""),
                "count": row["count"],
            })
    return pd.DataFrame(rows)


def validation_row(item: str, result: str, count: int, description: str) -> dict[str, Any]:
    return {"check_item": item, "result": result, "count": count, "description": description}


def build_validation(asset: pd.DataFrame, underlying: pd.DataFrame, domestic: pd.DataFrame, industry: pd.DataFrame, foreign_map: pd.DataFrame, foreign_asset: pd.DataFrame, contract: pd.DataFrame) -> pd.DataFrame:
    asset_ids = set(asset.get("asset_id", pd.Series(dtype=str)).astype(str))
    foreign_ids = set(foreign_asset.get("foreign_asset_id", pd.Series(dtype=str)).astype(str))
    rows = [
        validation_row("asset_master 行数", "PASS" if len(asset) else "FAIL", len(asset), "资产主表记录数。"),
        validation_row("asset_id 是否唯一", "PASS" if "asset_id" in asset and not asset["asset_id"].duplicated().any() else "FAIL", int(asset["asset_id"].duplicated().sum()) if "asset_id" in asset else -1, "count 为重复 asset_id 数。"),
        validation_row("underlying_mapping 行数", "PASS" if len(underlying) else "FAIL", len(underlying), "统一标的映射记录数。"),
        validation_row("domestic_cross_market_mapping 行数", "PASS" if len(domestic) else "WARN", len(domestic), "国内跨市场映射记录数。"),
        validation_row("domestic mapping_id 是否唯一", "PASS" if domestic.empty or not domestic["mapping_id"].duplicated().any() else "FAIL", int(domestic["mapping_id"].duplicated().sum()) if "mapping_id" in domestic else 0, "count 为重复 mapping_id 数。"),
        validation_row("domestic mapping 是否存在自映射", "PASS" if domestic.empty or int((domestic["asset_id_a"] == domestic["asset_id_b"]).sum()) == 0 else "FAIL", int((domestic["asset_id_a"] == domestic["asset_id_b"]).sum()) if {"asset_id_a", "asset_id_b"}.issubset(domestic.columns) else 0, "count 为自映射数量。"),
        validation_row("industry_chain_mapping 行数", "PASS" if len(industry) else "WARN", len(industry), "产业链映射记录数。"),
        validation_row("industry_chain 是否存在自映射", "PASS" if industry.empty or int((industry["asset_id_a"] == industry["asset_id_b"]).sum()) == 0 else "FAIL", int((industry["asset_id_a"] == industry["asset_id_b"]).sum()) if {"asset_id_a", "asset_id_b"}.issubset(industry.columns) else 0, "count 为自映射数量。"),
        validation_row("foreign_cross_market_mapping 行数", "PASS" if len(foreign_map) else "WARN", len(foreign_map), "国外跨市场映射记录数。"),
        validation_row("foreign mapping_id 是否唯一", "PASS" if foreign_map.empty or not foreign_map["mapping_id"].duplicated().any() else "FAIL", int(foreign_map["mapping_id"].duplicated().sum()) if "mapping_id" in foreign_map else 0, "count 为重复 mapping_id 数。"),
        validation_row("foreign mapping 国内资产是否能回连 asset_master", "PASS" if foreign_map.empty or not (set(foreign_map["domestic_asset_id"]) - asset_ids) else "FAIL", len(set(foreign_map["domestic_asset_id"]) - asset_ids) if "domestic_asset_id" in foreign_map else 0, "count 为无法回连国内 asset_master 的记录数。"),
        validation_row("foreign mapping 国外资产是否能回连 foreign_asset_master", "PASS" if foreign_map.empty or not (set(foreign_map["foreign_asset_id"]) - foreign_ids) else "FAIL", len(set(foreign_map["foreign_asset_id"]) - foreign_ids) if "foreign_asset_id" in foreign_map else 0, "count 为无法回连 foreign_asset_master 的记录数。"),
        validation_row("contract_rule_detail 行数", "PASS" if len(contract) else "WARN", len(contract), "期货和期权规则详情合并后记录数。"),
    ]
    return pd.DataFrame(rows)


def readme_df(missing_inputs: list[str]) -> pd.DataFrame:
    rows = [
        ("项目名称", "套利组合构建用资产映射表完整版"),
        ("文件说明", "本文件为完整开发版，包含资产主表、期货、期权、ETF、指数、统一标的映射、国内跨市场映射、产业链映射、国外跨市场映射和交易规则详情表。"),
        ("注意1", "本文件是完整版，供后续开发和检查使用。"),
        ("注意2", "提交用精简版仍为 output/mapping_info.xlsx。"),
        ("注意3", "本文件中无有效数据的列已自动删除。"),
        ("注意4", "暂未确认或无法识别的内容统一显示为空。"),
    ]
    if missing_inputs:
        rows.append(("缺失输入", "；".join(missing_inputs)))
    return pd.DataFrame(rows, columns=["item", "description"])


def write_excel(sheets: dict[str, pd.DataFrame]) -> None:
    EXCEL_OUT.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(EXCEL_OUT, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name[:31], index=False)
            ws = writer.book[name[:31]]
            ws.freeze_panes = "A2"
            if ws.max_row > 1 and ws.max_column > 1:
                ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.font = Font(bold=True)
            for col_cells in ws.columns:
                width = 10
                letter = col_cells[0].column_letter
                for cell in col_cells[:200]:
                    width = max(width, min(len("" if cell.value is None else str(cell.value)) + 2, 40))
                ws.column_dimensions[letter].width = width


def build_report(sheet_rows: dict[str, int], dropped: dict[str, list[str]], stats: dict[str, dict[str, int]], validation: pd.DataFrame, missing_inputs: list[str]) -> str:
    lines = [
        f"输出 Excel 路径: {EXCEL_OUT}",
        "",
        "缺失输入:",
        *(["- 无"] if not missing_inputs else [f"- {x}" for x in missing_inputs]),
        "",
        "每个 sheet 行数 / 保留列数 / 删除列:",
    ]
    for sheet, rows in sheet_rows.items():
        lines.append(f"- {sheet}: 行数={rows}, 保留列数={stats[sheet]['kept_columns']}, 删除列数={len(dropped.get(sheet, []))}")
        lines.append(f"  删除列名: {', '.join(dropped.get(sheet, [])) if dropped.get(sheet) else '无'}")
        lines.append(f"  完全空列数量: {stats[sheet]['empty_columns']}")
        lines.append(f"  CHECK/TODO/NEED_REVIEW 展示残留单元格数量: {stats[sheet]['invalid_token_cells']}")
    lines.append("")
    lines.append("validation_summary 结果:")
    for _, row in validation.iterrows():
        lines.append(f"- {row['result']} | {row['check_item']} | count={row['count']} | {row['description']}")
    return "\n".join(lines) + "\n"


def main() -> None:
    try:
        asset, asset_path = choose_asset_master()
        raw = {
            "asset_master": asset,
            "futures_master": read_csv(INPUTS["futures_master"]),
            "options_master": read_csv(INPUTS["options_master"]),
            "etf_master": read_csv(INPUTS["etf_master"]),
            "index_master": read_csv(INPUTS["index_master"]),
            "contract_rule_detail": concat_contract_rules(),
            "underlying_mapping": read_csv(INPUTS["underlying_mapping"]),
            "domestic_cross_market_mapping": read_csv(INPUTS["domestic_cross_market_mapping"]),
            "industry_chain_mapping": read_csv(INPUTS["industry_chain_mapping"]),
            "industry_chain_asset_pool": read_csv(INPUTS["industry_chain_asset_pool"]),
            "foreign_asset_master": read_csv(INPUTS["foreign_asset_master"]),
            "foreign_cross_market_mapping": read_csv(INPUTS["foreign_cross_market_mapping"]),
        }
        missing_inputs = [f"asset_master fallback used: {asset_path}"] if asset_path != ASSET_MASTER_PRIMARY else []
        for key, path in INPUTS.items():
            if key.startswith("contract_rule_detail_") or key.endswith("quality_report"):
                continue
            if not path.exists():
                missing_inputs.append(str(path))

        relation_summary = build_relation_summary(raw["domestic_cross_market_mapping"], raw["industry_chain_mapping"], raw["foreign_cross_market_mapping"])
        validation = build_validation(raw["asset_master"], raw["underlying_mapping"], raw["domestic_cross_market_mapping"], raw["industry_chain_mapping"], raw["foreign_cross_market_mapping"], raw["foreign_asset_master"], raw["contract_rule_detail"])
        raw["relation_summary"] = relation_summary
        raw["validation_summary"] = validation

        sheets: dict[str, pd.DataFrame] = {"README": readme_df(missing_inputs)}
        dropped: dict[str, list[str]] = {}
        stats: dict[str, dict[str, int]] = {}
        for sheet in [
            "asset_master", "futures_master", "options_master", "etf_master", "index_master",
            "contract_rule_detail", "underlying_mapping", "domestic_cross_market_mapping",
            "industry_chain_mapping", "industry_chain_asset_pool", "foreign_asset_master",
            "foreign_cross_market_mapping", "relation_summary", "validation_summary",
        ]:
            if raw[sheet].empty and sheet not in {"relation_summary", "validation_summary"}:
                missing_inputs.append(f"{sheet}: empty or missing")
                continue
            clean, drop_cols, sheet_stats = clean_sheet(sheet, raw[sheet])
            sheets[sheet] = clean
            dropped[sheet] = drop_cols
            sheet_stats["kept_columns"] = len(clean.columns)
            stats[sheet] = sheet_stats

        clean_readme, drop_cols, sheet_stats = clean_sheet("README", sheets["README"])
        sheets["README"] = clean_readme
        dropped["README"] = drop_cols
        sheet_stats["kept_columns"] = len(clean_readme.columns)
        stats["README"] = sheet_stats

        ordered_sheets = {name: sheets[name] for name in ["README"] + [s for s in sheets if s != "README"]}
        write_excel(ordered_sheets)

        sheet_rows = {name: len(df) for name, df in ordered_sheets.items()}
        report = build_report(sheet_rows, dropped, stats, sheets["validation_summary"], missing_inputs)
        REPORT_OUT.write_text(report, encoding="utf-8")

        fail_count = int((sheets["validation_summary"]["result"] == "FAIL").sum())
        warn_count = int((sheets["validation_summary"]["result"] == "WARN").sum())
        print(f"output/mapping_info_full.xlsx 路径: {EXCEL_OUT}")
        for name, rows in sheet_rows.items():
            print(f"{name}: 行数={rows}, 删除列数量={len(dropped.get(name, []))}")
        print(f"validation_summary FAIL 数量: {fail_count}")
        print(f"validation_summary WARN 数量: {warn_count}")
        print(f"检查报告路径: {REPORT_OUT}")
    except Exception as exc:
        print(f"mapping_info_full 导出失败: {exc}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
