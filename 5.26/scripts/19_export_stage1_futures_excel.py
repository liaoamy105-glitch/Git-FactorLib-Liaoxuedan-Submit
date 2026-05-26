from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font


PROJECT_ROOT = Path(__file__).resolve().parents[1]
OUTPUT_FILE = PROJECT_ROOT / "output" / "mapping_info_stage1_futures.xlsx"

REQUIRED_FILES = {
    "asset_master": PROJECT_ROOT / "data" / "processed" / "asset_master.csv",
    "futures_master": PROJECT_ROOT / "data" / "processed" / "futures_master.csv",
    "contract_rule_detail": PROJECT_ROOT / "data" / "processed" / "contract_rule_detail_futures_cleaned.csv",
    "remaining_review_list": PROJECT_ROOT / "data" / "processed" / "contract_rule_remaining_review_list.csv",
    "quality_summary": PROJECT_ROOT / "output" / "contract_rule_detail_clean_quality_summary.csv",
}

OPTIONAL_FILES = {
    "fill_log_akshare": PROJECT_ROOT / "data" / "processed" / "akshare_contract_rule_fill_log.csv",
    "infer_log_month_rule": PROJECT_ROOT / "data" / "processed" / "contract_month_rule_infer_log.csv",
    "applied_log_shfe_ine": PROJECT_ROOT / "data" / "processed" / "shfe_ine_rules_applied_log.csv",
}


def read_csv(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"未找到输入文件：{path}")
    return pd.read_csv(path, dtype=str, keep_default_na=False)


def build_readme() -> pd.DataFrame:
    rows = [
        ("项目名称", "套利组合构建用资产映射表与期货交易规则详情表"),
        ("当前阶段", "Stage 1 - 国内期货资产库与期货规则详情表"),
        ("完成内容 1", "国内六大期货交易所期货品种已覆盖。"),
        ("完成内容 2", "已生成 asset_master 和 futures_master。"),
        ("完成内容 3", "已生成 contract_rule_detail。"),
        ("完成内容 4", "CFFEX 8 个品种已按官网规则补充。"),
        ("完成内容 5", "SHFE/INE 22 个品种已从官网本地 HTML 解析。"),
        ("完成内容 6", "其余品种已通过 Tushare observed_delivery_months 推断合约月份规则，并通过 AKShare/OpenCTP 补充交易参数。"),
        ("完成内容 7", "特殊/历史/非标准品种已单独标记。"),
        ("完成内容 8", "仍需复核字段已集中到 remaining_review_list。"),
        ("注意事项 1", "HTML_PARSED_NEED_REVIEW 和 INFERRED_NEED_REVIEW 不是最终官方确认。"),
        ("注意事项 2", "动态字段如保证金、涨跌停、持仓限额不建议写死。"),
        ("注意事项 3", "remaining_review_list 是后续复核清单，不是失败清单。"),
        ("注意事项 4", "下一阶段将扩展期权、ETF、现货与跨市场映射。"),
    ]
    return pd.DataFrame(rows, columns=["item", "description"])


def build_source_status_dict() -> pd.DataFrame:
    rows = [
        ("OFFICIAL_COMPLETE", "最终规则状态：官方规则已确认且核心静态规则可解析。"),
        ("HTML_PARSED_NEED_REVIEW", "最终规则状态：来自官网本地 HTML 解析，规则可解析但仍需人工抽查。"),
        ("INFERRED_NEED_REVIEW", "最终规则状态：合约月份等来自 Tushare 上市合约观察值推断，需复核。"),
        ("SPECIAL_OR_INACTIVE_NEED_REVIEW", "最终规则状态：特殊、历史或非标准品种，需确认是否纳入有效套利资产池。"),
        ("OFFICIAL_REVIEWED", "来源状态：已按交易所官网规则人工确认。"),
        ("LOCAL_HTML_PARSED_NEED_REVIEW", "来源状态：通过浏览器保存官网 HTML 后本地解析，需人工复核。"),
        ("INFERRED_FROM_TUSHARE_LISTED_CONTRACTS_NEED_REVIEW", "来源状态：根据 Tushare observed_delivery_months 推断，需复核。"),
        ("AKSHARE_OPENCTP_NEED_REVIEW", "来源状态：根据 AKShare/OpenCTP 自动补充交易参数，需复核。"),
        ("MIXED", "来源状态：原始/自动信息混合，仍有关键规则未确认。"),
    ]
    return pd.DataFrame(rows, columns=["status", "meaning"])


def build_data_source_coverage() -> pd.DataFrame:
    rows = [
        ("Tushare", "用于期货基础信息、合约样本、observed_delivery_months、trading_hours。"),
        ("AKShare/OpenCTP", "用于 tick_size、contract_multiplier、margin_rule 等交易参数补充。"),
        ("CFFEX Official Website", "用于 CFFEX 8 个金融期货官方规则确认。"),
        ("SHFE/INE Local HTML", "用于 SHFE/INE 22 个期货品种本地 HTML 解析规则。"),
        ("DCE/CZCE/GFEX", "官网访问受限或尚未完成，当前通过 Tushare + AKShare/OpenCTP + 推断规则形成待复核版本。"),
    ]
    return pd.DataFrame(rows, columns=["data_source", "coverage"])


def collect_sheets() -> dict[str, pd.DataFrame]:
    sheets = {
        "README": build_readme(),
        "asset_master": read_csv(REQUIRED_FILES["asset_master"]),
        "futures_master": read_csv(REQUIRED_FILES["futures_master"]),
        "contract_rule_detail": read_csv(REQUIRED_FILES["contract_rule_detail"]),
        "remaining_review_list": read_csv(REQUIRED_FILES["remaining_review_list"]),
        "quality_summary": read_csv(REQUIRED_FILES["quality_summary"]),
        "source_status_dict": build_source_status_dict(),
        "data_source_coverage": build_data_source_coverage(),
    }

    for sheet_name, path in OPTIONAL_FILES.items():
        if path.exists():
            sheets[sheet_name] = read_csv(path)
        else:
            print(f"可选日志不存在，已跳过: {path}")
    return sheets


def write_excel(sheets: dict[str, pd.DataFrame]) -> None:
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)


def format_excel(sheets: dict[str, pd.DataFrame]) -> None:
    wb = load_workbook(OUTPUT_FILE)
    for sheet_name, df in sheets.items():
        ws = wb[sheet_name]
        if ws.max_row > 1:
            ws.freeze_panes = "A2"
        else:
            ws.freeze_panes = "A2"

        if ws.max_row >= 1 and ws.max_column >= 1:
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.font = Font(bold=True)

        for column_cells in ws.columns:
            header = column_cells[0].value
            max_length = len(str(header)) if header is not None else 0
            for cell in column_cells[1:]:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            width = min(max(max_length + 2, 10), 40)
            ws.column_dimensions[column_cells[0].column_letter].width = width

        if sheet_name in {"README", "source_status_dict", "data_source_coverage"}:
            for col in range(1, ws.max_column + 1):
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 40

    wb.save(OUTPUT_FILE)


def main() -> None:
    sheets = collect_sheets()
    write_excel(sheets)
    format_excel(sheets)

    print(f"输出 Excel 路径: {OUTPUT_FILE}")
    print("每个 sheet 的行数:")
    for sheet_name, df in sheets.items():
        print(f"  {sheet_name}: {len(df)}")
    print(f"contract_rule_detail 总行数: {len(sheets['contract_rule_detail'])}")
    print(f"remaining_review_list 行数: {len(sheets['remaining_review_list'])}")
    print(f"quality_summary 是否成功加入: {'quality_summary' in sheets}")


if __name__ == "__main__":
    main()
