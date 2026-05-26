from __future__ import annotations

import sys
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Font


ROOT = Path(__file__).resolve().parents[1]
ENCODING = "utf-8-sig"

ASSET_MASTER_PRIMARY = ROOT / "data" / "stage5_mapping" / "final" / "asset_master_for_mapping_patched.csv"
ASSET_MASTER_FALLBACK = ROOT / "data" / "stage4_etf_index" / "final" / "asset_master_with_etf_index_cleaned.csv"
UNDERLYING_MAPPING = ROOT / "data" / "stage5_mapping" / "final" / "underlying_mapping_fixed.csv"
DOMESTIC_MAPPING = ROOT / "data" / "stage5_mapping" / "final" / "domestic_cross_market_mapping_fixed.csv"
INDUSTRY_MAPPING = ROOT / "data" / "stage6_industry_chain" / "final" / "industry_chain_mapping.csv"

STAGE7_DIR = ROOT / "data" / "stage7_foreign_mapping"
FINAL_DIR = STAGE7_DIR / "final"
PROCESSED_DIR = STAGE7_DIR / "processed"
MANUAL_DIR = STAGE7_DIR / "manual"
OUTPUT_DIR = ROOT / "output" / "stage7_foreign_mapping"

FOREIGN_ASSET_OUT = FINAL_DIR / "foreign_asset_master.csv"
FOREIGN_MAPPING_OUT = FINAL_DIR / "foreign_cross_market_mapping.csv"
QUALITY_REPORT_OUT = PROCESSED_DIR / "foreign_mapping_quality_report.csv"
MISSING_DOMESTIC_OUT = PROCESSED_DIR / "foreign_mapping_missing_domestic_review.csv"
LOW_CONFIDENCE_OUT = PROCESSED_DIR / "foreign_mapping_low_confidence_review.csv"
EXCEL_OUT = ROOT / "output" / "mapping_info_stage7_foreign_mapping.xlsx"

MISSING_VALUES = {"", "CHECK", "TODO", "NA", "NAN", "NONE", "CHECK_WAIT_ETF_MASTER"}


def fa(
    foreign_asset_id: str,
    asset_type: str,
    subtype: str,
    symbol: str,
    name_en: str,
    name_cn: str,
    exchange_code: str,
    exchange_name: str,
    country: str,
    currency: str,
    underlying_group: str,
    sector: str,
    confidence: str,
    relation_type: str = "DOMESTIC_FOREIGN_SAME_UNDERLYING",
    strategy_type: str = "INTERNATIONAL_ARBITRAGE",
    contract_unit: str = "CHECK",
    quote_unit: str = "CHECK",
    tick_size: str = "CHECK",
) -> dict[str, str]:
    tradable = "N" if asset_type == "INDEX" else "Y"
    return {
        "foreign_asset_id": foreign_asset_id,
        "foreign_asset_type": asset_type,
        "foreign_subtype": subtype,
        "foreign_symbol": symbol,
        "foreign_name_en": name_en,
        "foreign_name_cn": name_cn,
        "foreign_exchange_code": exchange_code,
        "foreign_exchange_name": exchange_name,
        "foreign_country": country,
        "foreign_currency": currency,
        "foreign_contract_unit": contract_unit,
        "foreign_quote_unit": quote_unit,
        "foreign_tick_size": tick_size,
        "underlying_group": underlying_group,
        "sector": sector,
        "tradable": tradable,
        "can_long": tradable,
        "can_short": tradable,
        "data_source": "STEP7_BUILT_IN_MANUAL_RULES",
        "source_status": "MANUAL_RULE_CANDIDATE_NEED_REVIEW",
        "notes": "该国外资产为内置规则候选，后续可用交易所官网或数据源复核。",
        "_mapping_confidence": confidence,
        "_relation_type": relation_type,
        "_strategy_type": strategy_type,
    }


FOREIGN_RULES = [
    # 贵金属
    fa("FUT_COMEX_GC", "FUTURE", "COMMODITY_FUTURE", "GC", "COMEX Gold Futures", "COMEX黄金期货", "COMEX", "COMEX", "US", "USD", "GOLD", "贵金属", "HIGH"),
    fa("FUT_TOCOM_GOLD", "FUTURE", "COMMODITY_FUTURE", "Gold", "TOCOM Gold Futures", "TOCOM黄金期货", "TOCOM", "TOCOM", "JP", "JPY", "GOLD", "贵金属", "MEDIUM"),
    fa("ETF_US_GLD", "ETF", "COMMODITY_ETF", "GLD", "SPDR Gold Shares", "SPDR黄金ETF", "NYSEARCA", "NYSE Arca", "US", "USD", "GOLD", "贵金属", "MEDIUM", "DOMESTIC_FOREIGN_ETF_RELATED", "CROSS_MARKET_ARBITRAGE"),
    fa("FUT_COMEX_SI", "FUTURE", "COMMODITY_FUTURE", "SI", "COMEX Silver Futures", "COMEX白银期货", "COMEX", "COMEX", "US", "USD", "SILVER", "贵金属", "HIGH"),
    fa("ETF_US_SLV", "ETF", "COMMODITY_ETF", "SLV", "iShares Silver Trust", "iShares白银ETF", "NYSEARCA", "NYSE Arca", "US", "USD", "SILVER", "贵金属", "MEDIUM", "DOMESTIC_FOREIGN_ETF_RELATED", "CROSS_MARKET_ARBITRAGE"),
    fa("FUT_TOCOM_SILVER", "FUTURE", "COMMODITY_FUTURE", "Silver", "TOCOM Silver Futures", "TOCOM白银期货", "TOCOM", "TOCOM", "JP", "JPY", "SILVER", "贵金属", "LOW"),
    fa("FUT_NYMEX_PA", "FUTURE", "COMMODITY_FUTURE", "PA", "NYMEX Palladium Futures", "NYMEX钯期货", "NYMEX", "NYMEX", "US", "USD", "PALLADIUM", "贵金属", "HIGH"),
    fa("FUT_NYMEX_PL", "FUTURE", "COMMODITY_FUTURE", "PL", "NYMEX Platinum Futures", "NYMEX铂期货", "NYMEX", "NYMEX", "US", "USD", "PLATINUM", "贵金属", "HIGH"),
    # 有色
    fa("FUT_LME_CA", "FUTURE", "COMMODITY_FUTURE", "CA", "LME Copper", "LME铜", "LME", "London Metal Exchange", "UK", "USD", "COPPER", "有色", "HIGH"),
    fa("FUT_COMEX_HG", "FUTURE", "COMMODITY_FUTURE", "HG", "COMEX Copper", "COMEX铜", "COMEX", "COMEX", "US", "USD", "COPPER", "有色", "HIGH"),
    fa("FUT_LME_AH", "FUTURE", "COMMODITY_FUTURE", "AH", "LME Aluminum", "LME铝", "LME", "London Metal Exchange", "UK", "USD", "ALUMINUM", "有色", "HIGH"),
    fa("FUT_LME_ZS", "FUTURE", "COMMODITY_FUTURE", "ZS", "LME Zinc", "LME锌", "LME", "London Metal Exchange", "UK", "USD", "ZINC", "有色", "HIGH"),
    fa("FUT_LME_PB", "FUTURE", "COMMODITY_FUTURE", "PB", "LME Lead", "LME铅", "LME", "London Metal Exchange", "UK", "USD", "LEAD", "有色", "HIGH"),
    fa("FUT_LME_NI", "FUTURE", "COMMODITY_FUTURE", "NI", "LME Nickel", "LME镍", "LME", "London Metal Exchange", "UK", "USD", "NICKEL", "有色", "HIGH"),
    fa("FUT_LME_SN", "FUTURE", "COMMODITY_FUTURE", "SN", "LME Tin", "LME锡", "LME", "London Metal Exchange", "UK", "USD", "TIN", "有色", "HIGH"),
    fa("FUT_LME_ALUMINA", "FUTURE", "COMMODITY_FUTURE", "ALU", "LME Alumina", "LME氧化铝", "LME", "London Metal Exchange", "UK", "USD", "ALUMINA", "有色", "MEDIUM"),
    fa("FUT_SGX_ALUMINA", "FUTURE", "COMMODITY_FUTURE", "ALM", "SGX Alumina", "SGX氧化铝", "SGX", "Singapore Exchange", "SG", "USD", "ALUMINA", "有色", "MEDIUM"),
    fa("FUT_LME_STEEL_SCRAP", "FUTURE", "COMMODITY_FUTURE", "SCRAP", "LME Steel Scrap", "LME废钢", "LME", "London Metal Exchange", "UK", "USD", "STAINLESS_STEEL", "黑色", "LOW", "DOMESTIC_FOREIGN_RELATED_UNDERLYING", "REFERENCE_ONLY"),
    # 能源
    fa("FUT_NYMEX_CL", "FUTURE", "COMMODITY_FUTURE", "CL", "NYMEX WTI Crude Oil", "NYMEX WTI原油", "NYMEX", "NYMEX", "US", "USD", "CRUDE_OIL", "能源", "HIGH"),
    fa("FUT_ICE_B", "FUTURE", "COMMODITY_FUTURE", "B", "ICE Brent Crude Oil", "ICE布伦特原油", "ICE", "ICE Futures Europe", "UK", "USD", "CRUDE_OIL", "能源", "HIGH"),
    fa("FUT_NYMEX_HO", "FUTURE", "COMMODITY_FUTURE", "HO", "NYMEX Heating Oil", "NYMEX取暖油", "NYMEX", "NYMEX", "US", "USD", "FUEL_OIL", "能源", "MEDIUM", "DOMESTIC_FOREIGN_RELATED_UNDERLYING"),
    fa("FUT_ICE_GASOIL", "FUTURE", "COMMODITY_FUTURE", "G", "ICE Gasoil", "ICE柴油", "ICE", "ICE Futures Europe", "UK", "USD", "FUEL_OIL", "能源", "MEDIUM", "DOMESTIC_FOREIGN_RELATED_UNDERLYING"),
    fa("FUT_SGX_FO380", "FUTURE", "COMMODITY_FUTURE", "FO380", "Singapore Fuel Oil 380cst", "新加坡380燃料油", "SGX", "Singapore Exchange", "SG", "USD", "FUEL_OIL", "能源", "MEDIUM", "DOMESTIC_FOREIGN_RELATED_UNDERLYING"),
    fa("FUT_ICE_LSGO", "FUTURE", "COMMODITY_FUTURE", "LSGO", "ICE Low Sulphur Gasoil", "ICE低硫柴油", "ICE", "ICE Futures Europe", "UK", "USD", "LOW_SULFUR_FUEL_OIL", "能源", "MEDIUM", "DOMESTIC_FOREIGN_RELATED_UNDERLYING"),
    fa("FUT_SGX_BUNKER", "FUTURE", "COMMODITY_FUTURE", "BUNKER", "Singapore Marine Fuel", "新加坡船燃", "SGX", "Singapore Exchange", "SG", "USD", "LOW_SULFUR_FUEL_OIL", "能源", "MEDIUM", "DOMESTIC_FOREIGN_RELATED_UNDERLYING"),
    fa("REF_NYMEX_ASPHALT", "INDEX", "REFERENCE_INDEX", "ASPHALT", "NYMEX Asphalt Reference", "NYMEX沥青参考", "NYMEX", "NYMEX", "US", "USD", "BITUMEN", "能源", "LOW", "DOMESTIC_FOREIGN_RELATED_UNDERLYING", "REFERENCE_ONLY"),
    # 黑色
    fa("FUT_SGX_FE", "FUTURE", "COMMODITY_FUTURE", "FE", "SGX Iron Ore", "SGX铁矿石", "SGX", "Singapore Exchange", "SG", "USD", "IRON_ORE", "黑色", "HIGH"),
    fa("FUT_CME_TIO", "FUTURE", "COMMODITY_FUTURE", "TIO", "CME Iron Ore", "CME铁矿石", "CME", "CME", "US", "USD", "IRON_ORE", "黑色", "MEDIUM"),
    fa("FUT_SGX_COKING_COAL", "FUTURE", "COMMODITY_FUTURE", "CC", "SGX Coking Coal", "SGX焦煤", "SGX", "Singapore Exchange", "SG", "USD", "COKING_COAL", "黑色", "MEDIUM"),
    fa("FUT_SGX_COKE", "FUTURE", "COMMODITY_FUTURE", "COKE", "SGX Coke Reference", "SGX焦炭参考", "SGX", "Singapore Exchange", "SG", "USD", "COKE", "黑色", "LOW", "DOMESTIC_FOREIGN_RELATED_UNDERLYING", "REFERENCE_ONLY"),
    fa("FUT_LME_REBAR", "FUTURE", "COMMODITY_FUTURE", "SRR", "LME Steel Rebar", "LME螺纹钢", "LME", "London Metal Exchange", "UK", "USD", "REBAR", "黑色", "MEDIUM"),
    fa("FUT_SGX_REBAR", "FUTURE", "COMMODITY_FUTURE", "REBAR", "SGX Steel Rebar", "SGX螺纹钢", "SGX", "Singapore Exchange", "SG", "USD", "REBAR", "黑色", "MEDIUM"),
    fa("FUT_CME_HRC", "FUTURE", "COMMODITY_FUTURE", "HRC", "CME Hot-Rolled Coil Steel", "CME热卷", "CME", "CME", "US", "USD", "HOT_ROLLED_COIL", "黑色", "HIGH"),
    fa("FUT_LME_HRC", "FUTURE", "COMMODITY_FUTURE", "HRC", "LME HRC Steel", "LME热卷", "LME", "London Metal Exchange", "UK", "USD", "HOT_ROLLED_COIL", "黑色", "MEDIUM"),
    fa("REF_SGX_FERROSILICON", "INDEX", "REFERENCE_INDEX", "FeSi", "SGX Ferrosilicon Reference", "SGX硅铁参考", "SGX", "Singapore Exchange", "SG", "USD", "FERROSILICON", "黑色", "LOW", "DOMESTIC_FOREIGN_RELATED_UNDERLYING", "REFERENCE_ONLY"),
    fa("REF_GLOBAL_SILICOMANGANESE", "INDEX", "REFERENCE_INDEX", "SiMn", "Global Silicomanganese Index", "全球硅锰指数", "GLOBAL", "Global Reference", "GLOBAL", "USD", "SILICOMANGANESE", "黑色", "LOW", "DOMESTIC_FOREIGN_RELATED_UNDERLYING", "REFERENCE_ONLY"),
    # 农产品和油脂
    fa("FUT_CBOT_ZS", "FUTURE", "COMMODITY_FUTURE", "ZS", "CBOT Soybeans", "CBOT大豆", "CBOT", "CBOT", "US", "USD", "SOYBEAN", "农产品", "HIGH"),
    fa("FUT_CBOT_ZM", "FUTURE", "COMMODITY_FUTURE", "ZM", "CBOT Soybean Meal", "CBOT豆粕", "CBOT", "CBOT", "US", "USD", "SOYBEAN_MEAL", "农产品", "HIGH"),
    fa("FUT_CBOT_ZL", "FUTURE", "COMMODITY_FUTURE", "ZL", "CBOT Soybean Oil", "CBOT豆油", "CBOT", "CBOT", "US", "USD", "SOYBEAN_OIL", "农产品", "HIGH"),
    fa("FUT_CBOT_ZC", "FUTURE", "COMMODITY_FUTURE", "ZC", "CBOT Corn", "CBOT玉米", "CBOT", "CBOT", "US", "USD", "CORN", "农产品", "HIGH"),
    fa("FUT_CBOT_ZW", "FUTURE", "COMMODITY_FUTURE", "ZW", "CBOT Wheat", "CBOT小麦", "CBOT", "CBOT", "US", "USD", "COMMON_WHEAT", "农产品", "HIGH"),
    fa("FUT_KCBT_KE", "FUTURE", "COMMODITY_FUTURE", "KE", "KCBT Hard Red Winter Wheat", "KCBT硬红冬麦", "KCBT", "Kansas City Board of Trade", "US", "USD", "HARD_WHEAT", "农产品", "MEDIUM"),
    fa("FUT_CBOT_ZW_STRONG", "FUTURE", "COMMODITY_FUTURE", "ZW", "CBOT Wheat", "CBOT小麦", "CBOT", "CBOT", "US", "USD", "STRONG_WHEAT", "农产品", "MEDIUM"),
    fa("FUT_BMD_FCPO", "FUTURE", "COMMODITY_FUTURE", "FCPO", "BMD Crude Palm Oil", "BMD棕榈油", "BMD", "Bursa Malaysia Derivatives", "MY", "MYR", "PALM_OIL", "农产品", "HIGH"),
    fa("FUT_ICE_SB", "FUTURE", "COMMODITY_FUTURE", "SB", "ICE Raw Sugar No.11", "ICE原糖11号", "ICE", "ICE Futures US", "US", "USD", "SUGAR", "农产品", "HIGH"),
    fa("FUT_ICE_CT", "FUTURE", "COMMODITY_FUTURE", "CT", "ICE Cotton No.2", "ICE棉花2号", "ICE", "ICE Futures US", "US", "USD", "COTTON", "农产品", "HIGH"),
    fa("FUT_ICE_KC", "FUTURE", "COMMODITY_FUTURE", "KC", "ICE Coffee", "ICE咖啡", "ICE", "ICE Futures US", "US", "USD", "COFFEE", "农产品", "HIGH"),
    fa("FUT_ICE_CC", "FUTURE", "COMMODITY_FUTURE", "CC", "ICE Cocoa", "ICE可可", "ICE", "ICE Futures US", "US", "USD", "COCOA", "农产品", "HIGH"),
    fa("FUT_CME_HE", "FUTURE", "COMMODITY_FUTURE", "HE", "CME Lean Hogs", "CME瘦肉猪", "CME", "CME", "US", "USD", "LIVE_HOG", "农产品", "MEDIUM", "DOMESTIC_FOREIGN_RELATED_UNDERLYING"),
    fa("REF_GLOBAL_EGG", "INDEX", "REFERENCE_INDEX", "EGG", "Global Egg Reference", "全球鸡蛋参考", "GLOBAL", "Global Reference", "GLOBAL", "USD", "EGG", "农产品", "LOW", "DOMESTIC_FOREIGN_RELATED_UNDERLYING", "REFERENCE_ONLY"),
    fa("REF_OILSEED_PEANUT", "INDEX", "REFERENCE_INDEX", "PEANUT", "Oilseed Peanut Reference", "油籽花生参考", "GLOBAL", "Global Reference", "GLOBAL", "USD", "PEANUT", "农产品", "LOW", "DOMESTIC_FOREIGN_RELATED_UNDERLYING", "REFERENCE_ONLY"),
    fa("FUT_ICE_CANOLA", "FUTURE", "COMMODITY_FUTURE", "RS", "ICE Canola", "ICE油菜籽", "ICE", "ICE Futures Canada", "CA", "CAD", "RAPESEED", "农产品", "MEDIUM"),
    fa("FUT_MATIF_RAPESEED", "FUTURE", "COMMODITY_FUTURE", "RAP", "MATIF Rapeseed", "MATIF油菜籽", "MATIF", "Euronext MATIF", "FR", "EUR", "RAPESEED", "农产品", "MEDIUM"),
    fa("FUT_ICE_CANOLA_OIL", "FUTURE", "COMMODITY_FUTURE", "RSO", "Canola Oil Reference", "菜籽油参考", "ICE", "ICE Futures Canada", "CA", "CAD", "RAPESEED_OIL", "农产品", "MEDIUM", "DOMESTIC_FOREIGN_RELATED_UNDERLYING"),
    fa("FUT_ICE_CANOLA_MEAL", "FUTURE", "COMMODITY_FUTURE", "RSM", "Canola Meal Reference", "菜粕参考", "ICE", "ICE Futures Canada", "CA", "CAD", "RAPESEED_MEAL", "农产品", "MEDIUM", "DOMESTIC_FOREIGN_RELATED_UNDERLYING"),
    # 能源化工
    fa("REF_ASIA_PX", "INDEX", "REFERENCE_INDEX", "PX", "Asia Paraxylene Index", "亚洲PX指数", "ASIA", "Asia Reference", "GLOBAL", "USD", "PARAXYLENE", "化工", "MEDIUM", "DOMESTIC_FOREIGN_RELATED_UNDERLYING"),
    fa("REF_ASIA_PTA", "INDEX", "REFERENCE_INDEX", "PTA", "Asia PTA Reference", "亚洲PTA参考", "ASIA", "Asia Reference", "GLOBAL", "USD", "PTA", "化工", "MEDIUM", "DOMESTIC_FOREIGN_RELATED_UNDERLYING"),
    fa("REF_GLOBAL_METHANOL", "INDEX", "REFERENCE_INDEX", "MEOH", "Methanol Reference", "甲醇参考", "GLOBAL", "Global Reference", "GLOBAL", "USD", "METHANOL", "化工", "LOW", "DOMESTIC_FOREIGN_RELATED_UNDERLYING", "REFERENCE_ONLY"),
    fa("REF_CME_LLDPE", "INDEX", "REFERENCE_INDEX", "LLDPE", "CME LLDPE Reference", "CME线性低密度聚乙烯参考", "CME", "CME", "US", "USD", "LLDPE", "化工", "LOW", "DOMESTIC_FOREIGN_RELATED_UNDERLYING", "REFERENCE_ONLY"),
    fa("REF_ASIA_PVC", "INDEX", "REFERENCE_INDEX", "PVC", "Asia PVC Index", "亚洲PVC指数", "ASIA", "Asia Reference", "GLOBAL", "USD", "PVC", "化工", "LOW", "DOMESTIC_FOREIGN_RELATED_UNDERLYING", "REFERENCE_ONLY"),
    fa("REF_ASIA_PP", "INDEX", "REFERENCE_INDEX", "PP", "Asia Polypropylene Index", "亚洲PP指数", "ASIA", "Asia Reference", "GLOBAL", "USD", "POLYPROPYLENE", "化工", "LOW", "DOMESTIC_FOREIGN_RELATED_UNDERLYING", "REFERENCE_ONLY"),
    fa("REF_ASIA_MEG", "INDEX", "REFERENCE_INDEX", "MEG", "Asia MEG Index", "亚洲乙二醇指数", "ASIA", "Asia Reference", "GLOBAL", "USD", "ETHYLENE_GLYCOL", "化工", "LOW", "DOMESTIC_FOREIGN_RELATED_UNDERLYING", "REFERENCE_ONLY"),
    fa("REF_ASIA_STYRENE", "INDEX", "REFERENCE_INDEX", "SM", "Asia Styrene Index", "亚洲苯乙烯指数", "ASIA", "Asia Reference", "GLOBAL", "USD", "STYRENE", "化工", "LOW", "DOMESTIC_FOREIGN_RELATED_UNDERLYING", "REFERENCE_ONLY"),
    fa("FUT_OSE_RSS3", "FUTURE", "COMMODITY_FUTURE", "RSS3", "OSE Rubber RSS3", "OSE橡胶RSS3", "OSE", "Osaka Exchange", "JP", "JPY", "RUBBER", "化工", "HIGH"),
    fa("FUT_SICOM_TSR20", "FUTURE", "COMMODITY_FUTURE", "TSR20", "SICOM TSR20 Rubber", "SICOM TSR20橡胶", "SICOM", "Singapore Commodity Exchange", "SG", "USD", "RUBBER", "化工", "HIGH"),
    fa("REF_ASIA_BUTADIENE_RUBBER", "INDEX", "REFERENCE_INDEX", "BR", "Asia Butadiene Rubber Index", "亚洲丁二烯橡胶指数", "ASIA", "Asia Reference", "GLOBAL", "USD", "BUTADIENE_RUBBER", "化工", "MEDIUM", "DOMESTIC_FOREIGN_RELATED_UNDERLYING"),
    fa("REF_CME_PULP", "INDEX", "REFERENCE_INDEX", "PULP", "CME Pulp Reference", "CME纸浆参考", "CME", "CME", "US", "USD", "PULP", "化工", "MEDIUM", "DOMESTIC_FOREIGN_RELATED_UNDERLYING"),
    fa("REF_CME_UREA", "INDEX", "REFERENCE_INDEX", "UREA", "CME Urea Reference", "CME尿素参考", "CME", "CME", "US", "USD", "UREA", "化工", "MEDIUM", "DOMESTIC_FOREIGN_RELATED_UNDERLYING"),
    fa("REF_GLOBAL_SODA_ASH", "INDEX", "REFERENCE_INDEX", "SODA", "International Soda Ash Index", "国际纯碱指数", "GLOBAL", "Global Reference", "GLOBAL", "USD", "SODA_ASH", "化工", "LOW", "DOMESTIC_FOREIGN_RELATED_UNDERLYING", "REFERENCE_ONLY"),
    fa("REF_GLOBAL_GLASS", "INDEX", "REFERENCE_INDEX", "GLASS", "Global Glass Reference", "全球玻璃参考", "GLOBAL", "Global Reference", "GLOBAL", "USD", "GLASS", "建材", "LOW", "DOMESTIC_FOREIGN_RELATED_UNDERLYING", "REFERENCE_ONLY"),
    # 新能源
    fa("FUT_CME_LITHIUM_HYDROXIDE", "FUTURE", "COMMODITY_FUTURE", "LTH", "CME Lithium Hydroxide", "CME氢氧化锂", "CME", "CME", "US", "USD", "LITHIUM_CARBONATE", "新能源", "MEDIUM", "DOMESTIC_FOREIGN_RELATED_UNDERLYING"),
    fa("FUT_LME_LITHIUM_HYDROXIDE", "FUTURE", "COMMODITY_FUTURE", "LITH", "LME Lithium Hydroxide", "LME氢氧化锂", "LME", "London Metal Exchange", "UK", "USD", "LITHIUM_CARBONATE", "新能源", "MEDIUM", "DOMESTIC_FOREIGN_RELATED_UNDERLYING"),
    fa("REF_FASTMARKETS_LITHIUM", "INDEX", "REFERENCE_INDEX", "LITHIUM", "Fastmarkets Lithium Index", "Fastmarkets锂指数", "FASTMARKETS", "Fastmarkets", "GLOBAL", "USD", "LITHIUM_CARBONATE", "新能源", "MEDIUM", "DOMESTIC_FOREIGN_RELATED_UNDERLYING"),
    fa("REF_SILICON_METAL", "INDEX", "REFERENCE_INDEX", "SI", "Silicon Metal Europe US Index", "欧美金属硅指数", "GLOBAL", "Global Reference", "GLOBAL", "USD", "INDUSTRIAL_SILICON", "新能源", "LOW", "DOMESTIC_FOREIGN_RELATED_UNDERLYING", "REFERENCE_ONLY"),
    fa("REF_POLYSILICON", "INDEX", "REFERENCE_INDEX", "PS", "Polysilicon Price Index", "多晶硅价格指数", "GLOBAL", "Global Reference", "GLOBAL", "USD", "POLYSILICON", "新能源", "MEDIUM", "DOMESTIC_FOREIGN_RELATED_UNDERLYING"),
    # 股指 ETF
    fa("ETF_US_ASHR", "ETF", "INDEX_ETF", "ASHR", "Xtrackers Harvest CSI 300 China A-Shares ETF", "ASHR沪深300A股ETF", "NYSEARCA", "NYSE Arca", "US", "USD", "CSI300", "股指", "MEDIUM", "DOMESTIC_FOREIGN_ETF_RELATED", "GLOBAL_MACRO_HEDGE"),
    fa("ETF_US_KBA", "ETF", "INDEX_ETF", "KBA", "KraneShares Bosera MSCI China A 50 Connect ETF", "KBA中国A股ETF", "NYSEARCA", "NYSE Arca", "US", "USD", "CSI300", "股指", "MEDIUM", "DOMESTIC_FOREIGN_ETF_RELATED", "GLOBAL_MACRO_HEDGE"),
    fa("FUT_SGX_CN", "FUTURE", "INDEX_FUTURE", "CN", "SGX FTSE China A50 Futures", "SGX富时中国A50期货", "SGX", "Singapore Exchange", "SG", "USD", "CSI300", "股指", "MEDIUM", "DOMESTIC_FOREIGN_INDEX_RELATED", "GLOBAL_MACRO_HEDGE"),
    fa("FUT_SGX_CN_SSE50", "FUTURE", "INDEX_FUTURE", "CN", "SGX FTSE China A50 Futures", "SGX富时中国A50期货", "SGX", "Singapore Exchange", "SG", "USD", "SSE50", "股指", "MEDIUM", "DOMESTIC_FOREIGN_INDEX_RELATED", "GLOBAL_MACRO_HEDGE"),
    fa("ETF_US_CHINA_LARGECAP", "ETF", "INDEX_ETF", "FXI", "China Large-Cap ETF", "中国大盘ETF", "NYSEARCA", "NYSE Arca", "US", "USD", "SSE50", "股指", "LOW", "DOMESTIC_FOREIGN_ETF_RELATED", "REFERENCE_ONLY"),
    fa("ETF_US_CHINA_MIDCAP", "ETF", "INDEX_ETF", "ASHR_MID", "China Small Mid Cap ETF Reference", "中国中小盘ETF参考", "NYSEARCA", "NYSE Arca", "US", "USD", "CSI500", "股指", "LOW", "DOMESTIC_FOREIGN_ETF_RELATED", "REFERENCE_ONLY"),
    fa("ETF_US_CHINA_SMALLCAP", "ETF", "INDEX_ETF", "KURE", "China Small Cap ETF Reference", "中国小盘ETF参考", "NYSEARCA", "NYSE Arca", "US", "USD", "CSI1000", "股指", "LOW", "DOMESTIC_FOREIGN_ETF_RELATED", "REFERENCE_ONLY"),
    fa("FUT_CME_ES", "FUTURE", "INDEX_FUTURE", "ES", "CME E-mini S&P 500", "CME标普500期货", "CME", "CME", "US", "USD", "SPX", "股指", "HIGH", "DOMESTIC_FOREIGN_INDEX_RELATED", "GLOBAL_MACRO_HEDGE"),
    fa("ETF_US_SPY", "ETF", "INDEX_ETF", "SPY", "SPDR S&P 500 ETF", "标普500ETF", "NYSEARCA", "NYSE Arca", "US", "USD", "SPX", "股指", "HIGH", "DOMESTIC_FOREIGN_ETF_RELATED", "GLOBAL_MACRO_HEDGE"),
    fa("FUT_CME_NQ", "FUTURE", "INDEX_FUTURE", "NQ", "CME E-mini Nasdaq 100", "CME纳指100期货", "CME", "CME", "US", "USD", "NASDAQ", "股指", "HIGH", "DOMESTIC_FOREIGN_INDEX_RELATED", "GLOBAL_MACRO_HEDGE"),
    fa("ETF_US_QQQ", "ETF", "INDEX_ETF", "QQQ", "Invesco QQQ ETF", "纳指100ETF", "NASDAQ", "NASDAQ", "US", "USD", "NASDAQ", "股指", "HIGH", "DOMESTIC_FOREIGN_ETF_RELATED", "GLOBAL_MACRO_HEDGE"),
    # 国债利率
    fa("FUT_CME_ZT", "FUTURE", "BOND_FUTURE", "ZT", "CME 2-Year T-Note", "CME 2年美债期货", "CME", "CME", "US", "USD", "CGB_2Y", "利率", "MEDIUM", "DOMESTIC_FOREIGN_RELATED_UNDERLYING", "GLOBAL_MACRO_HEDGE"),
    fa("FUT_CME_ZF", "FUTURE", "BOND_FUTURE", "ZF", "CME 5-Year T-Note", "CME 5年美债期货", "CME", "CME", "US", "USD", "CGB_5Y", "利率", "MEDIUM", "DOMESTIC_FOREIGN_RELATED_UNDERLYING", "GLOBAL_MACRO_HEDGE"),
    fa("FUT_CME_ZN", "FUTURE", "BOND_FUTURE", "ZN", "CME 10-Year T-Note", "CME 10年美债期货", "CME", "CME", "US", "USD", "CGB_10Y", "利率", "MEDIUM", "DOMESTIC_FOREIGN_RELATED_UNDERLYING", "GLOBAL_MACRO_HEDGE"),
    fa("FUT_CME_ZB", "FUTURE", "BOND_FUTURE", "ZB", "CME 30-Year T-Bond", "CME 30年美债期货", "CME", "CME", "US", "USD", "CGB_30Y", "利率", "MEDIUM", "DOMESTIC_FOREIGN_RELATED_UNDERLYING", "GLOBAL_MACRO_HEDGE"),
    # FX/Crypto foreign assets only unless domestic appears later.
    fa("FUT_CME_CNH", "FUTURE", "FX_FUTURE", "CNH", "CME Chinese Renminbi Futures", "CME人民币期货", "CME", "CME", "US", "USD", "USD_CNY", "外汇", "MEDIUM", "DOMESTIC_FOREIGN_FX_RELATED", "GLOBAL_MACRO_HEDGE"),
    fa("FUT_CME_6E", "FUTURE", "FX_FUTURE", "6E", "CME Euro FX", "CME欧元期货", "CME", "CME", "US", "USD", "EUR_USD", "外汇", "HIGH", "DOMESTIC_FOREIGN_FX_RELATED", "GLOBAL_MACRO_HEDGE"),
    fa("FUT_CME_BTC", "FUTURE", "CRYPTO_FUTURE", "BTC", "CME Bitcoin Futures", "CME比特币期货", "CME", "CME", "US", "USD", "BTC", "加密货币", "HIGH", "DOMESTIC_FOREIGN_CRYPTO_RELATED", "REFERENCE_ONLY"),
    fa("FUT_CME_ETH", "FUTURE", "CRYPTO_FUTURE", "ETH", "CME Ether Futures", "CME以太坊期货", "CME", "CME", "US", "USD", "ETH", "加密货币", "HIGH", "DOMESTIC_FOREIGN_CRYPTO_RELATED", "REFERENCE_ONLY"),
]


def ensure_dirs() -> None:
    for path in [FINAL_DIR, PROCESSED_DIR, MANUAL_DIR, OUTPUT_DIR, EXCEL_OUT.parent]:
        path.mkdir(parents=True, exist_ok=True)


def clean_text(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    return str(value).strip()


def is_missing(value: Any) -> bool:
    return clean_text(value).upper() in MISSING_VALUES


def read_csv(path: Path, required: bool = True) -> pd.DataFrame:
    if not path.exists():
        if required:
            raise FileNotFoundError(f"缺少输入文件: {path}")
        return pd.DataFrame()
    return pd.read_csv(path, dtype=str, encoding=ENCODING).fillna("")


def save_csv(df: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(path, index=False, encoding=ENCODING)


def choose_asset_master() -> tuple[pd.DataFrame, str]:
    if ASSET_MASTER_PRIMARY.exists():
        return read_csv(ASSET_MASTER_PRIMARY), str(ASSET_MASTER_PRIMARY)
    return read_csv(ASSET_MASTER_FALLBACK), str(ASSET_MASTER_FALLBACK)


def build_foreign_asset_master() -> pd.DataFrame:
    df = pd.DataFrame(FOREIGN_RULES)
    public_cols = [
        "foreign_asset_id", "foreign_asset_type", "foreign_subtype", "foreign_symbol",
        "foreign_name_en", "foreign_name_cn", "foreign_exchange_code", "foreign_exchange_name",
        "foreign_country", "foreign_currency", "foreign_contract_unit", "foreign_quote_unit",
        "foreign_tick_size", "underlying_group", "sector", "tradable", "can_long", "can_short",
        "data_source", "source_status", "notes",
    ]
    save_csv(df[public_cols], FOREIGN_ASSET_OUT)
    return df


def asset_rank(row: pd.Series) -> tuple[int, str]:
    rank = {"FUTURE": 0, "ETF": 1, "INDEX": 2, "SPOT": 2, "OPTION": 3}.get(clean_text(row.get("asset_type")), 9)
    return rank, clean_text(row.get("asset_id"))


def build_domestic_preferred(asset_master: pd.DataFrame, underlying: pd.DataFrame) -> dict[str, pd.Series]:
    cols = ["asset_id", "symbol", "name_cn", "asset_type", "subtype", "exchange_code", "underlying_group"]
    base = underlying[cols].copy()
    enrich_cols = ["asset_id", "country", "currency", "tradable", "can_long", "can_short"]
    enrich = asset_master[[c for c in enrich_cols if c in asset_master.columns]].drop_duplicates("asset_id")
    base = base.merge(enrich, on="asset_id", how="left")
    base = base[~base["underlying_group"].map(is_missing)].copy()
    base = base[base["asset_type"].isin(["FUTURE", "ETF", "INDEX", "SPOT", "OPTION"])].copy()

    preferred: dict[str, pd.Series] = {}
    for group, group_df in base.groupby("underlying_group"):
        main = group_df[group_df["asset_type"].isin(["FUTURE", "ETF", "INDEX", "SPOT"])].copy()
        if main.empty:
            continue
        main["_rank"] = main.apply(asset_rank, axis=1)
        preferred[group] = main.sort_values("_rank").iloc[0].drop(labels=["_rank"], errors="ignore")
    return preferred


def unit_conversion_for(group: str) -> str:
    if group == "GOLD":
        return "gram_to_troy_ounce_conversion_needed"
    if group == "SILVER":
        return "kg_to_troy_ounce_conversion_needed"
    if group in {"COPPER", "ALUMINUM", "ZINC", "LEAD", "NICKEL", "TIN"}:
        return "metric_ton_to_lme_lot_conversion_needed"
    if group == "CRUDE_OIL":
        return "barrel_to_metric_ton_conversion_needed"
    if group in {"SOYBEAN", "CORN", "COMMON_WHEAT", "STRONG_WHEAT", "HARD_WHEAT"}:
        return "bushel_to_metric_ton_conversion_needed"
    if group in {"PALM_OIL", "IRON_ORE"}:
        return "metric_ton_conversion_needed"
    if group in {"CSI300", "SSE50", "CSI500", "CSI1000", "SPX", "NASDAQ"}:
        return "index_multiplier_conversion_needed"
    return "CHECK"


def trading_overlap(country: str) -> str:
    if country in {"SG", "JP", "MY", "HK"}:
        return "PARTIAL_ASIA_HOURS_OVERLAP"
    if country in {"US", "UK", "FR", "CA"}:
        return "LIMITED_OR_NIGHT_SESSION_OVERLAP"
    return "CHECK"


def build_mapping(foreign_rules: pd.DataFrame, preferred: dict[str, pd.Series]) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    rows = []
    missing = []
    for _, f in foreign_rules.iterrows():
        group = clean_text(f["underlying_group"])
        d = preferred.get(group)
        if d is None:
            missing.append(
                {
                    "foreign_asset_id": f["foreign_asset_id"],
                    "foreign_symbol": f["foreign_symbol"],
                    "foreign_name_en": f["foreign_name_en"],
                    "underlying_group": group,
                    "foreign_exchange_code": f["foreign_exchange_code"],
                    "mapping_confidence": f["_mapping_confidence"],
                    "issue_reason": "当前国内资产库未找到对应 underlying_group 的 FUTURE/ETF/INDEX 主映射资产。",
                    "suggested_next_step": "后续如新增国内相关资产，可重新运行 Step 7 生成国外跨市场映射。",
                }
            )
            continue
        domestic_currency = clean_text(d.get("currency")) or "CNY"
        foreign_currency = clean_text(f["foreign_currency"])
        relation_type = clean_text(f["_relation_type"])
        mapping_id = f"{relation_type}_{clean_text(d['asset_id'])}_{clean_text(f['foreign_asset_id'])}"
        rows.append(
            {
                "mapping_id": mapping_id,
                "domestic_asset_id": clean_text(d["asset_id"]),
                "foreign_asset_id": clean_text(f["foreign_asset_id"]),
                "domestic_symbol": clean_text(d["symbol"]),
                "foreign_symbol": clean_text(f["foreign_symbol"]),
                "domestic_name": clean_text(d["name_cn"]),
                "foreign_name": clean_text(f["foreign_name_cn"]),
                "domestic_asset_type": clean_text(d["asset_type"]),
                "foreign_asset_type": clean_text(f["foreign_asset_type"]),
                "domestic_exchange": clean_text(d["exchange_code"]),
                "foreign_exchange": clean_text(f["foreign_exchange_code"]),
                "domestic_country": clean_text(d.get("country")) or "CN",
                "foreign_country": clean_text(f["foreign_country"]),
                "underlying_group": group,
                "currency_domestic": domestic_currency,
                "currency_foreign": foreign_currency,
                "unit_domestic": "CHECK",
                "unit_foreign": clean_text(f["foreign_contract_unit"]) or "CHECK",
                "unit_conversion": unit_conversion_for(group),
                "fx_conversion_needed": "Y" if domestic_currency != foreign_currency else "N",
                "trading_hour_overlap": trading_overlap(clean_text(f["foreign_country"])),
                "relation_type": relation_type,
                "strategy_type": clean_text(f["_strategy_type"]),
                "direction_supported": "BIDIRECTIONAL_WITH_FX_AND_HOURS_CONSTRAINT",
                "mapping_confidence": clean_text(f["_mapping_confidence"]),
                "data_source": "STEP7_BUILT_IN_MANUAL_RULES",
                "source_status": "FOREIGN_MAPPING_CANDIDATE_NEED_REVIEW",
                "notes": "基于内置规则生成，单位换算、汇率换算和交易时间重叠需后续细化。",
            }
        )
    mapping = pd.DataFrame(rows)
    if not mapping.empty:
        mapping = mapping.drop_duplicates(["domestic_asset_id", "foreign_asset_id", "relation_type"]).drop_duplicates("mapping_id").reset_index(drop=True)
    missing_df = pd.DataFrame(missing)
    low_df = mapping[mapping["mapping_confidence"] == "LOW"].copy() if not mapping.empty else pd.DataFrame()
    save_csv(mapping, FOREIGN_MAPPING_OUT)
    save_csv(missing_df, MISSING_DOMESTIC_OUT)
    save_csv(low_df, LOW_CONFIDENCE_OUT)
    return mapping, missing_df, low_df


def build_quality(foreign_assets_public: pd.DataFrame, mapping: pd.DataFrame, missing: pd.DataFrame) -> pd.DataFrame:
    asset_ids = set(foreign_assets_public["foreign_asset_id"])
    domestic_ids = set(read_csv(UNDERLYING_MAPPING)["asset_id"])
    rows = [{
        "section": "overall",
        "underlying_group": "ALL",
        "foreign_exchange_code": "ALL",
        "relation_type": "ALL",
        "total_foreign_asset_count": len(foreign_assets_public),
        "total_foreign_mapping_count": len(mapping),
        "missing_domestic_count": len(missing),
        "low_confidence_mapping_count": int((mapping["mapping_confidence"] == "LOW").sum()) if not mapping.empty else 0,
        "high_confidence_mapping_count": int((mapping["mapping_confidence"] == "HIGH").sum()) if not mapping.empty else 0,
        "medium_confidence_mapping_count": int((mapping["mapping_confidence"] == "MEDIUM").sum()) if not mapping.empty else 0,
        "duplicate_mapping_id_count": int(mapping["mapping_id"].duplicated().sum()) if not mapping.empty else 0,
        "domestic_asset_missing_backlink_count": len(set(mapping["domestic_asset_id"]) - domestic_ids) if not mapping.empty else 0,
        "foreign_asset_missing_backlink_count": len(set(mapping["foreign_asset_id"]) - asset_ids) if not mapping.empty else 0,
        "foreign_asset_count": len(foreign_assets_public),
        "mapping_count": len(mapping),
        "count": len(mapping),
    }]
    for group, fa_group in foreign_assets_public.groupby("underlying_group"):
        m = mapping[mapping["underlying_group"] == group] if not mapping.empty else pd.DataFrame()
        rows.append({
            "section": "by_underlying_group", "underlying_group": group, "foreign_exchange_code": "", "relation_type": "",
            "foreign_asset_count": len(fa_group), "mapping_count": len(m),
            "high_confidence_count": int((m["mapping_confidence"] == "HIGH").sum()) if not m.empty else 0,
            "medium_confidence_count": int((m["mapping_confidence"] == "MEDIUM").sum()) if not m.empty else 0,
            "low_confidence_count": int((m["mapping_confidence"] == "LOW").sum()) if not m.empty else 0,
        })
    for exch, fa_group in foreign_assets_public.groupby("foreign_exchange_code"):
        m = mapping[mapping["foreign_exchange"] == exch] if not mapping.empty else pd.DataFrame()
        rows.append({"section": "by_foreign_exchange_code", "foreign_exchange_code": exch, "underlying_group": "", "relation_type": "", "foreign_asset_count": len(fa_group), "mapping_count": len(m)})
    if not mapping.empty:
        for rel, m in mapping.groupby("relation_type"):
            rows.append({"section": "by_relation_type", "relation_type": rel, "underlying_group": "", "foreign_exchange_code": "", "count": len(m)})
    report = pd.DataFrame(rows).fillna("")
    save_csv(report, QUALITY_REPORT_OUT)
    return report


def readme_df() -> pd.DataFrame:
    return pd.DataFrame([
        ("项目", "Step 7 国外跨市场映射候选表"),
        ("生成方式", "该表基于内置规则生成，不调用外部 API。"),
        ("用途", "主要用于后续内外盘套利、跨市场套利和全球相关品种映射。"),
        ("HIGH", "较明确的同标的关系。"),
        ("MEDIUM", "相关性较强但存在口径差异。"),
        ("LOW", "弱相关或参考关系。"),
        ("后续", "单位换算、汇率换算和交易时间重叠仍需后续细化。"),
    ], columns=["item", "description"])


def write_excel(foreign_assets: pd.DataFrame, mapping: pd.DataFrame, quality: pd.DataFrame, missing: pd.DataFrame, low: pd.DataFrame) -> None:
    sheets = {
        "README": readme_df(),
        "foreign_asset_master": foreign_assets,
        "foreign_cross_market_mapping": mapping,
        "foreign_mapping_quality_report": quality,
        "foreign_mapping_missing_domestic_review": missing,
        "foreign_mapping_low_confidence_review": low,
    }
    with pd.ExcelWriter(EXCEL_OUT, engine="openpyxl") as writer:
        for name, df in sheets.items():
            sheet = name[:31]
            df.to_excel(writer, sheet_name=sheet, index=False)
            ws = writer.book[sheet]
            ws.freeze_panes = "A2"
            if ws.max_row > 1 and ws.max_column > 1:
                ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.font = Font(bold=True)
            for col_cells in ws.columns:
                width = 10
                letter = col_cells[0].column_letter
                for cell in col_cells[:200]:
                    width = max(width, min(len("" if cell.value is None else str(cell.value)) + 2, 40))
                ws.column_dimensions[letter].width = width


def main() -> None:
    try:
        ensure_dirs()
        asset_master, source_path = choose_asset_master()
        underlying = read_csv(UNDERLYING_MAPPING)
        read_csv(DOMESTIC_MAPPING, required=False)
        read_csv(INDUSTRY_MAPPING, required=False)
        foreign_rules = build_foreign_asset_master()
        foreign_public = foreign_rules[[c for c in foreign_rules.columns if not c.startswith("_")]].copy()
        preferred = build_domestic_preferred(asset_master, underlying)
        mapping, missing, low = build_mapping(foreign_rules, preferred)
        quality = build_quality(foreign_public, mapping, missing)
        write_excel(foreign_public, mapping, quality, missing, low)

        duplicate_count = int(mapping["mapping_id"].duplicated().sum()) if not mapping.empty else 0
        domestic_backlink = int(quality.loc[quality["section"] == "overall", "domestic_asset_missing_backlink_count"].iloc[0])
        foreign_backlink = int(quality.loc[quality["section"] == "overall", "foreign_asset_missing_backlink_count"].iloc[0])
        print(f"使用的 asset_master 来源: {source_path}")
        print(f"foreign_asset_master 行数: {len(foreign_public)}")
        print(f"foreign_cross_market_mapping 行数: {len(mapping)}")
        print(f"missing domestic review 行数: {len(missing)}")
        print(f"low confidence review 行数: {len(low)}")
        print(f"HIGH confidence 数量: {int((mapping['mapping_confidence'] == 'HIGH').sum()) if not mapping.empty else 0}")
        print(f"MEDIUM confidence 数量: {int((mapping['mapping_confidence'] == 'MEDIUM').sum()) if not mapping.empty else 0}")
        print(f"LOW confidence 数量: {int((mapping['mapping_confidence'] == 'LOW').sum()) if not mapping.empty else 0}")
        print(f"duplicate_mapping_id_count: {duplicate_count}")
        print(f"domestic_asset_missing_backlink_count: {domestic_backlink}")
        print(f"foreign_asset_missing_backlink_count: {foreign_backlink}")
        print("各 foreign_exchange_code 数量:")
        for exch, count in foreign_public["foreign_exchange_code"].value_counts().sort_index().items():
            print(f"  {exch}: {count}")
        print(f"Excel 输出路径: {EXCEL_OUT}")
    except Exception as exc:
        print(f"Step 7 国外跨市场映射生成失败: {exc}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
