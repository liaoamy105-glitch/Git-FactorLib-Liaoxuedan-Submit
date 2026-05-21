from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment


MAPPING_SHEET_INPUT = Path("data/processed/mapping_sheet.csv")
DETAIL_SHEET_INPUT = Path("data/processed/detail_sheet.csv")
OUTPUT_EXCEL = Path("data/output/mapping_info.xlsx")
CHECK_OUTPUT = Path("data/processed/final_excel_check.csv")

SHEET_NAMES = ["映射表", "详情表"]


def add_check(checks, check_item, passed, message):
    checks.append(
        {
            "check_item": check_item,
            "status": "passed" if passed else "failed",
            "message": message,
        }
    )


def save_checks(checks):
    CHECK_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    pd.DataFrame(checks, columns=["check_item", "status", "message"]).to_csv(
        CHECK_OUTPUT, index=False, encoding="utf-8-sig"
    )


def fail_with_checks(checks):
    save_checks(checks)
    failed_items = [
        check["check_item"] for check in checks if check["status"] == "failed"
    ]
    print("final excel export failed")
    print("failed checks: " + ", ".join(failed_items))
    raise SystemExit(1)


def check_inputs():
    checks = []

    mapping_exists = MAPPING_SHEET_INPUT.exists()
    detail_exists = DETAIL_SHEET_INPUT.exists()

    add_check(
        checks,
        "input_mapping_sheet_exists",
        mapping_exists,
        "mapping_sheet.csv exists"
        if mapping_exists
        else "missing mapping_sheet.csv, please run scripts/04_build_mapping_sheet.py first",
    )
    add_check(
        checks,
        "input_detail_sheet_exists",
        detail_exists,
        "detail_sheet.csv exists"
        if detail_exists
        else "missing detail_sheet.csv, please run scripts/05_build_detail_sheet.py first",
    )

    if not mapping_exists or not detail_exists:
        fail_with_checks(checks)

    return checks


def autosize_width(value):
    if pd.isna(value):
        return 0
    text = str(value)
    lines = text.splitlines() or [text]
    return max(len(line) for line in lines)


def format_workbook(path):
    workbook = load_workbook(path)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.font = Font(bold=True)

        column_widths = {}
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                column_letter = cell.column_letter
                column_widths[column_letter] = max(
                    column_widths.get(column_letter, 0),
                    autosize_width(cell.value),
                )

        for column_letter, width in column_widths.items():
            worksheet.column_dimensions[column_letter].width = min(width + 2, 40)

    workbook.save(path)


def export_excel(mapping_sheet, detail_sheet):
    OUTPUT_EXCEL.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT_EXCEL, engine="openpyxl") as writer:
        mapping_sheet.to_excel(writer, sheet_name="映射表", index=False)
        detail_sheet.to_excel(writer, sheet_name="详情表", index=False)

    format_workbook(OUTPUT_EXCEL)


def validate_export(checks, mapping_sheet, detail_sheet):
    output_exists = OUTPUT_EXCEL.exists()
    add_check(
        checks,
        "output_excel_exists",
        output_exists,
        "output Excel exists" if output_exists else "output Excel was not created",
    )
    if not output_exists:
        return checks

    excel_file = pd.ExcelFile(OUTPUT_EXCEL)
    sheet_names_match = excel_file.sheet_names == SHEET_NAMES
    add_check(
        checks,
        "sheet_names_exactly_two",
        sheet_names_match,
        f"sheet names: {excel_file.sheet_names}",
    )

    exported_mapping = pd.read_excel(OUTPUT_EXCEL, sheet_name="映射表", dtype=str)
    exported_detail = pd.read_excel(OUTPUT_EXCEL, sheet_name="详情表", dtype=str)

    mapping_shape_match = exported_mapping.shape == mapping_sheet.shape
    detail_shape_match = exported_detail.shape == detail_sheet.shape

    add_check(
        checks,
        "mapping_sheet_shape_match",
        mapping_shape_match,
        f"Excel 映射表 shape {exported_mapping.shape}, CSV shape {mapping_sheet.shape}",
    )
    add_check(
        checks,
        "detail_sheet_shape_match",
        detail_shape_match,
        f"Excel 详情表 shape {exported_detail.shape}, CSV shape {detail_sheet.shape}",
    )

    return checks


def main():
    checks = check_inputs()

    mapping_sheet = pd.read_csv(MAPPING_SHEET_INPUT, dtype=str).fillna("")
    detail_sheet = pd.read_csv(DETAIL_SHEET_INPUT, dtype=str).fillna("")

    export_excel(mapping_sheet, detail_sheet)
    checks = validate_export(checks, mapping_sheet, detail_sheet)
    save_checks(checks)

    failed = [check for check in checks if check["status"] == "failed"]
    if failed:
        print("final excel export failed")
        print("failed checks: " + ", ".join(check["check_item"] for check in failed))
        raise SystemExit(1)

    print("final excel exported successfully")
    print(f"output file: {OUTPUT_EXCEL}")
    print("sheets: 映射表, 详情表")


if __name__ == "__main__":
    main()
