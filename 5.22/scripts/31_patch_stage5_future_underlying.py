from __future__ import annotations

import itertools
import re
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Font


ROOT = Path(__file__).resolve().parents[1]
ENCODING = "utf-8-sig"

ASSET_MASTER_IN = ROOT / "data" / "stage4_etf_index" / "final" / "asset_master_with_etf_index_cleaned.csv"
UNDERLYING_IN = ROOT / "data" / "stage5_mapping" / "final" / "underlying_mapping.csv"
DOMESTIC_IN = ROOT / "data" / "stage5_mapping" / "final" / "domestic_cross_market_mapping.csv"
OPTIONS_WITH_ETF = ROOT / "data" / "stage4_etf_index" / "final" / "options_master_with_etf_underlying.csv"

MANUAL_DIR = ROOT / "data" / "stage5_mapping" / "manual"
FINAL_DIR = ROOT / "data" / "stage5_mapping" / "final"
PROCESSED_DIR = ROOT / "data" / "stage5_mapping" / "processed"
OUTPUT_DIR = ROOT / "output"

PATCH_PATH = MANUAL_DIR / "manual_future_underlying_group_patch.csv"
ASSET_PATCHED_OUT = FINAL_DIR / "asset_master_for_mapping_patched.csv"
UNDERLYING_FIXED_OUT = FINAL_DIR / "underlying_mapping_fixed.csv"
DOMESTIC_FIXED_OUT = FINAL_DIR / "domestic_cross_market_mapping_fixed.csv"
PATCH_LOG_OUT = PROCESSED_DIR / "future_underlying_group_patch_log.csv"
REMAINING_REVIEW_OUT = PROCESSED_DIR / "future_underlying_group_remaining_review.csv"
UNDERLYING_QUALITY_FIXED_OUT = PROCESSED_DIR / "underlying_mapping_quality_report_fixed.csv"
DOMESTIC_QUALITY_FIXED_OUT = PROCESSED_DIR / "domestic_mapping_quality_report_fixed.csv"
MD_FIXED_OUT = OUTPUT_DIR / "stage5_mapping_quality_review_fixed.md"
EXCEL_FIXED_OUT = OUTPUT_DIR / "mapping_info_stage5_mapping_fixed.xlsx"

MISSING_VALUES = {"", "CHECK", "TODO", "NA", "NAN", "NONE", "CHECK_WAIT_ETF_MASTER"}
PATCH_NOTE = "future underlying_group patched for stage5 mapping quality, need review"


GROUP_MAP = {
    "IF": "CSI300", "IH": "SSE50", "IC": "CSI500", "IM": "CSI1000",
    "T": "CGB_10Y", "TF": "CGB_5Y", "TS": "CGB_2Y", "TL": "CGB_30Y",
    "AU": "GOLD", "AG": "SILVER", "CU": "COPPER", "BC": "COPPER", "AL": "ALUMINUM",
    "ZN": "ZINC", "PB": "LEAD", "NI": "NICKEL", "SN": "TIN", "AO": "ALUMINA",
    "AD": "CAST_ALUMINUM_ALLOY", "SS": "STAINLESS_STEEL", "WR": "WIRE_ROD",
    "RB": "REBAR", "HC": "HOT_ROLLED_COIL", "I": "IRON_ORE", "J": "COKE",
    "JM": "COKING_COAL", "SF": "FERROSILICON", "SM": "SILICOMANGANESE",
    "ZC": "THERMAL_COAL", "TC": "THERMAL_COAL",
    "SC": "CRUDE_OIL", "SCTAS": "CRUDE_OIL", "FU": "FUEL_OIL", "LU": "LOW_SULFUR_FUEL_OIL",
    "BU": "BITUMEN", "RU": "RUBBER", "NR": "RUBBER", "BR": "BUTADIENE_RUBBER",
    "SP": "PULP", "OP": "OFFSET_PAPER", "TA": "PTA", "PX": "PARAXYLENE",
    "MA": "METHANOL", "ME": "METHANOL", "L": "LLDPE", "V": "PVC", "PP": "POLYPROPYLENE",
    "EG": "ETHYLENE_GLYCOL", "EB": "STYRENE", "PF": "POLYESTER_STAPLE_FIBER",
    "PR": "BOTTLE_GRADE_PET", "PL": "PROPYLENE", "BZ": "BENZENE", "UR": "UREA",
    "SA": "SODA_ASH", "FG": "GLASS", "SH": "CAUSTIC_SODA", "PG": "LPG",
    "EC": "CONTAINER_SHIPPING_INDEX_EUROPE",
    "A": "SOYBEAN", "B": "SOYBEAN", "M": "SOYBEAN_MEAL", "Y": "SOYBEAN_OIL",
    "P": "PALM_OIL", "C": "CORN", "CS": "CORN_STARCH", "JD": "EGG", "LH": "LIVE_HOG",
    "RR": "JAPONICA_RICE", "AP": "APPLE", "CJ": "JUJUBE", "PK": "PEANUT", "SR": "SUGAR",
    "CF": "COTTON", "CY": "COTTON_YARN", "OI": "RAPESEED_OIL", "RO": "RAPESEED_OIL",
    "RM": "RAPESEED_MEAL", "RS": "RAPESEED", "ER": "EARLY_INDICA_RICE",
    "RI": "EARLY_INDICA_RICE", "JR": "JAPONICA_RICE", "LR": "LATE_INDICA_RICE",
    "PM": "COMMON_WHEAT", "WH": "STRONG_WHEAT", "WS": "STRONG_WHEAT", "WT": "HARD_WHEAT",
    "BB": "PLYWOOD", "FB": "FIBERBOARD", "LG": "LOG",
    "SI": "INDUSTRIAL_SILICON", "LC": "LITHIUM_CARBONATE", "PS": "POLYSILICON",
    "PD": "PALLADIUM", "PT": "PLATINUM",
}

GROUP_NAME = {
    "CSI300": ("沪深300", "CSI 300"), "SSE50": ("上证50", "SSE 50"),
    "CSI500": ("中证500", "CSI 500"), "CSI1000": ("中证1000", "CSI 1000"),
    "CGB_10Y": ("10年期国债", "10-year China Government Bond"),
    "CGB_5Y": ("5年期国债", "5-year China Government Bond"),
    "CGB_2Y": ("2年期国债", "2-year China Government Bond"),
    "CGB_30Y": ("30年期国债", "30-year China Government Bond"),
    "GOLD": ("黄金", "Gold"), "SILVER": ("白银", "Silver"), "COPPER": ("铜", "Copper"),
    "ALUMINUM": ("铝", "Aluminum"), "ZINC": ("锌", "Zinc"), "LEAD": ("铅", "Lead"),
    "NICKEL": ("镍", "Nickel"), "TIN": ("锡", "Tin"), "ALUMINA": ("氧化铝", "Alumina"),
    "CAST_ALUMINUM_ALLOY": ("铸造铝合金", "Cast Aluminum Alloy"),
    "STAINLESS_STEEL": ("不锈钢", "Stainless Steel"), "WIRE_ROD": ("线材", "Wire Rod"),
    "REBAR": ("螺纹钢", "Rebar"), "HOT_ROLLED_COIL": ("热轧卷板", "Hot Rolled Coil"),
    "IRON_ORE": ("铁矿石", "Iron Ore"), "COKE": ("焦炭", "Coke"),
    "COKING_COAL": ("焦煤", "Coking Coal"), "THERMAL_COAL": ("动力煤", "Thermal Coal"),
    "CRUDE_OIL": ("原油", "Crude Oil"), "FUEL_OIL": ("燃料油", "Fuel Oil"),
    "LOW_SULFUR_FUEL_OIL": ("低硫燃料油", "Low Sulfur Fuel Oil"), "BITUMEN": ("沥青", "Bitumen"),
    "RUBBER": ("橡胶", "Rubber"), "BUTADIENE_RUBBER": ("丁二烯橡胶", "Butadiene Rubber"),
    "PULP": ("纸浆", "Pulp"), "OFFSET_PAPER": ("胶版印刷纸", "Offset Paper"),
    "PTA": ("PTA", "PTA"), "PARAXYLENE": ("对二甲苯", "Paraxylene"),
    "METHANOL": ("甲醇", "Methanol"), "LLDPE": ("线性低密度聚乙烯", "LLDPE"),
    "PVC": ("PVC", "PVC"), "POLYPROPYLENE": ("聚丙烯", "Polypropylene"),
    "ETHYLENE_GLYCOL": ("乙二醇", "Ethylene Glycol"), "STYRENE": ("苯乙烯", "Styrene"),
    "POLYESTER_STAPLE_FIBER": ("短纤", "Polyester Staple Fiber"),
    "BOTTLE_GRADE_PET": ("瓶片", "Bottle-grade PET"), "PROPYLENE": ("丙烯", "Propylene"),
    "BENZENE": ("纯苯", "Benzene"), "UREA": ("尿素", "Urea"), "SODA_ASH": ("纯碱", "Soda Ash"),
    "GLASS": ("玻璃", "Glass"), "FERROSILICON": ("硅铁", "Ferrosilicon"),
    "SILICOMANGANESE": ("锰硅", "Silicomanganese"), "CAUSTIC_SODA": ("烧碱", "Caustic Soda"),
    "LPG": ("LPG", "LPG"), "CONTAINER_SHIPPING_INDEX_EUROPE": ("集运指数(欧线)", "Container Shipping Index Europe"),
    "SOYBEAN": ("大豆", "Soybean"), "SOYBEAN_MEAL": ("豆粕", "Soybean Meal"),
    "SOYBEAN_OIL": ("豆油", "Soybean Oil"), "PALM_OIL": ("棕榈油", "Palm Oil"),
    "CORN": ("玉米", "Corn"), "CORN_STARCH": ("玉米淀粉", "Corn Starch"),
    "EGG": ("鸡蛋", "Egg"), "LIVE_HOG": ("生猪", "Live Hog"),
    "JAPONICA_RICE": ("粳米", "Japonica Rice"), "EARLY_INDICA_RICE": ("早籼稻", "Early Indica Rice"),
    "LATE_INDICA_RICE": ("晚籼稻", "Late Indica Rice"), "COMMON_WHEAT": ("普麦", "Common Wheat"),
    "STRONG_WHEAT": ("强麦", "Strong Wheat"), "HARD_WHEAT": ("硬麦", "Hard Wheat"),
    "SUGAR": ("白糖", "Sugar"), "COTTON": ("棉花", "Cotton"), "COTTON_YARN": ("棉纱", "Cotton Yarn"),
    "RAPESEED_OIL": ("菜籽油", "Rapeseed Oil"), "RAPESEED_MEAL": ("菜粕", "Rapeseed Meal"),
    "RAPESEED": ("油菜籽", "Rapeseed"), "PEANUT": ("花生", "Peanut"), "APPLE": ("苹果", "Apple"),
    "JUJUBE": ("红枣", "Jujube"), "PLYWOOD": ("胶合板", "Plywood"), "FIBERBOARD": ("纤维板", "Fiberboard"),
    "LOG": ("原木", "Log"), "INDUSTRIAL_SILICON": ("工业硅", "Industrial Silicon"),
    "LITHIUM_CARBONATE": ("碳酸锂", "Lithium Carbonate"), "POLYSILICON": ("多晶硅", "Polysilicon"),
    "PALLADIUM": ("钯", "Palladium"), "PLATINUM": ("铂", "Platinum"),
}

PATCH_COLUMNS = [
    "asset_id", "symbol", "name_cn", "exchange_code", "old_underlying_group", "new_underlying_group",
    "underlying_name_cn", "underlying_name_en", "patch_method", "mapping_confidence", "review_status", "notes",
]


def ensure_dirs() -> None:
    for path in [MANUAL_DIR, FINAL_DIR, PROCESSED_DIR, OUTPUT_DIR]:
        path.mkdir(parents=True, exist_ok=True)


def clean_text(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def is_missing_value(value: Any) -> bool:
    return clean_text(value).upper() in MISSING_VALUES


def missing_mask(series: pd.Series) -> pd.Series:
    return series.fillna("").astype(str).str.strip().str.upper().isin(MISSING_VALUES)


def read_csv(path: Path, required: bool = True) -> pd.DataFrame:
    if not path.exists():
        if required:
            raise FileNotFoundError(f"缺少输入文件: {path}")
        return pd.DataFrame()
    return pd.read_csv(path, dtype=str, encoding=ENCODING).fillna("")


def save_csv(df: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(path, index=False, encoding=ENCODING)


def normalize_symbol(symbol: Any) -> str:
    text = clean_text(symbol).upper()
    text = re.sub(r"\s+", "", text)
    if text.endswith("_F"):
        text = text[:-2]
    return text


def infer_future_group(symbol: Any) -> tuple[str, str, str]:
    normalized = normalize_symbol(symbol)
    if normalized in GROUP_MAP:
        return GROUP_MAP[normalized], "CONSERVATIVE_SYMBOL_RULE_STAGE5_1", "HIGH"
    return "", "NEED_MANUAL_REVIEW", "CHECK"


def group_names(group: str) -> tuple[str, str]:
    return GROUP_NAME.get(group, ("CHECK", "CHECK"))


def identify_future_missing(underlying: pd.DataFrame, asset_master: pd.DataFrame) -> pd.DataFrame:
    mask = (
        underlying["asset_type"].astype(str).str.upper().eq("FUTURE")
        & missing_mask(underlying["underlying_group"])
    )
    missing = underlying.loc[mask, ["asset_id", "symbol", "name_cn", "exchange_code", "underlying_group"]].copy()
    missing = missing.rename(columns={"underlying_group": "old_underlying_group"})
    extra_cols = ["asset_id", "subtype", "sector", "notes"]
    missing = missing.merge(asset_master[extra_cols], on="asset_id", how="left")
    return missing


def build_manual_patch(missing: pd.DataFrame) -> pd.DataFrame:
    if PATCH_PATH.exists():
        patch = read_csv(PATCH_PATH)
        for col in PATCH_COLUMNS:
            if col not in patch.columns:
                patch[col] = ""
        patch = patch[PATCH_COLUMNS].copy()
    else:
        patch = pd.DataFrame(columns=PATCH_COLUMNS)

    existing_ids = set(patch["asset_id"].astype(str))
    new_rows = []
    for _, row in missing.iterrows():
        if row["asset_id"] in existing_ids:
            continue
        new_group, method, confidence = infer_future_group(row["symbol"])
        cn, en = group_names(new_group)
        new_rows.append(
            {
                "asset_id": row["asset_id"],
                "symbol": row["symbol"],
                "name_cn": row["name_cn"],
                "exchange_code": row["exchange_code"],
                "old_underlying_group": row["old_underlying_group"],
                "new_underlying_group": new_group,
                "underlying_name_cn": cn,
                "underlying_name_en": en,
                "patch_method": method,
                "mapping_confidence": confidence,
                "review_status": "AUTO_PATCH_NEED_REVIEW" if new_group else "NEED_MANUAL_REVIEW",
                "notes": f"Generated by Step 5.1; subtype={row.get('subtype', '')}; sector={row.get('sector', '')}",
            }
        )

    if new_rows:
        patch = pd.concat([patch, pd.DataFrame(new_rows)], ignore_index=True)

    # Refresh metadata for blank machine-generated fields only; preserve user-filled new_underlying_group.
    for idx, row in patch.iterrows():
        group = clean_text(row["new_underlying_group"])
        if group:
            cn, en = group_names(group)
            if is_missing_value(row["underlying_name_cn"]):
                patch.at[idx, "underlying_name_cn"] = cn
            if is_missing_value(row["underlying_name_en"]):
                patch.at[idx, "underlying_name_en"] = en
            if is_missing_value(row["mapping_confidence"]):
                patch.at[idx, "mapping_confidence"] = "HIGH"
            if is_missing_value(row["patch_method"]):
                patch.at[idx, "patch_method"] = "MANUAL_PATCH_EXISTING_VALUE"
            if is_missing_value(row["review_status"]):
                patch.at[idx, "review_status"] = "AUTO_PATCH_NEED_REVIEW"

    save_csv(patch, PATCH_PATH)
    return patch


def append_note(old_note: Any, note: str) -> str:
    text = clean_text(old_note)
    if note in text:
        return text
    return f"{text} | {note}" if text else note


def apply_patch_to_asset_master(asset_master: pd.DataFrame, patch: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    patched = asset_master.copy()
    valid_patch = patch[~missing_mask(patch["new_underlying_group"])].copy()
    patch_by_id = valid_patch.drop_duplicates("asset_id", keep="last").set_index("asset_id")
    log_rows = []

    for idx, row in patched.iterrows():
        asset_id = clean_text(row["asset_id"])
        should_patch = (
            clean_text(row["asset_type"]).upper() == "FUTURE"
            and is_missing_value(row.get("underlying_group", ""))
            and asset_id in patch_by_id.index
        )
        old_group = row.get("underlying_group", "")
        old_source_status = row.get("source_status", "")
        if should_patch:
            p = patch_by_id.loc[asset_id]
            new_group = clean_text(p["new_underlying_group"])
            patched.at[idx, "underlying_group"] = new_group
            patched.at[idx, "source_status"] = "FUTURE_UNDERLYING_GROUP_PATCHED_NEED_REVIEW"
            patched.at[idx, "notes"] = append_note(row.get("notes", ""), PATCH_NOTE)
            patch_applied = "Y"
            new_source_status = "FUTURE_UNDERLYING_GROUP_PATCHED_NEED_REVIEW"
            method = p.get("patch_method", "")
            confidence = p.get("mapping_confidence", "")
        else:
            new_group = old_group
            patch_applied = "N"
            new_source_status = old_source_status
            method = ""
            confidence = ""

        if clean_text(row["asset_type"]).upper() == "FUTURE" and (is_missing_value(old_group) or patch_applied == "Y"):
            log_rows.append(
                {
                    "asset_id": asset_id,
                    "symbol": row.get("symbol", ""),
                    "name_cn": row.get("name_cn", ""),
                    "exchange_code": row.get("exchange_code", ""),
                    "old_underlying_group": old_group,
                    "new_underlying_group": new_group,
                    "patch_applied": patch_applied,
                    "patch_method": method,
                    "mapping_confidence": confidence,
                    "old_source_status": old_source_status,
                    "new_source_status": new_source_status,
                    "notes": append_note(row.get("notes", ""), PATCH_NOTE) if patch_applied == "Y" else row.get("notes", ""),
                }
            )

    patch_log = pd.DataFrame(log_rows)
    save_csv(patched, ASSET_PATCHED_OUT)
    save_csv(patch_log, PATCH_LOG_OUT)
    return patched, patch_log


def infer_group_for_underlying(row: pd.Series, patched_ids: set[str]) -> tuple[str, str, str, str, str]:
    group = clean_text(row.get("underlying_group", ""))
    if not is_missing_value(group):
        cn, en = group_names(group)
        if row.get("asset_id", "") in patched_ids:
            return group, "HIGH", "FUTURE_UNDERLYING_GROUP_PATCH", cn, en
        return group, "HIGH", "EXISTING_UNDERLYING_GROUP", cn, en

    for candidate in [row.get("symbol", ""), clean_text(row.get("asset_id", "")).split("_")[-1]]:
        normalized = normalize_symbol(candidate)
        if normalized in GROUP_MAP:
            group = GROUP_MAP[normalized]
            cn, en = group_names(group)
            method = "FUTURE_UNDERLYING_GROUP_PATCH" if row.get("asset_id", "") in patched_ids else "CONSERVATIVE_SYMBOL_RULE"
            return group, "HIGH", method, cn, en

    name = clean_text(row.get("name_cn", ""))
    name_rules = [
        ("沪深300", "CSI300"), ("上证50", "SSE50"), ("中证500", "CSI500"), ("中证1000", "CSI1000"),
    ]
    for keyword, mapped_group in name_rules:
        if keyword in name:
            cn, en = group_names(mapped_group)
            return mapped_group, "HIGH", "CONSERVATIVE_NAME_RULE", cn, en

    return "CHECK", "CHECK", "NEED_MANUAL_REVIEW", "CHECK", "CHECK"


def build_underlying_mapping_fixed(asset_master: pd.DataFrame, patch_log: pd.DataFrame) -> pd.DataFrame:
    patched_ids = set(patch_log.loc[patch_log["patch_applied"] == "Y", "asset_id"]) if not patch_log.empty else set()
    rows = []
    for _, row in asset_master.iterrows():
        group, confidence, method, cn, en = infer_group_for_underlying(row, patched_ids)
        rows.append(
            {
                "asset_id": row.get("asset_id", ""),
                "symbol": row.get("symbol", ""),
                "name_cn": row.get("name_cn", ""),
                "asset_type": row.get("asset_type", ""),
                "subtype": row.get("subtype", ""),
                "exchange_code": row.get("exchange_code", ""),
                "underlying_group": group,
                "underlying_name_cn": cn,
                "underlying_name_en": en,
                "mapping_confidence": confidence,
                "mapping_method": method,
                "data_source_1": row.get("data_source_1", ""),
                "data_source_2": row.get("data_source_2", ""),
                "source_status": "UNDERLYING_MAPPING_FIXED_NEED_REVIEW" if confidence != "CHECK" else "UNDERLYING_MAPPING_NEED_REVIEW",
                "notes": "Step 5.1 fixed mapping generated from patched asset master.",
            }
        )
    fixed = pd.DataFrame(rows)
    save_csv(fixed, UNDERLYING_FIXED_OUT)
    return fixed


def norm_asset(row: pd.Series) -> dict[str, str]:
    keys = ["asset_id", "symbol", "name_cn", "asset_type", "subtype", "exchange_code", "underlying_group", "tradable"]
    return {k: clean_text(row.get(k, "")) for k in keys}


def relation_for(a: dict[str, str], b: dict[str, str], options_by_id: dict[str, pd.Series]) -> tuple[str, str, str] | None:
    types = {a["asset_type"], b["asset_type"]}
    subtypes = {a["subtype"], b["subtype"]}
    if types == {"FUTURE", "ETF"}:
        return "FUTURE_ETF", "BASIS_ARBITRAGE", "HIGH" if a["underlying_group"] == b["underlying_group"] else "MEDIUM"
    if types == {"FUTURE", "INDEX"}:
        return "FUTURE_INDEX", "BASIS_ARBITRAGE", "HIGH"
    if types == {"ETF", "INDEX"}:
        return "ETF_INDEX", "ETF_PREMIUM_DISCOUNT_ARBITRAGE", "HIGH"
    if "INDEX_OPTION" in subtypes and "INDEX" in types:
        return "OPTION_INDEX", "OPTION_ARBITRAGE", "HIGH"
    if "INDEX_OPTION" in subtypes and "FUTURE" in types:
        return "OPTION_FUTURE", "OPTION_FUTURE_ARBITRAGE", "HIGH"
    if "ETF_OPTION" in subtypes and "ETF" in types:
        option = a if a["subtype"] == "ETF_OPTION" else b
        other = b if a["subtype"] == "ETF_OPTION" else a
        opt_row = options_by_id.get(option["asset_id"])
        if opt_row is not None and clean_text(opt_row.get("underlying_asset_id")) == other["asset_id"]:
            return "OPTION_UNDERLYING", "OPTION_ARBITRAGE", "HIGH"
        return "OPTION_UNDERLYING", "OPTION_ARBITRAGE", "MEDIUM"
    if "COMMODITY_OPTION" in subtypes and "FUTURE" in types:
        option = a if a["subtype"] == "COMMODITY_OPTION" else b
        other = b if a["subtype"] == "COMMODITY_OPTION" else a
        opt_row = options_by_id.get(option["asset_id"])
        if opt_row is not None and clean_text(opt_row.get("underlying_asset_id")) == other["asset_id"]:
            return "OPTION_UNDERLYING", "OPTION_FUTURE_ARBITRAGE", "HIGH"
        return "OPTION_UNDERLYING", "OPTION_FUTURE_ARBITRAGE", "MEDIUM"
    if a["asset_type"] == "FUTURE" and b["asset_type"] == "FUTURE" and a["exchange_code"] != b["exchange_code"]:
        return "SAME_UNDERLYING_CROSS_EXCHANGE", "CROSS_EXCHANGE_ARBITRAGE", "MEDIUM"
    return None


def direction_for(relation: str, a_type: str, b_type: str) -> tuple[str, str]:
    if a_type == "FUTURE" and b_type == "FUTURE":
        return "BIDIRECTIONAL", "两个期货资产均可双向交易，跨交易所方向需结合合约规则确认。"
    if relation == "FUTURE_ETF":
        return "LONG_ETF_AND_SHORT_FUTURE_NEED_CHECK", "常见方向为多 ETF 空期货或反向，需结合融券、保证金和基差状态确认。"
    if relation in {"ETF_INDEX", "FUTURE_INDEX", "OPTION_INDEX"}:
        return "INDEX_REFERENCE_ONLY", "指数为参考标的，不直接交易。"
    if relation in {"OPTION_FUTURE", "OPTION_UNDERLYING"}:
        return "BIDIRECTIONAL_WITH_OPTION_MARGIN_CONSTRAINT", "期权买卖方向受保证金、权限和行权规则约束。"
    return "CHECK", "方向需后续按策略规则确认。"


def build_domestic_mapping_fixed(asset_master: pd.DataFrame, underlying_mapping: pd.DataFrame) -> pd.DataFrame:
    assets = asset_master.merge(
        underlying_mapping[["asset_id", "underlying_group"]],
        on="asset_id",
        suffixes=("", "_mapped"),
        how="left",
    )
    assets["underlying_group"] = assets["underlying_group_mapped"].where(
        ~missing_mask(assets["underlying_group_mapped"]), assets["underlying_group"]
    )
    assets = assets[~missing_mask(assets["underlying_group"])].copy()
    options_with_etf = read_csv(OPTIONS_WITH_ETF, required=False)
    options_by_id = {row["asset_id"]: row for _, row in options_with_etf.iterrows()} if not options_with_etf.empty and "asset_id" in options_with_etf else {}

    rows = []
    seen: set[tuple[str, str]] = set()
    for group, group_df in assets.groupby("underlying_group"):
        if is_missing_value(group):
            continue
        records = [norm_asset(row) for _, row in group_df.iterrows()]
        for a, b in itertools.combinations(records, 2):
            if a["asset_id"] == b["asset_id"]:
                continue
            pair = tuple(sorted([a["asset_id"], b["asset_id"]]))
            if pair in seen:
                continue
            relation = relation_for(a, b, options_by_id)
            if relation is None:
                continue
            seen.add(pair)
            relation_type, strategy_type, confidence = relation
            direction, note = direction_for(relation_type, a["asset_type"], b["asset_type"])
            tradable = "BOTH_TRADABLE" if a["tradable"] == "Y" and b["tradable"] == "Y" else "CHECK"
            rows.append(
                {
                    "mapping_id": f"{relation_type}_{pair[0]}_{pair[1]}",
                    "asset_id_a": a["asset_id"], "asset_id_b": b["asset_id"],
                    "symbol_a": a["symbol"], "symbol_b": b["symbol"],
                    "name_a": a["name_cn"], "name_b": b["name_cn"],
                    "asset_type_a": a["asset_type"], "asset_type_b": b["asset_type"],
                    "exchange_a": a["exchange_code"], "exchange_b": b["exchange_code"],
                    "underlying_group": group, "relation_type": relation_type,
                    "strategy_type": strategy_type, "market_relation": "CN_DOMESTIC",
                    "tradable_check": tradable, "direction_supported": direction,
                    "long_short_note": note, "mapping_confidence": confidence,
                    "data_source_1": "asset_master_for_mapping_patched",
                    "data_source_2": "underlying_mapping_fixed",
                    "source_status": "DOMESTIC_MAPPING_FIXED_CANDIDATE_NEED_REVIEW",
                    "notes": "根据 Step 5.1 patched asset master 和 fixed underlying_mapping 自动生成国内跨市场映射候选，非套利计算结果。",
                }
            )
    mapping = pd.DataFrame(rows)
    if not mapping.empty:
        mapping = mapping.drop_duplicates("mapping_id").sort_values("mapping_id").reset_index(drop=True)
    save_csv(mapping, DOMESTIC_FIXED_OUT)
    return mapping


def build_remaining_review(underlying_mapping: pd.DataFrame) -> pd.DataFrame:
    mask = underlying_mapping["asset_type"].astype(str).str.upper().eq("FUTURE") & missing_mask(underlying_mapping["underlying_group"])
    cols = ["asset_id", "symbol", "name_cn", "exchange_code", "underlying_group", "notes"]
    review = underlying_mapping.loc[mask, cols].copy()
    review["issue_reason"] = "FUTURE underlying_group remains blank/CHECK after Step 5.1 patch."
    review["suggested_next_step"] = "需要人工确认该期货品种对应的统一标的，或确认是否为历史/特殊/非标准品种。"
    review = review[["asset_id", "symbol", "name_cn", "exchange_code", "underlying_group", "issue_reason", "suggested_next_step", "notes"]]
    save_csv(review, REMAINING_REVIEW_OUT)
    return review


def build_underlying_quality(asset_master: pd.DataFrame, underlying_mapping: pd.DataFrame) -> pd.DataFrame:
    rows = [{
        "section": "overall", "asset_type": "ALL", "total_asset_count": len(asset_master),
        "underlying_mapping_count": len(underlying_mapping),
        "underlying_group_missing_count": int(missing_mask(underlying_mapping["underlying_group"]).sum()),
        "mapping_confidence_high_count": int((underlying_mapping["mapping_confidence"] == "HIGH").sum()),
        "mapping_confidence_medium_count": int((underlying_mapping["mapping_confidence"] == "MEDIUM").sum()),
        "mapping_confidence_check_count": int((underlying_mapping["mapping_confidence"] == "CHECK").sum()),
        "total_count": len(asset_master), "mapping_count": len(underlying_mapping),
        "high_confidence_count": int((underlying_mapping["mapping_confidence"] == "HIGH").sum()),
        "check_count": int((underlying_mapping["mapping_confidence"] == "CHECK").sum()),
    }]
    for asset_type, group in underlying_mapping.groupby("asset_type", dropna=False):
        rows.append({
            "section": "by_asset_type", "asset_type": asset_type, "total_asset_count": "",
            "underlying_mapping_count": "", "underlying_group_missing_count": int(missing_mask(group["underlying_group"]).sum()),
            "mapping_confidence_high_count": "", "mapping_confidence_medium_count": "",
            "mapping_confidence_check_count": "", "total_count": len(group), "mapping_count": len(group),
            "high_confidence_count": int((group["mapping_confidence"] == "HIGH").sum()),
            "check_count": int((group["mapping_confidence"] == "CHECK").sum()),
        })
    report = pd.DataFrame(rows)
    save_csv(report, UNDERLYING_QUALITY_FIXED_OUT)
    return report


def build_domestic_quality(mapping: pd.DataFrame) -> pd.DataFrame:
    rows = [{
        "section": "overall", "relation_type": "ALL", "count": len(mapping),
        "high_confidence_count": int((mapping["mapping_confidence"] == "HIGH").sum()) if not mapping.empty else 0,
        "medium_confidence_count": int((mapping["mapping_confidence"] == "MEDIUM").sum()) if not mapping.empty else 0,
        "domestic_mapping_count": len(mapping),
        "future_etf_mapping_count": int((mapping["relation_type"] == "FUTURE_ETF").sum()) if not mapping.empty else 0,
        "future_index_mapping_count": int((mapping["relation_type"] == "FUTURE_INDEX").sum()) if not mapping.empty else 0,
        "etf_index_mapping_count": int((mapping["relation_type"] == "ETF_INDEX").sum()) if not mapping.empty else 0,
        "option_index_mapping_count": int((mapping["relation_type"] == "OPTION_INDEX").sum()) if not mapping.empty else 0,
        "option_future_mapping_count": int((mapping["relation_type"] == "OPTION_FUTURE").sum()) if not mapping.empty else 0,
        "option_underlying_mapping_count": int((mapping["relation_type"] == "OPTION_UNDERLYING").sum()) if not mapping.empty else 0,
        "cross_exchange_mapping_count": int((mapping["relation_type"] == "SAME_UNDERLYING_CROSS_EXCHANGE").sum()) if not mapping.empty else 0,
    }]
    if not mapping.empty:
        for relation, group in mapping.groupby("relation_type"):
            rows.append({
                "section": "by_relation_type", "relation_type": relation, "count": len(group),
                "high_confidence_count": int((group["mapping_confidence"] == "HIGH").sum()),
                "medium_confidence_count": int((group["mapping_confidence"] == "MEDIUM").sum()),
            })
    report = pd.DataFrame(rows)
    save_csv(report, DOMESTIC_QUALITY_FIXED_OUT)
    return report


def pair_key(row: pd.Series) -> str:
    return "|".join(sorted([clean_text(row.get("asset_id_a", "")), clean_text(row.get("asset_id_b", ""))]))


def quality_stats(underlying: pd.DataFrame, domestic: pd.DataFrame, asset: pd.DataFrame) -> dict[str, Any]:
    asset_ids = set(asset["asset_id"]) if "asset_id" in asset else set()
    domestic_ids = set(domestic.get("asset_id_a", pd.Series(dtype=str))) | set(domestic.get("asset_id_b", pd.Series(dtype=str)))
    duplicate_mapping_id = int(domestic["mapping_id"].duplicated().sum()) if "mapping_id" in domestic else 0
    self_mapping = int((domestic["asset_id_a"] == domestic["asset_id_b"]).sum()) if {"asset_id_a", "asset_id_b"}.issubset(domestic.columns) else 0
    missing_domestic_ids = sorted([x for x in domestic_ids if x and x not in asset_ids])
    future_missing = int(missing_mask(underlying.loc[underlying["asset_type"] == "FUTURE", "underlying_group"]).sum())
    option_missing = int(missing_mask(underlying.loc[underlying["asset_type"] == "OPTION", "underlying_group"]).sum())
    index_missing = int(missing_mask(underlying.loc[underlying["asset_type"] == "INDEX", "underlying_group"]).sum())
    high_issues = []
    if future_missing or option_missing or index_missing:
        high_issues.append("FUTURE/OPTION/INDEX underlying_group missing")
    if duplicate_mapping_id:
        high_issues.append("duplicate mapping_id")
    if self_mapping:
        high_issues.append("self mapping")
    if missing_domestic_ids:
        high_issues.append("domestic asset_id cannot link asset_master")
    conclusion = "NEED_MAJOR_FIX" if high_issues else "NEED_MINOR_FIX"
    if not high_issues and int((underlying["mapping_confidence"] == "CHECK").sum()) == 0:
        conclusion = "PASS"
    return {
        "future_missing": future_missing, "option_missing": option_missing, "index_missing": index_missing,
        "duplicate_mapping_id": duplicate_mapping_id, "self_mapping": self_mapping,
        "missing_domestic_asset_id_count": len(missing_domestic_ids),
        "high_issues": high_issues, "conclusion": conclusion,
    }


def write_markdown(before_future_missing: int, patch_log: pd.DataFrame, remaining_review: pd.DataFrame, underlying: pd.DataFrame, domestic: pd.DataFrame, asset: pd.DataFrame) -> dict[str, Any]:
    stats = quality_stats(underlying, domestic, asset)
    relation_dist = domestic["relation_type"].value_counts().sort_index() if not domestic.empty else pd.Series(dtype=int)
    patch_applied_count = int((patch_log["patch_applied"] == "Y").sum()) if not patch_log.empty else 0
    lines = [
        "# Step 5.1 映射表质量修复报告",
        "",
        f"- 修复前 FUTURE underlying_group 缺失数量：{before_future_missing}",
        f"- 修复后 FUTURE underlying_group 缺失数量：{stats['future_missing']}",
        f"- patch_applied_count：{patch_applied_count}",
        f"- remaining_review_count：{len(remaining_review)}",
        f"- domestic mapping fixed 行数：{len(domestic)}",
        "",
        "## relation_type 分布",
    ]
    if relation_dist.empty:
        lines.append("_无数据_")
    else:
        lines.extend([f"- {k}: {v}" for k, v in relation_dist.items()])
    lines.extend([
        "",
        "## HIGH 问题",
        "- 未发现 HIGH 问题。" if not stats["high_issues"] else "\n".join([f"- {x}" for x in stats["high_issues"]]),
        "",
        "## 判断",
        f"- FUTURE underlying_group 缺失：{stats['future_missing']}",
        f"- OPTION underlying_group 缺失：{stats['option_missing']}",
        f"- INDEX underlying_group 缺失：{stats['index_missing']}",
        f"- domestic mapping mapping_id 重复：{stats['duplicate_mapping_id']}",
        f"- domestic mapping 自映射：{stats['self_mapping']}",
        f"- domestic mapping asset_id 无法回连 asset_master：{stats['missing_domestic_asset_id_count']}",
        "- ETF 的大量非核心 underlying_group 缺失不视为当前阶段严重问题。",
        "",
        "## 结论",
        f"**{stats['conclusion']}**",
    ])
    MD_FIXED_OUT.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return stats


def readme_df() -> pd.DataFrame:
    return pd.DataFrame(
        [
            ("说明", "当前为 Step 5.1 修复 FUTURE underlying_group 后的映射表。"),
            ("修复范围", "修复对象仅限 Step 5 质量检查中识别出的 FUTURE underlying_group 缺失项。"),
            ("输入保护", "未修改 Stage 1 / Stage 2 / Step 4 原始文件。"),
            ("patched asset master", "patched asset master 仅用于 Step 5 映射质量修复。"),
            ("ETF 缺失说明", "ETF 的大量非核心 underlying_group 缺失不视为当前阶段严重问题。"),
            ("后续阶段", "后续可进入产业链映射或国外跨市场映射。"),
        ],
        columns=["item", "description"],
    )


def write_excel(sheets: dict[str, pd.DataFrame]) -> None:
    with pd.ExcelWriter(EXCEL_FIXED_OUT, engine="openpyxl") as writer:
        for name, df in sheets.items():
            sheet = name[:31]
            df.to_excel(writer, sheet_name=sheet, index=False)
            ws = writer.book[sheet]
            ws.freeze_panes = "A2"
            if ws.max_row > 1 and ws.max_column > 1:
                ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.font = Font(bold=True)
            for col_cells in ws.columns:
                width = 10
                letter = col_cells[0].column_letter
                for cell in col_cells[:200]:
                    width = max(width, min(len("" if cell.value is None else str(cell.value)) + 2, 40))
                ws.column_dimensions[letter].width = width


def main() -> None:
    ensure_dirs()
    asset_master = read_csv(ASSET_MASTER_IN)
    underlying_in = read_csv(UNDERLYING_IN)
    read_csv(DOMESTIC_IN)
    read_csv(PROCESSED_DIR / "underlying_mapping_quality_report.csv", required=False)
    read_csv(PROCESSED_DIR / "domestic_mapping_quality_report.csv", required=False)

    future_missing = identify_future_missing(underlying_in, asset_master)
    before_future_missing = len(future_missing)
    patch = build_manual_patch(future_missing)
    asset_patched, patch_log = apply_patch_to_asset_master(asset_master, patch)
    underlying_fixed = build_underlying_mapping_fixed(asset_patched, patch_log)
    domestic_fixed = build_domestic_mapping_fixed(asset_patched, underlying_fixed)
    remaining_review = build_remaining_review(underlying_fixed)
    underlying_quality = build_underlying_quality(asset_patched, underlying_fixed)
    domestic_quality = build_domestic_quality(domestic_fixed)
    stats = write_markdown(before_future_missing, patch_log, remaining_review, underlying_fixed, domestic_fixed, asset_patched)

    write_excel(
        {
            "README": readme_df(),
            "asset_master_for_mapping_patched": asset_patched,
            "underlying_mapping_fixed": underlying_fixed,
            "domestic_cross_market_mapping_fixed": domestic_fixed,
            "future_underlying_group_patch_log": patch_log,
            "future_underlying_group_remaining_review": remaining_review,
            "underlying_mapping_quality_report_fixed": underlying_quality,
            "domestic_mapping_quality_report_fixed": domestic_quality,
        }
    )

    relation_counts = domestic_fixed["relation_type"].value_counts().sort_index() if not domestic_fixed.empty else pd.Series(dtype=int)
    print(f"FUTURE underlying_group 缺失修复前数量: {before_future_missing}")
    print(f"FUTURE underlying_group 缺失修复后数量: {stats['future_missing']}")
    print(f"patch_applied_count: {int((patch_log['patch_applied'] == 'Y').sum()) if not patch_log.empty else 0}")
    print(f"remaining_review_count: {len(remaining_review)}")
    print(f"underlying_mapping_fixed 行数: {len(underlying_fixed)}")
    print(f"domestic_cross_market_mapping_fixed 行数: {len(domestic_fixed)}")
    print("各 relation_type 数量:")
    for relation, count in relation_counts.items():
        print(f"  {relation}: {count}")
    print(f"fixed 质量结论: {stats['conclusion']}")
    print(f"Excel 输出路径: {EXCEL_FIXED_OUT}")
    print(f"markdown 输出路径: {MD_FIXED_OUT}")


if __name__ == "__main__":
    main()
